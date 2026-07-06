# -*- coding: utf-8 -*-
"""Build a filtered first-pass checklist for final result-table review.

This intentionally does not write to the institution result-table workbooks.
It reads the structured candidate sheet from ``00_여기부터_검토시작.xlsx``
and creates a smaller checklist where first-pass rows are candidates that can
plausibly become final result-table entries.
"""

from __future__ import annotations

import argparse
import datetime as dt
import math
import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def reviewer_defaults() -> list[str]:
    raw = os.getenv("DISCLOSURE_REVIEWERS", "A,B,C,D")
    reviewers = [part.strip() for part in raw.split(",") if part.strip()]
    return reviewers or ["A", "B", "C", "D"]


REVIEWERS = reviewer_defaults()


def include_reviewers(values: Iterable[object]) -> None:
    for value in values:
        reviewer = str(value or "").strip()
        if reviewer and reviewer not in REVIEWERS:
            REVIEWERS.append(reviewer)


STATUS_VALUES = "확인전,이상없음,수정요청,판단보류,제외"
JUDGMENT_VALUES = "미입력,결과표반영,자료보완,설명요청,제외,이상없음"
RECHECK_VALUES = "N,Y"


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def flat(value: Any, limit: int = 260) -> str:
    value = re.sub(r"\s+", " ", text(value).replace("\n", " ")).strip()
    return value if len(value) <= limit else value[: limit - 1] + "…"


def compact(value: Any) -> str:
    value = text(value)
    value = value.replace("(주)", "").replace("㈜", "")
    value = re.sub(r"^\(?수정\)?\s*", "", value)
    value = re.sub(r"[\s\(\)\[\]{}·ㆍ._\-/]", "", value)
    return value


def normalized_agency(value: Any) -> str:
    value = compact(value)
    value = re.sub(r"^\d{3}", "", value)
    for token in ("공기업시장형", "공기업준시장형", "준정부기관기금관리형", "준정부기관위탁집행형", "기타공공기관", "부설기관"):
        value = value.replace(token, "")
    return value


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            return float(value)
        return None
    raw = str(value).replace(",", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def equal_number(left: Any, right: Any, tolerance: float = 0.0001) -> bool:
    lnum = number(left)
    rnum = number(right)
    if lnum is None or rnum is None:
        return False
    return abs(lnum - rnum) <= tolerance


def find_dir(parent: Path, prefix: str) -> Path:
    matches = [p for p in parent.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not matches:
        raise FileNotFoundError(f"Missing directory {prefix}* under {parent}")
    return sorted(matches, key=lambda p: p.name)[0]


def find_report(core_dir: Path) -> Path:
    matches = [
        p
        for p in core_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and p.name.startswith("00_")
    ]
    if not matches:
        raise FileNotFoundError(f"Missing 00_* report workbook under {core_dir}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def find_core_dir(out_dir: Path) -> Path:
    matches = [
        p
        for p in out_dir.iterdir()
        if p.is_dir() and p.name.startswith("00_") and "핵심산출물" in p.name
    ]
    if not matches:
        matches = [
            p
            for p in out_dir.iterdir()
            if p.is_dir() and p.name.startswith("00_") and "먼저보기" not in p.name
        ]
    if not matches:
        raise FileNotFoundError(f"Missing core output folder under {out_dir}")
    return sorted(matches, key=lambda p: p.name)[0]


def find_admin_queue(core_dir: Path) -> Path:
    matches = [
        p
        for p in core_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and p.name.startswith("01_") and "관리자" in p.name
        and "필터링" not in p.name
    ]
    if not matches:
        raise FileNotFoundError(f"Missing administrator queue under {core_dir}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def sheet_by_name_or_index(wb, title: str, index: int):
    if title in wb.sheetnames:
        return wb[title]
    return wb[wb.sheetnames[index]]


def read_sheet_rows(path: Path, sheet_title: str, index: int) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = sheet_by_name_or_index(wb, sheet_title, index)
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    result: list[dict[str, Any]] = []
    for row in rows:
        if not any(text(v) for v in row):
            continue
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    wb.close()
    return result


def build_assignment_map(admin_queue: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(admin_queue, read_only=True, data_only=True)
    sheet_name = "전체후보" if "전체후보" in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}
    mapping: dict[str, dict[str, str]] = {}
    for row in rows:
        reviewer = text(row[idx.get("담당자", -1)] if "담당자" in idx else "")
        agency = text(row[idx.get("기관명", -1)] if "기관명" in idx else "")
        check_no = text(row[idx.get("기관번호", -1)] if "기관번호" in idx else "")
        if not reviewer or not agency:
            continue
        key = normalized_agency(agency)
        if key and key not in mapping:
            mapping[key] = {"reviewer": reviewer, "check_no": check_no, "agency": agency}
    wb.close()
    return mapping


def lookup_assignment(agency: str, assignment: dict[str, dict[str, str]]) -> dict[str, str]:
    key = normalized_agency(agency)
    if key in assignment:
        return assignment[key]
    for known, value in assignment.items():
        if key and (key in known or known in key):
            return value
    return {"reviewer": "", "check_no": "", "agency": agency}


def resolve_path(root: Path, raw: Any) -> Path | None:
    first = text(raw).splitlines()[0] if text(raw) else ""
    first = re.sub(r"^ALIO 공시자료:\s*", "", first).strip().strip('"')
    if not first:
        return None
    candidate = Path(first)
    candidates = [candidate]
    if not candidate.is_absolute():
        candidates.extend([root / candidate, root.parent / candidate])
        candidates.extend(child / candidate for child in root.iterdir() if child.is_dir())
    for path in candidates:
        try:
            if path.exists():
                return path
        except OSError:
            continue
    return None


def norm_item(value: Any) -> str:
    return re.sub(r"\s+", "", text(value))


def find_row_by_item(ws, item_name: str, row_range: range) -> int | None:
    key = norm_item(item_name)
    for row in row_range:
        if norm_item(ws.cell(row, 3).value) == key:
            return row
    return None


def find_sheet_token(wb, token: str):
    return next((wb[name] for name in wb.sheetnames if token in name), None)


def classify_13_welfare_crosscheck(root: Path, row: dict[str, Any]) -> tuple[str | None, str]:
    """Return (exclude_reason, evidence_note) for 13-1/13-5 cross-check rows."""
    if text(row.get("항목")) != "13-1/13-5":
        return None, ""
    if "예산상 복리후생비와 1인당 복리후생비" not in text(row.get("발견 요약")):
        return None, ""
    key = text(row.get("상세키/미매칭"))
    parts = key.split("/")
    if len(parts) < 2:
        return None, ""
    emp_type, welfare_item = parts[0], parts[1]
    path = resolve_path(root, row.get("대조군 자료") or row.get("확인자료"))
    if not path:
        return None, "제출엑셀 경로 확인 실패"

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws13_1 = find_sheet_token(wb, "13-1")
        ws13_5 = find_sheet_token(wb, "13-5")
        if not ws13_1 or not ws13_5:
            wb.close()
            return None, "13-1/13-5 시트 확인 실패"

        emp_cols = {"임원": 4, "일반정규직": 5, "무기계약직": 6, "비정규직": 7}
        emp_blocks = {
            "임원": range(13, 26),
            "일반정규직": range(28, 41),
            "무기계약직": range(43, 56),
            "비정규직": range(58, 71),
        }
        col = emp_cols.get(emp_type)
        block = emp_blocks.get(emp_type)
        if not col or not block:
            wb.close()
            return None, ""
        pay_row = find_row_by_item(ws13_1, welfare_item, range(14, 27))
        nonpay_row = find_row_by_item(ws13_1, welfare_item, range(28, 41))
        total_row = find_row_by_item(ws13_5, welfare_item, block)
        if total_row is None:
            wb.close()
            return None, ""
        left = (number(ws13_1.cell(pay_row, col).value) if pay_row else 0) or 0
        right = (number(ws13_1.cell(nonpay_row, col).value) if nonpay_row else 0) or 0
        total = (number(ws13_5.cell(total_row, 5).value) or 0)
        note = (
            f"13-1 급여성 {left:,.0f} + 비급여성 {right:,.0f} = {left + right:,.0f}; "
            f"13-5 예산총합 {total:,.0f}"
        )
        wb.close()
        if abs((left + right) - total) <= 0.0001:
            return "13-1 급여성+비급여성 합산값이 13-5 예산총합과 일치하여 자동 불일치 후보에서 제외", note
        return None, note
    except Exception as exc:
        return None, f"합산검산 실패: {exc}"


def row_blob(row: dict[str, Any]) -> str:
    keys = [
        "항목",
        "검토유형",
        "검토 질문",
        "발견 요약",
        "차이/판정",
        "확인자료",
        "상세키/미매칭",
        "자동점수/원천근거",
    ]
    return " ".join(text(row.get(key)) for key in keys)


def has_any(value: str, tokens: tuple[str, ...]) -> bool:
    return any(token in value for token in tokens)


def numeric_difference(row: dict[str, Any]) -> float | None:
    match = re.search(r"차이\s+(-?[\d,]+(?:\.\d+)?)", text(row.get("차이/판정")))
    if not match:
        return None
    try:
        return abs(float(match.group(1).replace(",", "")))
    except ValueError:
        return None


def classify_second_pass_filter(row: dict[str, Any], evidence_note: str) -> dict[str, str] | None:
    review_type = text(row.get("검토유형"))
    item = text(row.get("항목"))
    finding = text(row.get("발견 요약"))
    judgment = text(row.get("차이/판정"))
    blob = row_blob(row)

    if has_any(blob, ("지적 X", "지적X", "지적 제외", "지적제외", "개선권고")):
        return {
            "stage": "2차",
            "priority": "LOW",
            "possibility": "하",
            "filter_reason": "2024년 기준상 지적 X·지적 제외·개선권고 성격이 섞일 수 있어 1차에서 후순위 처리",
            "review_point": "기준 엑셀의 비고/정평 의견과 같은 사안인지 확인 후 필요 시 제외",
            "evidence_note": evidence_note,
        }

    if has_any(blob, ("수정공시", "정오표")):
        return {
            "stage": "2차",
            "priority": "MEDIUM",
            "possibility": "중",
            "filter_reason": "수정공시·정오표 이슈는 ALIO 이력/정오표 확인 전 자동 확정 불가",
            "review_point": "수정공시 일자, 정오표 반영 여부, 점검기간 중 수정 인정 여부를 별도로 확인",
            "evidence_note": evidence_note,
        }

    if review_type == "공시-제출 수치" and has_any(judgment, ("1,000배 단위차 의심", "1000배 단위차 의심")):
        return {
            "stage": "2차",
            "priority": "MEDIUM",
            "possibility": "중",
            "filter_reason": "1,000배 단위차만 있는 수치 후보는 원/천원 단위 보정 후 판단해야 하므로 2차",
            "review_point": "원/천원 단위 보정으로 일치하는지 먼저 확인하고, 보정 후에도 차이가 남으면 1차 후보로 승격",
            "evidence_note": evidence_note,
        }

    if review_type == "공시-제출 수치" and item != "5-1":
        diff = numeric_difference(row)
        if diff is not None and diff <= 1:
            return {
                "stage": "2차",
                "priority": "LOW",
                "possibility": "하",
                "filter_reason": "금액성 수치 차이가 1 이하라 반올림·입력단위 오차 가능성이 커 2차",
                "review_point": "입력단위, 소수점, 반올림 기준으로 설명 가능한지 확인하고 설명 불가 시 1차 후보로 승격",
                "evidence_note": evidence_note,
            }

    if review_type == "공시-제출 수치" and has_any(
        blob,
        ("반올림", "소수점", "단수차이", "입력단위", "1원 차이", "소액 오차", "허용오차"),
    ):
        return {
            "stage": "2차",
            "priority": "LOW",
            "possibility": "하",
            "filter_reason": "소액·반올림·입력단위 오차는 기준상 지적 제외 가능성이 있어 2차",
            "review_point": "허용오차, 입력단위, 반올림 기준으로 설명 가능한지 확인",
            "evidence_note": evidence_note,
        }

    if review_type == "제출-근거자료" and finding != "근거자료 없음":
        if "제목" in finding:
            reason = "근거자료는 있으나 자동 제목 매칭률이 낮아 증빙 위치 확인이 필요한 후보"
            point = "위반 후보로 바로 보지 말고 제목 정규화, 수정공고, 날짜가 다른 반복공고, 파일명 차이 때문에 낮게 나온 것인지 확인"
        elif "숫자" in finding:
            reason = "근거자료는 있으나 자동 숫자 매칭률이 낮아 증빙 위치 확인이 필요한 후보"
            point = "위반 후보로 바로 보지 말고 PDF/HWP 미파싱, 원/천원 단위, 표 영역 외 일반 숫자 때문에 낮게 나온 것인지 확인"
        else:
            reason = "근거자료는 있으나 자동 매칭률이 낮거나 부분 매칭이라 증빙 위치 확인이 필요한 후보"
            point = "위반 후보로 바로 보지 말고 파일명·단위·HWP/PDF 미파싱 때문에 낮게 나온 것인지 확인"
        return {
            "stage": "2차",
            "priority": "MEDIUM",
            "possibility": "하",
            "filter_reason": reason,
            "review_point": point,
            "evidence_note": evidence_note,
        }

    if review_type == "불성실공시 키워드":
        return {
            "stage": "2차",
            "priority": "LOW",
            "possibility": "하",
            "filter_reason": "본문 키워드 후보는 양식 안내문구와 실제 지적문구가 섞일 수 있어 1차에서 제외",
            "review_point": "실제 기관 소명/불성실공시 내역인지, 단순 양식 문구인지 원문 확인",
            "evidence_note": evidence_note,
        }

    if item == "공통/본문" or "본문/OCR" in review_type or "본문" in review_type:
        return {
            "stage": "2차",
            "priority": "LOW",
            "possibility": "하",
            "filter_reason": "본문/OCR 점수만 높은 후보는 문맥 확인 전 자동 확정 불가",
            "review_point": "실제 오류 문구인지, 양식 안내문구·참고문구·OCR 오인식인지 원문 확인",
            "evidence_note": evidence_note,
        }

    return None


def classify_row(root: Path, row: dict[str, Any]) -> dict[str, str]:
    review_type = text(row.get("검토유형"))
    item = text(row.get("항목"))
    finding = text(row.get("발견 요약"))
    judgment = text(row.get("차이/판정"))

    exclude_reason, evidence_note = classify_13_welfare_crosscheck(root, row)
    if exclude_reason:
        return {
            "stage": "제외",
            "priority": "EXCLUDE",
            "possibility": "제외",
            "filter_reason": exclude_reason,
            "review_point": "결과표 기재 대상 아님",
            "evidence_note": evidence_note,
        }

    second_pass = classify_second_pass_filter(row, evidence_note)
    if second_pass:
        return second_pass

    if review_type == "공시-ALIO 매칭확인":
        return {
            "stage": "2차",
            "priority": "MEDIUM",
            "possibility": "중",
            "filter_reason": "ALIO 원천 파일에는 있으나 기관명/고용형태/항목/연도 키 매칭이 되지 않은 후보",
            "review_point": "ALIO 원천 파일 존재, 기관명 정규화, 항목명 prefix, 고용형태 표기 차이를 확인",
            "evidence_note": evidence_note,
        }

    if review_type == "공시-제출 수치":
        return {
            "stage": "1차",
            "priority": "HIGH",
            "possibility": "상",
            "filter_reason": "공시값·제출값·셀 위치가 특정된 수치 후보",
            "review_point": "ALIO 값과 제출엑셀 셀값이 실제 다른지 확인",
            "evidence_note": evidence_note,
        }

    if review_type == "공시-제출 목록":
        return {
            "stage": "1차",
            "priority": "HIGH",
            "possibility": "상",
            "filter_reason": "6-2 목록 누락/초과 후보로 공고명 단위 확인 가능",
            "review_point": "제목 정규화((주)/㈜, 공백, (수정)) 후에도 누락인지 확인",
            "evidence_note": evidence_note,
        }

    if review_type == "제출엑셀 내부":
        return {
            "stage": "1차",
            "priority": "HIGH",
            "possibility": "중",
            "filter_reason": "제출엑셀 내부 교차검증 후보이며 관련 셀 위치가 특정됨",
            "review_point": "같은 기준연도·고용유형·항목끼리 비교한 것인지 확인",
            "evidence_note": evidence_note,
        }

    if review_type == "제출-근거자료" and finding == "근거자료 없음":
        return {
            "stage": "자료보완",
            "priority": "MEDIUM",
            "possibility": "자료보완",
            "filter_reason": "제출값은 있으나 항목별 근거자료 파일이 없어 산출근거 확인이 불가",
            "review_point": "다른 폴더명·압축파일·오제출 여부 확인 후, 없으면 자료보완 후보로 판단",
            "evidence_note": evidence_note,
        }

    return {
        "stage": "2차",
        "priority": "LOW",
        "possibility": "하",
        "filter_reason": "기본 2차 후보",
        "review_point": "필요 시 원천자료 확인",
        "evidence_note": evidence_note,
    }


OUTPUT_HEADERS = [
    "상태",
    "판단 결과",
    "재확인 필요",
    "검토 메모",
    "한줄요약",
    "검토단계",
    "최종표 반영가능성",
    "우선순위",
    "담당자",
    "기관번호",
    "기관",
    "항목",
    "검토유형",
    "검토포인트",
    "발견 요약",
    "대조군 값",
    "비교군 값",
    "차이/판정",
    "필터링 기준",
    "검산/보정 메모",
    "확인자료",
    "상세키",
    "원천순번",
]

COL_STAGE = OUTPUT_HEADERS.index("검토단계")
COL_PRIORITY = OUTPUT_HEADERS.index("우선순위")
COL_REVIEWER = OUTPUT_HEADERS.index("담당자")
COL_CHECK_NO = OUTPUT_HEADERS.index("기관번호")
COL_ITEM = OUTPUT_HEADERS.index("항목")
COL_SOURCE_NO = OUTPUT_HEADERS.index("원천순번")


def make_review_summary(row: dict[str, Any]) -> str:
    parts = [
        text(row.get("기관")),
        text(row.get("항목")),
        text(row.get("검토유형")),
        text(row.get("발견 요약")),
    ]
    summary = " | ".join(part for part in parts if part)
    verdict = text(row.get("차이/판정"))
    if verdict:
        summary = f"{summary} | {verdict}" if summary else verdict
    return summary


def output_row(row: dict[str, Any], classification: dict[str, str], assignment: dict[str, str]) -> list[Any]:
    return [
        "확인전",
        "미입력",
        "N",
        "",
        make_review_summary(row),
        classification["stage"],
        classification["possibility"],
        classification["priority"],
        assignment.get("reviewer", ""),
        assignment.get("check_no", ""),
        text(row.get("기관")),
        text(row.get("항목")),
        text(row.get("검토유형")),
        classification["review_point"],
        text(row.get("발견 요약")),
        text(row.get("대조군 값")),
        text(row.get("비교군 값")),
        text(row.get("차이/판정")),
        classification["filter_reason"],
        classification.get("evidence_note", ""),
        text(row.get("확인자료")),
        text(row.get("상세키/미매칭")),
        row.get("순번"),
    ]


def style_table(ws, review_sheet: bool = True) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.column <= 4:
                cell.protection = Protection(locked=False)
    widths = {
        "A": 12,
        "B": 14,
        "C": 12,
        "D": 34,
        "E": 72,
        "F": 10,
        "G": 14,
        "H": 10,
        "I": 8,
        "J": 10,
        "K": 30,
        "L": 12,
        "M": 18,
        "N": 44,
        "O": 34,
        "P": 36,
        "Q": 36,
        "R": 32,
        "S": 44,
        "T": 38,
        "U": 58,
        "V": 24,
        "W": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in ("W",):
        ws.column_dimensions[col].hidden = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "F2" if review_sheet else "A2"
    ws.auto_filter.ref = ws.dimensions
    if review_sheet:
        stage_col = get_column_letter(COL_STAGE + 1)
        priority_col = get_column_letter(COL_PRIORITY + 1)
        ws.conditional_formatting.add(
            f"{stage_col}2:{stage_col}{max(ws.max_row, 2)}",
            CellIsRule(operator="equal", formula=['"1차"'], fill=PatternFill("solid", fgColor="FFF2CC")),
        )
        ws.conditional_formatting.add(
            f"{stage_col}2:{stage_col}{max(ws.max_row, 2)}",
            CellIsRule(operator="equal", formula=['"자료보완"'], fill=PatternFill("solid", fgColor="D9EAF7")),
        )
        ws.conditional_formatting.add(
            f"{stage_col}2:{stage_col}{max(ws.max_row, 2)}",
            CellIsRule(operator="equal", formula=['"2차"'], fill=PatternFill("solid", fgColor="E2F0D9")),
        )
        ws.conditional_formatting.add(
            f"{stage_col}2:{stage_col}{max(ws.max_row, 2)}",
            CellIsRule(operator="equal", formula=['"제외"'], fill=PatternFill("solid", fgColor="D9D9D9")),
        )
        ws.conditional_formatting.add(
            f"{priority_col}2:{priority_col}{max(ws.max_row, 2)}",
            CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill("solid", fgColor="F8CBAD")),
        )
        add_dropdowns(ws)
    # Keep generated sheets unprotected so reviewers can filter, sort, copy, and paste without Excel prompts.
    ws.protection.sheet = False


def add_dropdowns(ws) -> None:
    end_row = max(ws.max_row, 2)
    status = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=False)
    judgment = DataValidation(type="list", formula1=f'"{JUDGMENT_VALUES}"', allow_blank=False)
    recheck = DataValidation(type="list", formula1=f'"{RECHECK_VALUES}"', allow_blank=False)
    ws.add_data_validation(status)
    ws.add_data_validation(judgment)
    ws.add_data_validation(recheck)
    status.add(f"A2:A{end_row}")
    judgment.add(f"B2:B{end_row}")
    recheck.add(f"C2:C{end_row}")


def add_sheet(wb: Workbook, title: str, rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append(row)
    style_table(ws, review_sheet=True)


def add_intro(wb: Workbook, generated_at: str, rows_by_stage: Counter[str], rows_by_reviewer: dict[str, Counter[str]]) -> None:
    ws = wb.active
    ws.title = "읽는법"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "1차 검토 체크리스트 필터링본"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"생성시각: {generated_at}"
    ws["A4"] = "바로 시작"
    ws["B4"] = "담당자 파일은 검토시작_1차 시트가 먼저 열립니다. A-D열만 입력하고, E열 한줄요약으로 사안을 먼저 파악합니다."
    ws["A5"] = "이 파일의 목적"
    ws["B5"] = "최종 결과표에 바로 옮길 가능성이 있는 후보만 1차로 올리고, 자료보완·단위보정·반올림·본문키워드·낮은매칭·수정공시 등은 별도 시트로 분리합니다."
    ws["A7"] = "1차 기준"
    ws["B7"] = "값·목록·셀 위치·내부교차검증처럼 확인 대상이 특정되는 후보"
    ws["A8"] = "자료보완 기준"
    ws["B8"] = "제출값은 있으나 항목별 근거자료 파일이 없어 증빙 확인이 막힌 후보"
    ws["A9"] = "2차 기준"
    ws["B9"] = "1,000배 단위차, 소액/반올림, 낮은 증빙매칭, 불성실공시/본문 키워드, 지적X/개선권고, 수정공시/정오표"
    ws["A11"] = "전체 건수"
    ws["B11"] = rows_by_stage["1차"]
    ws["C11"] = "1차"
    ws["B12"] = rows_by_stage["자료보완"]
    ws["C12"] = "자료보완"
    ws["B13"] = rows_by_stage["2차"]
    ws["C13"] = "2차"
    ws["B14"] = rows_by_stage["제외"]
    ws["C14"] = "제외"
    ws["A16"] = "담당자별 건수"
    ws.append([])
    ws.append(["담당자", "1차", "자료보완", "2차", "제외"])
    start = ws.max_row
    for reviewer in REVIEWERS:
        ws.append([
            reviewer,
            rows_by_reviewer[reviewer]["1차"],
            rows_by_reviewer[reviewer]["자료보완"],
            rows_by_reviewer[reviewer]["2차"],
            rows_by_reviewer[reviewer]["제외"],
        ])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in enumerate([14, 58, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for cell in ws[start]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)


def add_criteria_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("필터링기준")
    ws.append(["구분", "필터 단계", "기준", "검토자가 볼 포인트"])
    rows = [
        ["공시-ALIO 매칭확인", "2차", "ALIO 원천은 있으나 기관명·고용형태·항목·연도 키가 매칭되지 않은 후보", "ALIO 원천 파일, 기관명 정규화, 항목명 prefix, 고용형태 표기 차이 확인"],
        ["공시-제출 수치", "포함", "ALIO 값, 제출엑셀 값, 셀 위치가 특정됨", "단위 보정 후에도 차이가 남는지 확인"],
        ["공시-제출 수치: 1,000배 단위차", "2차", "비율이 약 1,000배 또는 1/1,000배", "원/천원 단위 보정 후에도 차이가 남으면 1차 후보로 승격"],
        ["공시-제출 수치: 소액/반올림/입력단위", "2차", "반올림, 소수점, 단수차이, 입력단위, 1원 차이 등", "허용오차와 기준상 지적 제외 가능성 확인"],
        ["공시-제출 목록", "포함", "6-2 공고명 단위 누락/초과 후보", "(주)/㈜, 공백, (수정) 표기 정규화 후 확인"],
        ["제출엑셀 내부", "포함", "같은 제출엑셀 안의 논리검증 후보", "같은 기준연도·고용유형·항목끼리 비교했는지 확인"],
        ["제출-근거자료: 근거자료 없음", "자료보완", "제출값은 있으나 항목별 증빙 파일이 없음", "다른 폴더명·압축파일·오제출 여부 확인 후 자료보완 판단"],
        ["제출-근거자료: 낮은 숫자 매칭", "2차", "근거파일은 있으나 자동 숫자 매칭률이 낮은 증빙 위치 확인 후보", "PDF/HWP 미파싱, 단위, 표 영역 외 숫자 여부 확인"],
        ["제출-근거자료: 낮은 제목 매칭", "2차", "6-2 근거파일은 있으나 제목 매칭률이 낮은 증빙 위치 확인 후보", "수정공고, 반복공고, 날짜·파일명 차이 여부 확인"],
        ["불성실공시 키워드", "2차", "양식 안내문구와 실제 지적문구가 섞임", "실제 불성실공시 내역인지 원문 확인"],
        ["본문/OCR 점수만 높은 건", "2차", "본문/OCR 키워드나 점수만 높고 값·셀 위치가 특정되지 않음", "문맥상 실제 오류인지, OCR 오인식인지 확인"],
        ["2024 기준상 지적 X/개선권고", "2차", "기준 엑셀 비고 또는 후보 문구가 지적 X, 지적 제외, 개선권고 성격", "같은 선례인지 확인 후 제외 또는 참고 처리"],
        ["수정공시/정오표 이슈", "2차", "수정공시·정오표는 ALIO 이력 확인 전 자동 확정 불가", "수정공시 일자와 정오표 반영 여부 확인"],
        ["13-1/13-5 합산 일치", "제외", "13-1 급여성+비급여성 합산값이 13-5 예산총합과 일치", "자동 불일치 후보에서 제외"],
        ["6-2 필수 첨부파일 없음", "추가필터", "공고문·입사지원서·직무기술서 파일명/첨부구분 미확인", "공고문/첨부 구조화 후 1차 후보로 추가"],
        ["6-2 근무분야 공란/기타", "추가필터", "2024년 4월 이후 공개 건 중 근무분야 미기재/기타", "공고문 구조화 후 1차 후보로 추가"],
        ["11-1 직원 평균보수", "추가필터", "상시종업원수, 평균근속연수, 성별 평균임금, 수당 첨부파일 대조", "11-1 파서 추가 후 1차/2차 분리"],
        ["5-1 개인별 속성", "추가필터", "지역인재, 고졸, 직종, 계약유형, 임용일 판단", "개인별 로데이터 속성 파싱 후 1차 후보 추가"],
        ["10-1 비고/성과상여금/상임현원", "추가필터", "연봉제 비고, 성과상여금 미공시, 상임현원 있는데 10-1 미공시", "10-1 전용 규칙 추가 후 1차/2차 분리"],
    ]
    for row in rows:
        ws.append(row)
    style_table(ws, review_sheet=False)
    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = {"A": 28, "B": 16, "C": 54, "D": 54}[col]


def sort_output(rows: list[list[Any]]) -> list[list[Any]]:
    stage_rank = {"1차": 0, "자료보완": 1, "2차": 2, "제외": 3}
    priority_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "EXCLUDE": 9}
    reviewer_rank = {reviewer: idx for idx, reviewer in enumerate(REVIEWERS)}
    return sorted(
        rows,
        key=lambda r: (
            stage_rank.get(text(r[COL_STAGE]), 9),
            reviewer_rank.get(text(r[COL_REVIEWER]), 9),
            priority_rank.get(text(r[COL_PRIORITY]), 9),
            text(r[COL_CHECK_NO]),
            text(r[COL_ITEM]),
            number(r[COL_SOURCE_NO]) or 999999,
        ),
    )


def save_workbook(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_{dt.datetime.now().strftime('%H%M%S')}{path.suffix}")
        wb.save(fallback)
        return fallback


def set_active_sheet(wb: Workbook, sheet_name: str) -> None:
    if sheet_name in wb.sheetnames:
        wb.active = wb.sheetnames.index(sheet_name)


def write_text(path: Path, text_value: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text_value.rstrip() + "\n", encoding="utf-8-sig")
    return path


def write_start_guides(
    out_dir: Path,
    start_dir: Path,
    generated_at: str,
    stage_counts: Counter[str],
    by_reviewer: dict[str, Counter[str]],
) -> list[Path]:
    total = sum(stage_counts.values())
    summary_rows = "\n".join(
        f"| {stage} | {stage_counts.get(stage, 0)} |"
        for stage in ("1차", "자료보완", "2차", "제외")
    )
    reviewer_labels = [
        *REVIEWERS,
        *sorted(reviewer for reviewer in by_reviewer.keys() if reviewer not in REVIEWERS),
    ]
    reviewer_rows = "\n".join(
        f"| {reviewer} | {by_reviewer[reviewer].get('1차', 0)} | {by_reviewer[reviewer].get('자료보완', 0)} | {by_reviewer[reviewer].get('2차', 0)} | {by_reviewer[reviewer].get('제외', 0)} |"
        for reviewer in reviewer_labels
    )
    reviewer_queue_rows = "\n".join(
        f"| {reviewer} | `01_{reviewer}_검토큐_필터링.xlsx` | `검토시작_1차` |"
        for reviewer in REVIEWERS
    )
    reviewer_queue_list = ", ".join(
        f"`30_검토산출물\\00_00_먼저보기\\01_{reviewer}_검토큐_필터링.xlsx`"
        for reviewer in REVIEWERS
    )

    top_start = f"""# 검토 시작 안내

생성시각: {generated_at}

최종 결과표에 옮길 후보 검토는 `30_검토산출물\\00_00_먼저보기`에서 시작합니다.
나머지 폴더는 근거 추적, 자동검토 원천, 로그, 압축본입니다.

## 먼저 열 파일

| 사용자 | 먼저 열 파일 | 먼저 볼 시트/문서 | 목적 |
|---|---|---|---|
| 전체 | `30_검토산출물\\00_00_먼저보기\\00_읽는순서.md` | 문서 전체 | 자기 역할별 시작점 확인 |
| 관리자 | `30_검토산출물\\00_00_먼저보기\\01_관리자용_전체검토현황_필터링.xlsx` | `1차_필터링체크리스트` | 전체 배분, 미배정, 단계별 물량 확인 |
| 담당자 | 담당자별 `01_*_검토큐_필터링.xlsx` | `검토시작_1차` | 자기 1차 검토 대상 처리 |
| 결과표 작성자 | `30_검토산출물\\00_00_먼저보기\\04_결과표_기재초안.xlsx` | `결과표_기재후보` | 확인 완료 후보를 결과표 문안으로 전사 |
| 기준 점검자 | `30_검토산출물\\00_00_먼저보기\\02_2024유형대비_필터커버리지.xlsx` | 첫 시트 | 현행 필터가 못 줄이는 2024년 지적유형 확인 |
| 근거 추적자 | `30_검토산출물\\00_검토용_핵심산출물\\00_여기부터_검토시작.xlsx` | `상세후보_값대조` | 필터링 큐의 원천 근거 추적 |

## 필터링 결과

| 구분 | 건수 |
|---|---:|
{summary_rows}
| 전체 | {total} |

| 담당자 | 1차 | 자료보완 | 2차 | 제외 |
|---|---:|---:|---:|---:|
{reviewer_rows}

## 기준

- `1차`: 공시-제출 수치, 공시-제출 목록, 제출엑셀 내부 불일치처럼 값·목록·셀 위치가 특정되어 최종 결과표 기재 가능성이 높은 후보입니다.
- `자료보완`: 제출값은 있으나 항목별 근거자료 파일이 없어 산출근거 확인이 막힌 후보입니다.
- `2차`: 1,000배 단위차, 소액/반올림, 낮은 증빙매칭, 불성실공시/본문 키워드, 지적X/개선권고, 수정공시/정오표처럼 1차 후 참고할 후보입니다.
- `제외`: 합산검산 등으로 이미 설명 가능한 오탐 후보입니다.

## 나머지 폴더

- `90_원천_자동검토`: 원천 PDF, 숫자대조/본문검토 산출물, 자동검토 원본입니다.
- `80_압축본`: 전달/보관용 ZIP입니다.
- `99_보관_로그`: 백업, 기타 작업목록, 실행 로그입니다.
"""

    read_order = f"""# 먼저 볼 파일

이 폴더만 먼저 보면 됩니다. 다른 폴더는 근거 추적이나 재생성용입니다.

생성시각: {generated_at}

## 필터링 결과

| 구분 | 건수 |
|---|---:|
{summary_rows}
| 전체 | {total} |

## 사용자별 시작점

| 사용자 | 먼저 열 파일 | 먼저 볼 시트/문서 |
|---|---|---|
| 관리자 | `01_관리자용_전체검토현황_필터링.xlsx` | `1차_필터링체크리스트` |
| 담당자 | 아래 담당자별 파일 | `검토시작_1차` |
| 결과표 작성자 | `04_결과표_기재초안.xlsx` | `결과표_기재후보` |
| 시스템 점검자 | `05_자동화시스템_검토보고서.md` | 문서 전체 |
| 기준 점검자 | `02_2024유형대비_필터커버리지.xlsx` | 첫 시트 |
| 근거 추적자 | `..\\00_검토용_핵심산출물\\00_여기부터_검토시작.xlsx` | `상세후보_값대조` |

## 담당자별 검토

| 담당자 | 먼저 열 파일 | 먼저 볼 시트 |
|---|---|---|
{reviewer_queue_rows}

필터가 줄이지 못하는 2024년 1차 지적유형은 `02_2024유형대비_필터커버리지.xlsx`에서 확인합니다.

## 읽는 순서

1. 엑셀을 열면 실제 작업 시트가 먼저 뜹니다. 담당자는 `검토시작_1차`, 관리자는 `1차_필터링체크리스트`부터 봅니다.
2. `한줄요약`으로 기관, 항목, 검토유형, 발견내용, 차이/판정을 먼저 봅니다.
3. 필요하면 오른쪽의 `검토포인트`, `대조군 값`, `비교군 값`, `필터링 기준`을 봅니다.
4. 후보가 맞으면 `확인자료`의 제출엑셀/ALIO/근거자료를 엽니다.
5. 판단 후 앞쪽 `상태`, `판단 결과`, `재확인 필요`, `검토 메모`만 입력합니다.
6. `검토시작_1차`가 끝난 뒤 `자료보완_근거자료없음`, `2차_참고` 순서로 봅니다. `제외_오탐후보`는 원칙적으로 확인용입니다.
7. 결과표에 옮길 때는 `04_결과표_기재초안.xlsx`의 문안 초안을 사용하되, 불성실 유형과 위반수준은 원문 확인 후 확정합니다.

## 시트 의미

| 시트 | 의미 |
|---|---|
| `검토시작_1차` | 최종 결과표에 옮길 가능성이 있는 후보 |
| `자료보완_근거자료없음` | 제출값은 있으나 근거자료 파일이 없어 자료보완 여부를 볼 후보 |
| `2차_참고` | 단위보정, 반올림, 낮은 증빙매칭, 본문 키워드, 수정공시 등 1차 후 볼 후보 |
| `제외_오탐후보` | 합산검산 등으로 1차에서 제외한 후보와 제외 사유 |
| `필터링기준` | 왜 1차/2차/제외로 나눴는지 기준 |
| `04_결과표_기재초안.xlsx` | 결과표 양식으로 옮기기 위한 문안 초안. 자동 확정표는 아님 |

## 분류 예시

예시 기관은 1차에 `6-2 목록 누락 가능`을 남기고, `13-5 근거자료 없음`은 `자료보완_근거자료없음`으로 분리했습니다.
13-1/13-5 내부 불일치 후보는 급여성+비급여성 합산값이 13-5 예산총합과 맞아 `제외_오탐후보`로 내렸습니다.
"""
    folder_map = """# 30_검토산출물 폴더 지도

## 일반 검토자가 여는 곳

| 위치 | 용도 |
|---|---|
| `00_00_먼저보기` | 실제 검토 시작 지점. 필터링된 1차 체크리스트만 모음 |

## 필요할 때만 여는 곳

| 위치 | 용도 | 여는 경우 |
|---|---|---|
| `00_검토용_핵심산출물` | 필터링 전 구조화 후보와 기존 큐 | 1차 체크리스트의 근거를 더 추적할 때 |
| `90_원천_자동검토\\01_ALIO_원문PDF` | 기관별 ALIO 원문 PDF | ALIO 화면 원문 확인이 필요할 때 |
| `90_원천_자동검토\\02_ALIO_숫자대조` | PDF/엑셀 숫자 추출 원천 | 숫자 후보의 원천 추적 |
| `90_원천_자동검토\\03_숫자대조_숏리스트` | 숫자대조 CSV 후보 | 필터링 전 숫자 후보를 더 볼 때 |
| `90_원천_자동검토\\04_증빙본문_텍스트추출` | 증빙 PDF/HWP/OCR 텍스트 | 본문 키워드 근거 확인 |
| `90_원천_자동검토\\05_본문검토_숏리스트` | 본문 키워드 후보 CSV | 2차 참고 후보를 더 볼 때 |
| `90_원천_자동검토\\06_auto_review` | 자동검토 원본 엑셀 | 필터링 파일을 다시 만들거나 원천 로직 확인 |
| `90_원천_자동검토\\06_검토리포트` | 과거 리포트 산출물 | 기존 상세 리포트 확인이 필요할 때 |

## 열지 않아도 되는 곳

| 위치 | 용도 |
|---|---|
| `80_압축본` | 전달/참고용 압축본 |
| `99_보관_로그\\97_이전정규화자료_백업` | 과거 정규화 자료 백업 |
| `99_보관_로그\\98_기타작업목록` | 보조 작업 목록 |
| `99_보관_로그\\99_실행로그` | 재현용 실행 로그 |

## 재생성 명령

필터링 체크리스트만 다시 만들 때:

```powershell
python .\\90_tools\\filter_first_pass_checklist.py --root .
```

전체 검토 패키지를 다시 만들 때:

```powershell
python .\\90_tools\\run_review_pipeline.py --root . --use-existing-auto
```
"""
    flow_audit = f"""# 사용자 플로우 가독성 점검

생성시각: {generated_at}

## 결론

최종 검토산출물만 보고 1차 검토를 시작할 수 있도록 구성했습니다.
검토자는 `00_00_먼저보기`만 열고, 자기 역할에 맞는 파일에서 시작하면 됩니다.

## 플로우별 점검

| 단계 | 사용자 | 확인할 것 | 점검 결과 |
|---|---|---|---|
| 1 | 전체 | `30_검토산출물` 진입 후 시작 위치가 보이는가 | `README_먼저읽기.md`, `00_검토_시작.md`, `00_00_먼저보기`로 시작점이 분리됨 |
| 2 | 관리자 | 전체 물량과 담당자 배분을 바로 볼 수 있는가 | 관리자 파일은 `1차_필터링체크리스트`가 활성 시트로 열림 |
| 3 | 담당자 | 자기 작업 파일과 1차 시트를 바로 찾는가 | 담당자 파일은 `검토시작_1차`가 활성 시트로 열림 |
| 4 | 담당자 | 행마다 판단할 내용을 빠르게 이해하는가 | A-D열은 입력, E열은 `한줄요약`, 오른쪽은 상세 근거로 분리됨 |
| 5 | 담당자 | 1차 이후 무엇을 볼지 알 수 있는가 | `자료보완_근거자료없음` → `2차_참고` → `제외_오탐후보` 순서로 안내 |
| 6 | 근거 추적자 | 원천 근거를 어디서 찾아야 하는가 | `확인자료` 열과 `00_검토용_핵심산출물\\00_여기부터_검토시작.xlsx`를 연결 |
| 7 | 기준 점검자 | 2024년 지적유형 대비 필터 한계를 볼 수 있는가 | `02_2024유형대비_필터커버리지.xlsx`와 `.md`를 별도 제공 |
| 8 | 결과표 작성자 | 결과표 양식에 옮길 문안 초안이 있는가 | `04_결과표_기재초안.xlsx`에서 결과표 행, 불성실 유형 추정, 문안 초안을 제공 |

## 현재 물량

| 구분 | 건수 |
|---|---:|
{summary_rows}
| 전체 | {total} |

## 담당자별 물량

| 담당자 | 1차 | 자료보완 | 2차 | 제외 |
|---|---:|---:|---:|---:|
{reviewer_rows}

## 남은 리스크

- 이 산출물은 자동 확정표가 아니라 사람이 먼저 볼 행을 줄이는 1차 큐입니다.
- `04_결과표_기재초안.xlsx`도 자동 확정표가 아니며, 확인 완료 건만 결과표에 반영해야 합니다.
- `자료보완_근거자료없음`은 실제 오류 확정이 아니라 증빙 경로, 압축파일, 오제출 여부 확인이 먼저 필요합니다.
- 2024년 1차 지적유형 중 6-2 첨부파일/근무분야, 11-1 보수, 5-1 개인별 속성, 10-1 비고성 항목은 추가 파서 없이는 자동 축소가 제한됩니다.
- 원천자료에 긴 경로 또는 읽기 불가 폴더가 있으면 자동검토가 일부 증빙을 건너뛸 수 있습니다.
"""
    return [
        write_text(out_dir / "00_검토_시작.md", top_start),
        write_text(start_dir / "00_읽는순서.md", read_order),
        write_text(start_dir / "03_사용자플로우_가독성점검.md", flow_audit),
        write_text(start_dir / "99_폴더지도.md", folder_map),
    ]


def find_2024_type_book(root: Path) -> Path | None:
    patterns = ["*24*1차*공시점검*유형*.xlsx", "*24*점검*유형*.xlsx"]
    for pattern in patterns:
        matches = [p for p in root.glob(pattern) if p.is_file() and not p.name.startswith("~$")]
        if matches:
            return sorted(matches, key=lambda p: p.name)[0]
    return None


def is_type_noise(value: str) -> bool:
    value = flat(value, 400)
    return (
        not value
        or value in {" ", "-"}
        or "2차 점검" in value
        or "3차 점검" in value
        or "주요 이의신청 유형" in value
        or value.startswith("1차 점검 _")
        or value.startswith("점검항목")
    )


def extract_2024_first_round_rules(root: Path) -> list[dict[str, str]]:
    path = find_2024_type_book(root)
    if path is None:
        return []
    wb = load_workbook(path, read_only=False, data_only=True)
    rules: list[dict[str, str]] = []
    for ws in wb.worksheets:
        max_row, max_col = ws.max_row or 0, ws.max_column or 0
        group_starts: list[int] = []
        for row_idx in range(1, min(max_row, 8) + 1):
            for col_idx in range(1, max_col + 1):
                if "24년도 1차 점검" in text(ws.cell(row_idx, col_idx).value):
                    group_starts.append(col_idx)
        if not group_starts:
            continue

        header_row = None
        for row_idx in range(1, min(max_row, 10) + 1):
            if any("지적 유형" in text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, max_col + 1)):
                header_row = row_idx
                break
        if header_row is None:
            continue

        for start_col in sorted(set(group_starts)):
            issue_col = None
            for col_idx in range(start_col, min(max_col, start_col + 4) + 1):
                if "지적 유형" in text(ws.cell(header_row, col_idx).value):
                    issue_col = col_idx
                    break
            if issue_col is None:
                issue_col = start_col

            marker_seen = False
            for row_idx in range(header_row + 1, max_row + 1):
                issue = text(ws.cell(row_idx, issue_col).value)
                issue_flat = flat(issue, 500)
                if "2차 점검" in issue_flat or "3차 점검" in issue_flat or "주요 이의신청 유형" in issue_flat:
                    marker_seen = True
                    continue
                if marker_seen or is_type_noise(issue):
                    continue
                rules.append(
                    {
                        "sheet": ws.title,
                        "row": str(row_idx),
                        "issue": issue_flat,
                        "opinion": flat(ws.cell(row_idx, issue_col + 1).value, 260),
                        "note": flat(ws.cell(row_idx, issue_col + 2).value, 160),
                    }
                )
    wb.close()
    return rules


def classify_2024_coverage(rule: dict[str, str]) -> tuple[str, str, str]:
    sheet = text(rule.get("sheet"))
    issue = text(rule.get("issue"))
    combined = f"{sheet} {issue}"

    if sheet.startswith("11-1."):
        return (
            "미커버",
            "현행 필터는 11-1 직원 평균보수 파서를 사용하지 않음",
            "상시종업원수, 평균근속연수, 성별 평균임금, 수당 첨부파일 대조 로직 추가 필요",
        )

    if sheet.startswith("6-2."):
        if "목록" in issue or "미공시" in issue:
            return (
                "부분커버",
                "현행 필터는 ALIO-제출 채용공고 목록 누락/초과만 1차로 올림",
                "근무분야, 결과확정일, 공개일, 필수첨부파일은 공고문/첨부 구조화 필요",
            )
        return (
            "미커버",
            "2024년 6-2 주요 유형은 근무분야·결과확정일·공개일·첨부파일 내용 검토",
            "채용공고 본문, 내부결재일, 첨부파일 3종을 구조화하지 않으면 자동 필터로 줄이기 어려움",
        )

    if sheet.startswith("5-1."):
        if any(token in issue for token in ("채용인원", "인원", "공시", "불일치")):
            return (
                "부분커버",
                "ALIO-제출 수치 차이는 1차로 잡지만 지역인재·고졸·직종 속성 판단은 제한적",
                "개인별 로데이터의 생년, 학력, 지역인재, 직종/계약유형 속성 파싱 필요",
            )
        return (
            "미커버",
            "5-1 유형은 제출 로데이터 속성 판단이 중심",
            "개인별 로데이터 규칙화 필요",
        )

    if sheet.startswith("10-1."):
        if "급여성 복리후생비" in issue and "13-1" in issue:
            return (
                "부분커버",
                "10-1/13-1 금액 불일치 성격은 일부 자동 후보화 가능",
                "비고란 연봉제, 성과상여금 미공시, 만근 환산, 성별 보수 표기 등은 별도 규칙 필요",
            )
        if any(token in combined for token in ("평균보수", "평균연봉", "합계연봉", "기본급", "금액", "성과상여금")):
            return (
                "부분커버",
                "공시-제출 수치 차이는 1차로 잡을 수 있으나 항목 의미 판단은 제한적",
                "임원 현원, 성별 보수, 연봉제 비고, 성과상여금 항목 존재 여부 규칙 추가 필요",
            )
        return (
            "미커버",
            "10-1 텍스트·비고·해당사항 없음 판단은 현행 필터 밖",
            "10-1 전용 규칙 추가 필요",
        )

    if sheet.startswith("13-1."):
        if "13-5" in issue:
            return (
                "직접커버",
                "13-1/13-5 예산 재원 불일치 후보와 합산 일치 오탐 제외를 반영",
                "소요재원 구분 자체가 맞는지는 수동 확인 필요",
            )
        if "수정공시" in issue or "정오표" in issue:
            return (
                "미커버",
                "수정공시·정오표 누락은 현재 ALIO 이력/정오표 구조화가 없음",
                "수정공시 이력과 정오표 수집 로직 필요",
            )
        return (
            "부분커버",
            "13-1 수치 차이와 근거자료 없음은 1차로 잡힘",
            "사복기금/예산 소요재원 구분은 항목 의미 규칙 추가 필요",
        )

    if sheet.startswith("13-5."):
        if "11-1" in issue or "상시종업원수" in issue:
            return (
                "부분커버",
                "13-1/13-5 총액 불일치는 잡지만 11-1 상시종업원수 연계는 미구현",
                "11-1 상시종업원수 파서와 13-5 평균인원 비교 로직 추가 필요",
            )
        return (
            "직접커버",
            "13-5와 13-1 총액 불일치 후보 및 합산 일치 제외를 반영",
            "평균인원 산식과 소수점/반올림은 추가 확인 필요",
        )

    return (
        "미커버",
        "현행 필터의 직접 대상 항목 밖",
        "해당 항목 전용 파서 또는 수동 체크리스트 필요",
    )


def write_coverage_audit(root: Path, start_dir: Path, generated_at: str) -> list[Path]:
    rules = extract_2024_first_round_rules(root)
    if not rules:
        return []

    coverage_rows: list[list[Any]] = []
    status_counts: Counter[str] = Counter()
    sheet_counts: dict[str, Counter[str]] = defaultdict(Counter)
    for rule in rules:
        status, coverage, action = classify_2024_coverage(rule)
        status_counts[status] += 1
        sheet_counts[rule["sheet"]][status] += 1
        coverage_rows.append([
            status,
            rule["sheet"],
            rule["row"],
            rule["issue"],
            rule["note"],
            coverage,
            action,
        ])

    xlsx_path = start_dir / "02_2024유형대비_필터커버리지.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    ws.append(["생성시각", generated_at])
    ws.append(["기준", "'24년도 1차 공시점검 점검 유형 정리.xlsx의 24년도 1차 점검 지적유형"])
    ws.append([])
    ws.append(["판정", "건수"])
    for status in ("직접커버", "부분커버", "미커버"):
        ws.append([status, status_counts.get(status, 0)])
    ws.append(["전체", len(rules)])
    ws.append([])
    ws.append(["항목", "직접커버", "부분커버", "미커버", "전체"])
    for sheet in sorted(sheet_counts):
        counter = sheet_counts[sheet]
        ws.append([
            sheet,
            counter.get("직접커버", 0),
            counter.get("부분커버", 0),
            counter.get("미커버", 0),
            sum(counter.values()),
        ])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in {"A": 24, "B": 70, "C": 14, "D": 14, "E": 14}.items():
        ws.column_dimensions[col].width = width

    detail = wb.create_sheet("유형별커버리지")
    detail.append(["판정", "항목", "원본행", "2024 지적유형", "2024 비고", "현행 필터 커버리지", "보완 필요 작업"])
    for row in coverage_rows:
        detail.append(row)
    for row in detail.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {"A": 14, "B": 28, "C": 10, "D": 80, "E": 22, "F": 46, "G": 54}
    for col, width in widths.items():
        detail.column_dimensions[col].width = width
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    saved_xlsx = save_workbook(wb, xlsx_path)

    sheet_lines = "\n".join(
        f"| {sheet} | {counter.get('직접커버', 0)} | {counter.get('부분커버', 0)} | {counter.get('미커버', 0)} | {sum(counter.values())} |"
        for sheet, counter in sorted(sheet_counts.items())
    )
    md = f"""# 2024 유형 대비 필터 커버리지

생성시각: {generated_at}

## 결론

현행 필터는 전수검사 전 `우선순위 큐`로는 쓸 수 있지만, 전수검사를 크게 줄이는 `누락방지형 필터`로는 부족합니다.
값·목록·근거자료 부재처럼 기계적으로 특정 가능한 후보는 줄여주지만, 2024년 주요 지적유형 중 11-1 직원 평균보수, 6-2 채용공고 내용, 5-1 개인별 속성 판단은 아직 수동 확인이 필요합니다.

## 전체 판정

| 판정 | 건수 |
|---|---:|
| 직접커버 | {status_counts.get('직접커버', 0)} |
| 부분커버 | {status_counts.get('부분커버', 0)} |
| 미커버 | {status_counts.get('미커버', 0)} |
| 전체 | {len(rules)} |

## 항목별 커버리지

| 항목 | 직접커버 | 부분커버 | 미커버 | 전체 |
|---|---:|---:|---:|---:|
{sheet_lines}

## 사용 기준

- `직접커버`: 현행 1차 필터에서 같은 성격의 후보를 직접 올리거나, 합산검산으로 제외합니다.
- `부분커버`: 숫자 차이 등 일부 징후는 잡지만 2024년 지적유형의 판단요소 전체를 커버하지 못합니다.
- `미커버`: 전용 파서나 로데이터 속성 판단 없이 현재 필터로는 업무 경감 효과가 작습니다.

상세 행별 판단은 `02_2024유형대비_필터커버리지.xlsx`의 `유형별커버리지` 시트를 봅니다.
"""
    saved_md = write_text(start_dir / "02_2024유형대비_필터커버리지.md", md)
    return [saved_xlsx, saved_md]


def build_filtered_workbooks(root: Path) -> list[Path]:
    out_dir = find_dir(root, "30_")
    core_dir = find_core_dir(out_dir)
    start_dir = out_dir / "00_00_먼저보기"
    start_dir.mkdir(parents=True, exist_ok=True)
    report = find_report(core_dir)
    admin_queue = find_admin_queue(core_dir)
    assignment = build_assignment_map(admin_queue)
    detail_rows = read_sheet_rows(report, "상세후보_값대조", 2)

    output_rows: list[list[Any]] = []
    by_reviewer: dict[str, Counter[str]] = defaultdict(Counter)
    for row in detail_rows:
        agency = text(row.get("기관"))
        assign = lookup_assignment(agency, assignment)
        classification = classify_row(root, row)
        out = output_row(row, classification, assign)
        output_rows.append(out)
        reviewer = text(out[COL_REVIEWER]) or "미배정"
        by_reviewer[reviewer][text(out[COL_STAGE])] += 1

    include_reviewers(by_reviewer.keys())
    output_rows = sort_output(output_rows)
    first_rows = [row for row in output_rows if row[COL_STAGE] == "1차"]
    supplement_rows = [row for row in output_rows if row[COL_STAGE] == "자료보완"]
    second_rows = [row for row in output_rows if row[COL_STAGE] == "2차"]
    excluded_rows = [row for row in output_rows if row[COL_STAGE] == "제외"]
    stage_counts = Counter(row[COL_STAGE] for row in output_rows)
    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    written: list[Path] = []
    admin_path = start_dir / "01_관리자용_전체검토현황_필터링.xlsx"
    wb = Workbook()
    add_intro(wb, generated_at, stage_counts, by_reviewer)
    add_sheet(wb, "1차_필터링체크리스트", first_rows)
    add_sheet(wb, "자료보완_근거자료없음", supplement_rows)
    add_sheet(wb, "2차_참고", second_rows)
    add_sheet(wb, "제외_오탐후보", excluded_rows)
    for reviewer in REVIEWERS:
        add_sheet(wb, f"{reviewer}_1차", [row for row in first_rows if row[COL_REVIEWER] == reviewer])
    add_criteria_sheet(wb)
    set_active_sheet(wb, "1차_필터링체크리스트")
    written.append(save_workbook(wb, admin_path))

    for reviewer in REVIEWERS:
        reviewer_path = start_dir / f"01_{reviewer}_검토큐_필터링.xlsx"
        wb = Workbook()
        reviewer_rows = [row for row in output_rows if row[COL_REVIEWER] == reviewer]
        reviewer_stage_counts = Counter(row[COL_STAGE] for row in reviewer_rows)
        reviewer_counts = defaultdict(Counter)
        reviewer_counts[reviewer].update(reviewer_stage_counts)
        add_intro(wb, generated_at, reviewer_stage_counts, reviewer_counts)
        add_sheet(wb, "검토시작_1차", [row for row in reviewer_rows if row[COL_STAGE] == "1차"])
        add_sheet(wb, "자료보완_근거자료없음", [row for row in reviewer_rows if row[COL_STAGE] == "자료보완"])
        add_sheet(wb, "2차_참고", [row for row in reviewer_rows if row[COL_STAGE] == "2차"])
        add_sheet(wb, "제외_오탐후보", [row for row in reviewer_rows if row[COL_STAGE] == "제외"])
        add_criteria_sheet(wb)
        set_active_sheet(wb, "검토시작_1차")
        written.append(save_workbook(wb, reviewer_path))

    written.extend(write_start_guides(out_dir, start_dir, generated_at, stage_counts, by_reviewer))
    written.extend(write_coverage_audit(root, start_dir, generated_at))
    return written


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".", help="Project root")
    args = parser.parse_args()
    root = Path(args.root).resolve()
    written = build_filtered_workbooks(root)
    for path in written:
        print(path)


if __name__ == "__main__":
    main()
