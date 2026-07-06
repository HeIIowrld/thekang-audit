# -*- coding: utf-8 -*-
"""
Match institution-submitted standard workbooks to supporting evidence files.

This is a lightweight local process. It does not try to prove every item from
every PDF/HWP, but it does the repeatable first pass:
- finds submitted standard workbooks;
- finds evidence folders/files for 5-1, 6-2, 10-1, 13-1, 13-5;
- checks whether evidence exists per institution/item;
- compares numeric signatures in submitted sheets against evidence Excel files;
- checks 6-2 submitted job titles against evidence file names/text snippets;
- writes a review workbook with status and drill-down rows.

Usage:
    python .\\90_tools\\evidence_match_review.py --root .
    python .\\90_tools\\evidence_match_review.py --root . --limit 20
"""

from __future__ import annotations

import argparse
import datetime as dt
import math
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Install it with: python -m pip install openpyxl"
    ) from exc


TARGET_ITEMS = ("5-1", "6-2", "10-1", "13-1", "13-5")
EVIDENCE_EXTS = {".xlsx", ".xls", ".xlsm", ".pdf", ".hwp", ".hwpx", ".docx", ".doc", ".zip"}
MAX_UNMATCHED_VALUES_PER_ROW = 25
DEFAULT_MAX_EVIDENCE_EXCELS = 2
DEFAULT_MAX_CELLS_PER_EVIDENCE = 5000
DEFAULT_WORKBOOK_KEYWORDS = ("review", "disclosure", "checklist", "노무", "통합공시", "점검표")
DEFAULT_STEM_STRIP_PATTERNS = (
    r"_?노무.*$",
    r"_?통합공시.*$",
    r"_?점검표.*$",
    r"_?disclosure.*$",
    r"_?review.*$",
    r"_?checklist.*$",
    r"20\d{2}년도.*$",
)
GENERIC_ORG_KEYS = {
    "공기업",
    "시장형",
    "준시장형",
    "준정부기관",
    "기금관리형",
    "위탁집행형",
    "기타공공기관",
    "부설기관",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def flat_text(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value).replace("\n", " ")).strip()


def parse_keyword_args(values: list[str] | None) -> tuple[str, ...]:
    if values is None:
        return DEFAULT_WORKBOOK_KEYWORDS
    keywords: list[str] = []
    for value in values:
        keywords.extend(token.strip() for token in value.split(",") if token.strip())
    return tuple(keywords)


def name_matches_keywords(path: Path, keywords: tuple[str, ...]) -> bool:
    if not keywords:
        return True
    name = path.name.lower()
    return any(keyword.lower() in name for keyword in keywords)


def strip_generic_suffixes(stem: str) -> str:
    for pattern in DEFAULT_STEM_STRIP_PATTERNS:
        stem = re.sub(pattern, "", stem, flags=re.IGNORECASE)
    return stem.strip(" _-")


def compact(value: Any) -> str:
    return re.sub(r"\s+", "", text(value))


def normalize_org_name(value: Any) -> str:
    s = compact(value)
    replacements = [
        "주식회사", "(주)", "㈜", "재단법인", "사단법인", "(재)", "(사)",
        "의료법인", "학교법인",
    ]
    for token in replacements:
        s = s.replace(token, "")
    return re.sub(r"[\(\)\[\]{}·ㆍ\.,_\-\s]", "", s)


def usable_org_key(value: Any) -> str:
    key = normalize_org_name(value)
    return "" if len(key) < 3 or key in GENERIC_ORG_KEYS else key


def org_key_in_path(org_key: str, path: Path) -> bool:
    if not org_key:
        return False
    return org_key in normalize_org_name(str(path))


def normalize_title(value: Any) -> str:
    s = flat_text(value).lower()
    s = re.sub(r"^\(수정\)\s*", "", s)
    s = re.sub(r"\s+", "", s)
    return re.sub(r"[\[\]\(\){}·ㆍ,._\-]", "", s)


def parse_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    s = flat_text(value)
    if not s or s in {"-", "－", "해당없음", "해당 없음"}:
        return None
    s = s.replace(",", "").replace("천원", "").replace("명", "").replace("%", "")
    s = re.sub(r"^\(([-+]?\d+(?:\.\d+)?)\)$", r"-\1", s)
    try:
        return float(s)
    except ValueError:
        return None


def number_key(value: Any) -> str | None:
    num = parse_number(value)
    if num is None:
        return None
    if abs(num) < 1e-9:
        return None
    # Drop obvious years and date fragments; they create noise in evidence matching.
    if float(num).is_integer() and 1900 <= int(num) <= 2100:
        return None
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    return f"{num:.4f}".rstrip("0").rstrip(".")


def find_cells(ws: Any, needle: str, max_rows: int = 80) -> list[tuple[int, int]]:
    hits = []
    for r in range(1, min(ws.max_row or 0, max_rows) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            if needle in flat_text(ws.cell(r, c).value):
                hits.append((r, c))
    return hits


def sheet_has_token(workbook_path: Path, token: str) -> bool:
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return False
    return any(token in name for name in wb.sheetnames)


def workbook_has_target_sheets(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    joined = "|".join(wb.sheetnames)
    return sum(1 for item in TARGET_ITEMS if item in joined) >= 4


def find_project_dirs(root: Path) -> dict[str, Path]:
    base20 = next(
        (p for p in root.iterdir() if p.is_dir() and p.name.startswith("20_") and p.name != "20_materials"),
        None,
    )
    submitted_root = root
    if base20:
        submitted_root = next((p for p in base20.iterdir() if p.is_dir() and p.name.startswith("02_")), base20)
    base30 = next((p for p in root.iterdir() if p.is_dir() and p.name.startswith("30_")), root)
    return {"submitted": submitted_root, "output": base30 / "90_원천_자동검토" / "06_auto_review"}


def iter_xlsx_files(root: Path) -> Iterable[Path]:
    def onerror(exc: OSError) -> None:
        print(f"  skip unreadable directory: {exc.filename}", flush=True)

    for dirpath, dirnames, filenames in os.walk(root, onerror=onerror):
        dirnames[:] = [name for name in dirnames if "_백업_" not in name and not name.startswith("_assigned_sources")]
        for filename in filenames:
            if filename.lower().endswith(".xlsx"):
                yield Path(dirpath) / filename


def iter_all_files(root: Path) -> Iterable[Path]:
    def onerror(exc: OSError) -> None:
        print(f"  skip unreadable directory: {exc.filename}", flush=True)

    for dirpath, dirnames, filenames in os.walk(root, onerror=onerror):
        dirnames[:] = [name for name in dirnames if "_백업_" not in name and not name.startswith("_assigned_sources")]
        for filename in filenames:
            yield Path(dirpath) / filename


def iter_all_dirs(root: Path) -> Iterable[Path]:
    def onerror(exc: OSError) -> None:
        print(f"  skip unreadable directory: {exc.filename}", flush=True)

    for dirpath, dirnames, _filenames in os.walk(root, onerror=onerror):
        dirnames[:] = [name for name in dirnames if "_백업_" not in name and not name.startswith("_assigned_sources")]
        for dirname in dirnames:
            yield Path(dirpath) / dirname


def find_submitted_workbooks(
    submitted_root: Path,
    limit: int | None = None,
    name_keywords: tuple[str, ...] = DEFAULT_WORKBOOK_KEYWORDS,
) -> list[Path]:
    paths = []
    for path in iter_xlsx_files(submitted_root):
        if path.name.startswith("~$"):
            continue
        if "_백업_" in str(path):
            continue
        if any(part.startswith("_assigned_sources") for part in path.parts):
            continue
        if not name_matches_keywords(path, name_keywords):
            continue
        if workbook_has_target_sheets(path):
            paths.append(path)
            if limit and len(paths) >= limit:
                break
    return paths


def extract_institution_name(wb: Any, path: Path) -> str:
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 25)):
            for cell in row:
                if flat_text(cell.value) == "기관명":
                    for col in range(cell.column + 1, min(ws.max_column or cell.column + 5, cell.column + 8) + 1):
                        value = flat_text(ws.cell(cell.row, col).value)
                        if value:
                            return value
    return fallback_institution_name(path)


def fallback_institution_name(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"^\d+[_\.\s-]*", "", stem)
    return strip_generic_suffixes(stem)


def find_sheet(wb: Any, token: str) -> Any | None:
    return next((wb[name] for name in wb.sheetnames if token in name), None)


def collect_sheet_numbers(ws: Any, item: str) -> Counter:
    counts: Counter[str] = Counter()
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        row_text = " ".join(flat_text(v) for v in row if v is not None)
        if "작성방법" in row_text or "검증 자료" in row_text or "제출해주시기" in row_text:
            # Most standard forms put supporting-data request text below this point.
            # Keep parsing for 6-2 titles elsewhere, but numeric comparison should not
            # use these instruction sections.
            break
        for value in row:
            key = number_key(value)
            if key:
                counts[key] += 1
    return counts


def collect_6_2_titles(ws: Any) -> list[str]:
    title_hits = find_cells(ws, "제목", max_rows=35)
    title_col = next(
        (c for r, c in title_hits if any("날짜" in flat_text(ws.cell(r, cc).value) for cc in range(1, c + 1))),
        3,
    )
    data_start = next((r + 1 for r, c in title_hits if c == title_col), 1)
    titles = []
    for r in range(data_start, (ws.max_row or 0) + 1):
        row_text = " ".join(
            flat_text(ws.cell(r, c).value)
            for c in range(1, (ws.max_column or 0) + 1)
        )
        if "작성방법" in row_text or "검증 자료" in row_text or "제출해주시기" in row_text:
            break
        title = flat_text(ws.cell(r, title_col).value)
        if not title or title.startswith("제목") or "작성방법" in title or "검증 자료" in title:
            continue
        key = normalize_title(title)
        if key:
            titles.append(key)
    return titles


def parse_submitted_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=False, data_only=True)
    institution = extract_institution_name(wb, path)
    if not usable_org_key(institution):
        institution = fallback_institution_name(path)
    parsed = {
        "path": path,
        "institution": institution,
        "org_key": usable_org_key(institution) or usable_org_key(path.stem),
        "numbers": {},
        "titles_6_2": [],
    }
    for item in TARGET_ITEMS:
        ws = find_sheet(wb, item)
        if not ws:
            parsed["numbers"][item] = Counter()
            continue
        parsed["numbers"][item] = collect_sheet_numbers(ws, item)
        if item == "6-2":
            parsed["titles_6_2"] = collect_6_2_titles(ws)
    return parsed


def item_from_path(path: Path) -> str | None:
    path_text = str(path)
    for item in TARGET_ITEMS:
        if item in path_text:
            return item
    if "복리후생비" in path_text and "13." in path_text:
        return "13-1"
    return None


def path_matches_item(path: Path, item: str) -> bool:
    s = str(path)
    compact_s = compact(s)
    if item in s:
        return True
    if item == "5-1":
        return "신규채용" in compact_s
    if item == "6-2":
        return "직원채용" in compact_s or "채용정보" in compact_s
    if item == "10-1":
        return "임원연봉" in compact_s
    if item == "13-1":
        return "예산상복리후생" in compact_s or ("13." in s and "복리후생" in compact_s)
    if item == "13-5":
        return "1인당복리후생" in compact_s or ("13." in s and "복리후생" in compact_s)
    return False


def index_item_folders(submitted_root: Path) -> dict[str, list[Path]]:
    folders: dict[str, list[Path]] = {item: [] for item in TARGET_ITEMS}
    for path in iter_all_dirs(submitted_root):
        if "_백업_" in str(path) or any(part.startswith("_assigned_sources") for part in path.parts):
            continue
        for item in TARGET_ITEMS:
            if path_matches_item(path, item):
                folders[item].append(path)
    return folders


def candidate_evidence_dirs(
    parsed: dict[str, Any],
    item: str,
    folder_index: dict[str, list[tuple[Path, str]]],
    submitted_root: Path,
) -> list[Path]:
    org_key = parsed["org_key"]
    workbook_path = parsed["path"]
    dirs = []

    # Local item folders near nested submitted workbooks are the strongest signal.
    # Avoid scanning the whole submitted root for a root-level workbook. That
    # creates cross-institution false matches such as one agency picking up
    # another agency's 6-2 evidence folder.
    for base in [workbook_path.parent, workbook_path.parent.parent]:
        if (
            base.exists()
            and base != workbook_path.anchor
            and base.resolve() != submitted_root.resolve()
            and (org_key_in_path(org_key, base) or org_key_in_path(org_key, base.parent))
        ):
            try:
                children = list(base.iterdir())
            except OSError as exc:
                print(f"  skip unreadable directory: {exc.filename}", flush=True)
                continue
            for child in children:
                if child.is_dir() and path_matches_item(child, item):
                    dirs.append(child)

    for folder, norm_path in folder_index.get(item, []):
        if org_key and org_key in norm_path:
            dirs.append(folder)

    seen = set()
    unique = []
    for d in dirs:
        key = str(d.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(d)
    return unique


def evidence_files_for_folder(folder: Path, folder_file_cache: dict[Path, list[Path]]) -> list[Path]:
    if folder in folder_file_cache:
        return folder_file_cache[folder]
    files = []
    for path in iter_all_files(folder):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() in EVIDENCE_EXTS:
            files.append(path)
    folder_file_cache[folder] = files
    return files


def evidence_files_for_dirs(dirs: list[Path], submitted_path: Path, folder_file_cache: dict[Path, list[Path]]) -> list[Path]:
    files = []
    submitted_resolved = submitted_path.resolve()
    for folder in dirs:
        for path in evidence_files_for_folder(folder, folder_file_cache):
            if path.resolve() != submitted_resolved:
                files.append(path)
    # De-duplicate nested duplicate discoveries.
    seen = set()
    unique = []
    for path in files:
        key = str(path.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def collect_xlsx_numbers(path: Path, max_cells: int = DEFAULT_MAX_CELLS_PER_EVIDENCE) -> Counter:
    counts: Counter[str] = Counter()
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return counts
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return counts
    visited = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                visited += 1
                if visited > max_cells:
                    return counts
                key = number_key(value)
                if key:
                    counts[key] += 1
    return counts


def collect_xlsx_text(path: Path, max_chars: int = 20000) -> str:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return ""
    pieces = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is not None:
                    pieces.append(flat_text(value))
                    if sum(len(p) for p in pieces) > max_chars:
                        return " ".join(pieces)
    return " ".join(pieces)


def compare_counters(submitted: Counter, evidence: Counter) -> tuple[int, int, float, list[str]]:
    keys = list(submitted.keys())
    total = len(keys)
    matched = sum(1 for key in keys if evidence.get(key, 0) > 0)
    coverage = matched / total if total else 1.0
    unmatched = [key for key in keys if evidence.get(key, 0) == 0]
    return total, matched, coverage, unmatched


def file_counts(files: list[Path]) -> Counter:
    return Counter(path.suffix.lower() or "(no_ext)" for path in files)


def match_title_coverage(titles: list[str], files: list[Path], scan_excel_text: bool = False) -> tuple[int, int, float]:
    if not titles:
        return 0, 0, 1.0
    haystack_parts = []
    for path in files:
        haystack_parts.append(normalize_title(path.name))
    # Excel evidence sometimes has the exact title in cells, but scanning every
    # workbook is expensive. Keep it optional for targeted reruns.
    if scan_excel_text:
        for path in files:
            if path.suffix.lower() in {".xlsx", ".xlsm"}:
                haystack_parts.append(normalize_title(collect_xlsx_text(path, max_chars=12000)))
    haystack = " ".join(haystack_parts)
    matched = sum(1 for title in set(titles) if title and title in haystack)
    total = len(set(titles))
    return total, matched, matched / total if total else 1.0


def status_for(files: list[Path], numeric_total: int, numeric_coverage: float, title_total: int, title_coverage: float) -> str:
    if not files:
        return "근거자료 없음"
    if numeric_total and numeric_coverage < 0.35:
        return "숫자 매칭 낮음"
    if title_total and title_coverage < 0.35:
        return "6-2 제목 매칭 낮음"
    if numeric_total and numeric_coverage < 0.7:
        return "부분 매칭"
    if title_total and title_coverage < 0.7:
        return "부분 매칭"
    return "매칭 양호"


def build_match_rows(
    submitted: list[dict[str, Any]],
    submitted_root: Path,
    max_evidence_excels: int = DEFAULT_MAX_EVIDENCE_EXCELS,
    max_cells_per_evidence: int = DEFAULT_MAX_CELLS_PER_EVIDENCE,
    scan_excel_text_for_titles: bool = False,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    raw_folder_index = index_item_folders(submitted_root)
    folder_index = {
        item: [(folder, normalize_org_name(str(folder))) for folder in folders]
        for item, folders in raw_folder_index.items()
    }
    summary_rows = []
    file_rows = []
    unmatched_rows = []
    number_cache: dict[Path, Counter] = {}
    folder_file_cache: dict[Path, list[Path]] = {}

    for parsed_index, parsed in enumerate(submitted, start=1):
        if parsed_index == 1 or parsed_index % 25 == 0 or parsed_index == len(submitted):
            print(f"[match {parsed_index}/{len(submitted)}] {parsed['institution']}", flush=True)
        for item in TARGET_ITEMS:
            dirs = candidate_evidence_dirs(parsed, item, folder_index, submitted_root)
            files = evidence_files_for_dirs(dirs, parsed["path"], folder_file_cache)
            counts = file_counts(files)
            evidence_numbers: Counter[str] = Counter()
            excel_files = [p for p in files if p.suffix.lower() in {".xlsx", ".xlsm"}]
            excel_files = sorted(excel_files, key=lambda p: (p.stat().st_size if p.exists() else 0, str(p)))[:max_evidence_excels]
            for path in excel_files:
                if path not in number_cache:
                    number_cache[path] = collect_xlsx_numbers(path, max_cells=max_cells_per_evidence)
                evidence_numbers.update(number_cache[path])

            numeric_total, numeric_matched, numeric_coverage, unmatched = compare_counters(
                parsed["numbers"].get(item, Counter()),
                evidence_numbers,
            )
            title_total = title_matched = 0
            title_coverage = 1.0
            if item == "6-2":
                title_total, title_matched, title_coverage = match_title_coverage(
                    parsed["titles_6_2"],
                    files,
                    scan_excel_text=scan_excel_text_for_titles,
                )

            status = status_for(files, numeric_total, numeric_coverage, title_total, title_coverage)
            summary_rows.append({
                "status": status,
                "institution": parsed["institution"],
                "item": item,
                "submitted_file": str(parsed["path"]),
                "evidence_folder_count": len(dirs),
                "evidence_file_count": len(files),
                "xlsx": counts.get(".xlsx", 0) + counts.get(".xlsm", 0),
                "xls": counts.get(".xls", 0),
                "pdf": counts.get(".pdf", 0),
                "hwp_hwpx": counts.get(".hwp", 0) + counts.get(".hwpx", 0),
                "zip": counts.get(".zip", 0),
                "numeric_submitted_distinct": numeric_total,
                "numeric_matched_distinct": numeric_matched,
                "numeric_coverage": round(numeric_coverage, 4),
                "numeric_excel_scanned": len(excel_files),
                "title_submitted_distinct": title_total,
                "title_matched_distinct": title_matched,
                "title_coverage": round(title_coverage, 4),
                "unmatched_values_sample": ", ".join(unmatched[:MAX_UNMATCHED_VALUES_PER_ROW]),
                "evidence_dirs_sample": "\n".join(str(d) for d in dirs[:5]),
                "evidence_files_sample": "\n".join(str(f) for f in files[:8]),
            })

            for value in unmatched[:MAX_UNMATCHED_VALUES_PER_ROW]:
                unmatched_rows.append({
                    "institution": parsed["institution"],
                    "item": item,
                    "submitted_file": str(parsed["path"]),
                    "unmatched_numeric_value": value,
                })

            for path in files:
                file_rows.append({
                    "institution": parsed["institution"],
                    "item": item,
                    "suffix": path.suffix.lower(),
                    "file": str(path),
                })

    return summary_rows, file_rows, unmatched_rows


def write_output(output_path: Path, summary_rows: list[dict[str, Any]], file_rows: list[dict[str, Any]], unmatched_rows: list[dict[str, Any]], submitted_count: int) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "summary"
    ws.append(["section", "key", "count_or_value"])
    ws.append(["run", "generated_at", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["input", "submitted_workbooks", submitted_count])
    ws.append(["input", "institution_item_rows", len(summary_rows)])
    for status, count in Counter(row["status"] for row in summary_rows).most_common():
        ws.append(["status", status, count])
    for item, count in Counter(row["item"] for row in summary_rows if row["status"] != "매칭 양호").most_common():
        ws.append(["non_ok_by_item", item, count])

    ws = wb.create_sheet("evidence_match")
    headers = [
        "status", "institution", "item", "submitted_file",
        "evidence_folder_count", "evidence_file_count",
        "xlsx", "xls", "pdf", "hwp_hwpx", "zip",
        "numeric_submitted_distinct", "numeric_matched_distinct", "numeric_coverage", "numeric_excel_scanned",
        "title_submitted_distinct", "title_matched_distinct", "title_coverage",
        "unmatched_values_sample", "evidence_dirs_sample", "evidence_files_sample",
    ]
    ws.append(headers)
    for row in summary_rows:
        ws.append([row.get(h, "") for h in headers])

    ws = wb.create_sheet("evidence_files")
    headers = ["institution", "item", "suffix", "file"]
    ws.append(headers)
    for row in file_rows:
        ws.append([row.get(h, "") for h in headers])

    ws = wb.create_sheet("unmatched_numeric_values")
    headers = ["institution", "item", "submitted_file", "unmatched_numeric_value"]
    ws.append(headers)
    for row in unmatched_rows:
        ws.append([row.get(h, "") for h in headers])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            width = 18
            if sheet.title == "evidence_match" and col in {4, 18, 19, 20}:
                width = 55
            elif sheet.title == "evidence_files" and col == 4:
                width = 80
            elif sheet.title == "unmatched_numeric_values" and col == 3:
                width = 70
            sheet.column_dimensions[get_column_letter(col)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)


def parse_args(argv: Iterable[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Match submitted workbooks to evidence files.")
    parser.add_argument("--root", default=".", help="Project root. Defaults to current directory.")
    parser.add_argument("--submitted-root", default=None, help="Root containing institution submitted materials.")
    parser.add_argument(
        "--submitted-name-keyword",
        action="append",
        default=None,
        help="Submitted workbook filename keyword. Repeat or comma-separate. Defaults to generic and Korean legacy keywords.",
    )
    parser.add_argument(
        "--disable-submitted-name-filter",
        action="store_true",
        help="Find submitted workbooks by sheet structure only, without filename keywords.",
    )
    parser.add_argument("--output", default=None, help="Output xlsx path.")
    parser.add_argument("--limit", type=int, default=None, help="Process only first N submitted standard workbooks.")
    parser.add_argument("--max-evidence-excels", type=int, default=DEFAULT_MAX_EVIDENCE_EXCELS, help="Max evidence xlsx/xlsm files scanned per institution/item.")
    parser.add_argument("--max-cells-per-evidence", type=int, default=DEFAULT_MAX_CELLS_PER_EVIDENCE, help="Max cells scanned per evidence workbook.")
    parser.add_argument("--scan-excel-text-for-titles", action="store_true", help="Also scan evidence Excel cell text when matching 6-2 titles. Slower.")
    return parser.parse_args(list(argv))


def main(argv: Iterable[str] = sys.argv[1:]) -> int:
    args = parse_args(argv)
    root = Path(args.root).resolve()
    dirs = find_project_dirs(root)
    submitted_root = Path(args.submitted_root).resolve() if args.submitted_root else dirs["submitted"]
    if not submitted_root.exists():
        raise SystemExit("Could not find submitted root. Pass --submitted-root.")

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(args.output).resolve() if args.output else dirs["output"] / f"evidence_match_review_{stamp}.xlsx"

    name_keywords = () if args.disable_submitted_name_filter else parse_keyword_args(args.submitted_name_keyword)
    submitted_paths = find_submitted_workbooks(submitted_root, args.limit, name_keywords)
    submitted = []
    seen_org_keys: set[str] = set()
    for i, path in enumerate(submitted_paths, start=1):
        print(f"[{i}/{len(submitted_paths)}] {path}", flush=True)
        try:
            parsed = parse_submitted_workbook(path)
        except Exception as exc:
            print(f"  parse failed: {exc!r}", flush=True)
            continue
        if parsed["org_key"] in seen_org_keys:
            continue
        seen_org_keys.add(parsed["org_key"])
        submitted.append(parsed)

    summary_rows, file_rows, unmatched_rows = build_match_rows(
        submitted,
        submitted_root,
        max_evidence_excels=args.max_evidence_excels,
        max_cells_per_evidence=args.max_cells_per_evidence,
        scan_excel_text_for_titles=args.scan_excel_text_for_titles,
    )
    write_output(output_path, summary_rows, file_rows, unmatched_rows, len(submitted))
    print(f"Done: {output_path}")
    print(f"Submitted workbooks: {len(submitted)}")
    print(f"Institution/item rows: {len(summary_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
