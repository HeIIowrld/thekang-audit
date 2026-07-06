# -*- coding: utf-8 -*-
"""
Automated first-pass review for public disclosure check data.

What this script does:
- Finds public disclosure export workbooks in the current project.
- Finds institution-submitted standard review workbooks.
- Compares 5-1, 6-2, 10-1, 13-1, and 13-5 where the structure is parseable.
- Runs cross-checks inspired by the 2024 check-type workbook:
  * public disclosure values vs submitted values.
  * 10-1 salary welfare amount vs 13-1 officer salary welfare amount.
  * 13-1 budget welfare amount vs 13-5 budget total amount.
  * 6-2 job posting list matching by title/date.
  * Evidence-file inventory for each institution/item folder.

Usage:
    python .\\90_tools\\disclosure_auto_review.py --root .

Optional:
    python .\\90_tools\\disclosure_auto_review.py --root . --limit 20
    python .\\90_tools\\disclosure_auto_review.py --output .\\30_검토산출물\\06_자동검토\\review.xlsx

Dependencies:
    python -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import datetime as dt
import math
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Install it with: python -m pip install openpyxl"
    ) from exc


TARGET_ITEMS = ("5-1", "6-2", "10-1", "13-1", "13-5")
NUMERIC_TOLERANCE = 0.0001
DEFAULT_WORKBOOK_KEYWORDS = ("review", "disclosure", "checklist", "노무", "통합공시", "점검표")
DEFAULT_STEM_STRIP_PATTERNS = (
    r"_?노무.*$",
    r"_?통합공시.*$",
    r"_?점검표.*$",
    r"_?disclosure.*$",
    r"_?review.*$",
    r"_?checklist.*$",
    r"20\d{2}년도.*$",
)


@dataclass
class Finding:
    severity: str
    item: str
    institution: str
    check_type: str
    key: str
    submitted_value: Any = ""
    alio_value: Any = ""
    difference: Any = ""
    source_file: str = ""
    source_cell: str = ""
    note: str = ""


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def flat_text(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value).replace("\n", " ")).strip()


def parse_keyword_args(values: list[str] | None) -> tuple[str, ...]:
    if values is None:
        return DEFAULT_WORKBOOK_KEYWORDS
    keywords: list[str] = []
    for value in values:
        keywords.extend(token.strip() for token in value.split(",") if token.strip())
    return tuple(keywords)


def name_matches_keywords(path: Path, keywords: tuple[str, ...]) -> bool:
    if not keywords:
        return True
    name = path.name.lower()
    return any(keyword.lower() in name for keyword in keywords)


def strip_generic_suffixes(stem: str) -> str:
    for pattern in DEFAULT_STEM_STRIP_PATTERNS:
        stem = re.sub(pattern, "", stem, flags=re.IGNORECASE)
    return stem.strip(" _-")


def norm_space(value: Any) -> str:
    return re.sub(r"\s+", "", text(value))


def normalize_org_name(value: Any) -> str:
    s = norm_space(value)
    s = s.replace("주식회사", "").replace("(주)", "").replace("㈜", "")
    s = s.replace("(재)", "").replace("재단법인", "")
    s = s.replace("(사)", "").replace("사단법인", "")
    s = re.sub(r"[\(\)\[\]{}·ㆍ\.,_-]", "", s)
    return s


def normalize_key(value: Any) -> str:
    s = norm_space(value)
    s = s.replace("(", "").replace(")", "")
    s = s.replace("복리후생비", "")
    s = s.replace("정규직일반정규직", "일반정규직")
    s = s.replace("정규직무기계약직", "무기계약직")
    return s


def normalize_title(value: Any) -> str:
    s = flat_text(value).lower()
    s = re.sub(r"^\(수정\)\s*", "", s)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[\[\]\(\){}·ㆍ,._\-]", "", s)
    return s


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    s = flat_text(value)
    if not s or s in {"-", "－", "해당없음", "해당 없음"}:
        return None
    s = s.replace(",", "").replace("천원", "").replace("명", "").replace("%", "")
    s = re.sub(r"^\(([-+]?\d+(?:\.\d+)?)\)$", r"-\1", s)
    try:
        return float(s)
    except ValueError:
        return None


def value_equal(left: Any, right: Any) -> bool:
    if flat_text(left) == "" and parse_number(right) == 0:
        return True
    if flat_text(right) == "" and parse_number(left) == 0:
        return True
    ln = parse_number(left)
    rn = parse_number(right)
    if ln is not None and rn is not None:
        return abs(ln - rn) <= NUMERIC_TOLERANCE
    return flat_text(left) == flat_text(right)


def value_diff(left: Any, right: Any) -> str:
    ln = parse_number(left)
    rn = parse_number(right)
    if ln is not None and rn is not None:
        diff = ln - rn
        if abs(diff - round(diff)) <= NUMERIC_TOLERANCE:
            return str(int(round(diff)))
        return f"{diff:.4f}"
    return ""


def cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def find_cells(ws: Any, needle: str, max_rows: int = 80) -> list[tuple[int, int]]:
    hits = []
    for r in range(1, min(ws.max_row or 0, max_rows) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            if needle in flat_text(ws.cell(r, c).value):
                hits.append((r, c))
    return hits


def rows_as_dicts(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    row_iter = ws.iter_rows(values_only=True)
    try:
        headers = [flat_text(v) for v in next(row_iter)]
    except StopIteration:
        return []
    records = []
    for row in row_iter:
        if not any(v is not None and text(v) != "" for v in row):
            continue
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return records


def alio_dir_score(path: Path) -> tuple[int, int, float]:
    names = [file.name for file in path.glob("*.xlsx")]
    required_hits = sum(
        1
        for prefix in ("5-1", "6-2", "10", "13-1", "13-5")
        if any(name.startswith(prefix) for name in names)
    )
    name_score = 0
    if "공시점검" in path.name:
        name_score += 10
    if "ALIO" in path.name.upper() or "알리오" in path.name:
        name_score += 5
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0.0
    return required_hits, name_score, mtime


def find_project_dirs(root: Path) -> dict[str, Path]:
    alio_dirs = [
        p for p in root.iterdir()
        if p.is_dir() and len(list(p.glob("*.xlsx"))) >= 5
    ]
    alio_dir = max(alio_dirs, key=alio_dir_score) if alio_dirs else None

    base20 = next(
        (p for p in root.iterdir() if p.is_dir() and p.name.startswith("20_") and p.name != "20_materials"),
        None,
    )
    submitted_root = None
    if base20:
        submitted_root = next((p for p in base20.iterdir() if p.is_dir() and p.name.startswith("02_")), None)
    if not submitted_root:
        submitted_root = root

    base30 = next((p for p in root.iterdir() if p.is_dir() and p.name.startswith("30_")), root)
    output_root = base30 / "90_원천_자동검토" / "06_auto_review"
    return {
        "alio": alio_dir,
        "submitted": submitted_root,
        "output": output_root,
    }


def find_alio_files(alio_dir: Path) -> dict[str, Path]:
    files = {}
    for path in alio_dir.glob("*.xlsx"):
        name = path.name
        if name.startswith("5-1"):
            files["5-1"] = path
        elif name.startswith("6-2") and "최초" not in name:
            files["6-2-reg"] = path
        elif name.startswith("6-2") and "최초" in name:
            files["6-2-first"] = path
        elif name.startswith("10"):
            files["10-1"] = path
        elif name.startswith("13-1"):
            files["13-1"] = path
        elif name.startswith("13-5"):
            files["13-5"] = path
    return files


def load_alio_indexes(alio_files: dict[str, Path]) -> dict[str, dict[tuple, Any]]:
    indexes: dict[str, dict[tuple, Any]] = {
        "5-1": {},
        "6-2": {},
        "10-1": {},
        "13-1": {},
        "13-5": {},
    }

    if "5-1" in alio_files:
        for row in rows_as_dicts(alio_files["5-1"]):
            org = normalize_org_name(row.get("기관명"))
            cls = normalize_key(row.get("분류"))
            indexes["5-1"][(org, cls, "2025total")] = row.get("2025년")
            indexes["5-1"][(org, cls, "2026q1")] = row.get("2026년 1/4분기")

    for key in ("6-2-reg", "6-2-first"):
        if key not in alio_files:
            continue
        for row in rows_as_dicts(alio_files[key]):
            org = normalize_org_name(row.get("기관명"))
            title = normalize_title(row.get("제목"))
            if not org or not title:
                continue
            record = {
                "기관명": row.get("기관명"),
                "제목": row.get("제목"),
                "최초공시일": row.get("최초공시일"),
                "등록일": row.get("등록일"),
                "마감일": row.get("마감일"),
                "source": key,
            }
            indexes["6-2"][(org, title)] = record

    if "10-1" in alio_files:
        for row in rows_as_dicts(alio_files["10-1"]):
            org = normalize_org_name(row.get("기관명"))
            role = normalize_key(row.get("구분"))
            category = normalize_key(row.get("분류"))
            indexes["10-1"][(org, role, category, "2025")] = row.get("2025년결산")
            indexes["10-1"][(org, role, category, "2026")] = row.get("2026년예산")

    if "13-1" in alio_files:
        for row in rows_as_dicts(alio_files["13-1"]):
            org = normalize_org_name(row.get("기관명"))
            employment = normalize_key(row.get("고용형태"))
            item = normalize_key(row.get("항목"))
            indexes["13-1"][(org, employment, item, "2025")] = row.get("2025년")

    if "13-5" in alio_files:
        for row in rows_as_dicts(alio_files["13-5"]):
            org = normalize_org_name(row.get("기관명"))
            employment = normalize_key(row.get("고용형태"))
            benefit = normalize_key(row.get("구분"))
            source = normalize_key(row.get("항목"))
            indexes["13-5"][(org, employment, benefit, source, "2025")] = row.get("2025년")

    return indexes


def workbook_has_target_sheets(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    sheet_joined = "|".join(wb.sheetnames)
    return sum(1 for item in TARGET_ITEMS if item in sheet_joined) >= 4


def iter_xlsx_files(root: Path) -> Iterable[Path]:
    def onerror(exc: OSError) -> None:
        print(f"  skip unreadable directory: {exc.filename}", flush=True)

    for dirpath, dirnames, filenames in os.walk(root, onerror=onerror):
        dirnames[:] = [name for name in dirnames if "_백업_" not in name and not name.startswith("_assigned_sources")]
        for filename in filenames:
            if filename.lower().endswith(".xlsx"):
                yield Path(dirpath) / filename


def iter_all_files(root: Path) -> Iterable[Path]:
    def onerror(exc: OSError) -> None:
        print(f"  skip unreadable directory: {exc.filename}", flush=True)

    for dirpath, dirnames, filenames in os.walk(root, onerror=onerror):
        dirnames[:] = [name for name in dirnames if "_백업_" not in name and not name.startswith("_assigned_sources")]
        for filename in filenames:
            yield Path(dirpath) / filename


def find_submitted_workbooks(
    submitted_root: Path,
    limit: int | None = None,
    name_keywords: tuple[str, ...] = DEFAULT_WORKBOOK_KEYWORDS,
) -> list[Path]:
    candidates = []
    for path in iter_xlsx_files(submitted_root):
        if path.name.startswith("~$"):
            continue
        if "_백업_" in str(path):
            continue
        if any(part.startswith("_assigned_sources") for part in path.parts):
            continue
        if not name_matches_keywords(path, name_keywords):
            continue
        if workbook_has_target_sheets(path):
            candidates.append(path)
            if limit and len(candidates) >= limit:
                break
    return candidates


def find_sheet(wb: Any, token: str) -> Any | None:
    return next((wb[name] for name in wb.sheetnames if token in name), None)


def extract_institution_name(wb: Any, path: Path) -> str:
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 20)):
            for cell in row:
                if flat_text(cell.value) == "기관명":
                    for col in range(cell.column + 1, min(ws.max_column or cell.column + 5, cell.column + 8) + 1):
                        value = flat_text(ws.cell(cell.row, col).value)
                        if value:
                            return value
    stem = path.stem
    stem = re.sub(r"^\d+[_\.\s-]*", "", stem)
    return strip_generic_suffixes(stem)


def read_submitted_5_1(ws: Any) -> dict[tuple, tuple[Any, str]]:
    records: dict[tuple, tuple[Any, str]] = {}
    for r in range(1, (ws.max_row or 0) + 1):
        values = [flat_text(ws.cell(r, c).value) for c in range(1, (ws.max_column or 0) + 1)]
        if "구분" not in values:
            continue
        label_col = values.index("구분") + 1
        col_2025 = next((i + 1 for i, v in enumerate(values) if "2025년 계" in v), None)
        col_2026 = next((i + 1 for i, v in enumerate(values) if "2026년 1분기" in v), None)
        if not col_2025 and not col_2026:
            continue
        for rr in range(r + 1, min((ws.max_row or 0) + 1, r + 14)):
            label = flat_text(ws.cell(rr, label_col).value)
            if not label or label.startswith("※") or "전일제" in label:
                continue
            has_numeric = False
            if col_2025 and parse_number(ws.cell(rr, col_2025).value) is not None:
                has_numeric = True
            if col_2026 and parse_number(ws.cell(rr, col_2026).value) is not None:
                has_numeric = True
            if not has_numeric:
                continue
            cls = normalize_key(label)
            if col_2025:
                records[(cls, "2025total")] = (ws.cell(rr, col_2025).value, cell_addr(rr, col_2025))
            if col_2026:
                records[(cls, "2026q1")] = (ws.cell(rr, col_2026).value, cell_addr(rr, col_2026))
    return records


def role_year_columns_10_1(ws: Any) -> dict[tuple[str, str], int]:
    role_row = None
    for r in range(1, min(ws.max_row or 1, 25) + 1):
        row_text = "|".join(flat_text(ws.cell(r, c).value) for c in range(1, (ws.max_column or 0) + 1))
        if "상임기관장" in row_text and "상임감사" in row_text:
            role_row = r
            break
    if role_row is None:
        return {
            ("상임기관장", "2025"): 2,
            ("상임기관장", "2026"): 3,
            ("상임감사", "2025"): 4,
            ("상임감사", "2026"): 6,
            ("상임이사", "2025"): 7,
            ("상임이사", "2026"): 9,
        }

    role_starts = []
    for c in range(1, (ws.max_column or 0) + 1):
        v = normalize_key(ws.cell(role_row, c).value)
        if v in {"상임기관장", "상임감사", "상임이사"}:
            role_starts.append((v, c))
    if not role_starts:
        return {}

    columns: dict[tuple[str, str], int] = {}
    for i, (role, start_col) in enumerate(role_starts):
        end_col = role_starts[i + 1][1] - 1 if i + 1 < len(role_starts) else min(start_col + 4, ws.max_column or start_col)
        for c in range(start_col, end_col + 1):
            label = flat_text(ws.cell(role_row + 1, c).value)
            if "2025" in label:
                columns[(role, "2025")] = c
            elif "2026" in label:
                columns[(role, "2026")] = c
        if (role, "2025") not in columns:
            columns[(role, "2025")] = start_col
        if (role, "2026") not in columns:
            columns[(role, "2026")] = min(start_col + 1, end_col)
    return columns


def read_submitted_10_1(ws: Any) -> dict[tuple, tuple[Any, str]]:
    columns = role_year_columns_10_1(ws)
    records: dict[tuple, tuple[Any, str]] = {}
    for r in range(1, (ws.max_row or 0) + 1):
        category = ""
        for c in range(1, min(ws.max_column or 0, 6) + 1):
            candidate = normalize_key(ws.cell(r, c).value)
            if candidate in {
                "기본급", "고정수당", "실적수당", "급여성", "성과상여금",
                "경영평가성과급", "기타", "합계", "비고",
            }:
                category = candidate
                break
        if category not in {
            "기본급", "고정수당", "실적수당", "급여성", "성과상여금",
            "경영평가성과급", "기타", "합계", "비고",
        }:
            continue
        for (role, year), col in columns.items():
            records[(role, category, year)] = (ws.cell(r, col).value, cell_addr(r, col))
    return records


def benefit_prefix(value: Any) -> str:
    s = flat_text(value)
    if "비급여성" in s:
        return "비급여성"
    if "급여성" in s:
        return "급여성"
    return normalize_key(s)


def read_submitted_13_1(ws: Any) -> dict[tuple, tuple[Any, str]]:
    records: dict[tuple, tuple[Any, str]] = {}
    header_hits = find_cells(ws, "항목")
    header_row = next((r for r, c in header_hits if any("구분" in flat_text(ws.cell(r, cc).value) for cc in range(1, (ws.max_column or 0) + 1))), 12)
    item_col = next((c for r, c in header_hits if r == header_row), 3)
    group_col = max(1, item_col - 1)
    employment_cols: dict[int, str] = {}
    for c in range(item_col + 1, min(ws.max_column or item_col + 6, item_col + 8) + 1):
        label = flat_text(ws.cell(header_row + 1, c).value).replace("*", "")
        if label in {"임원", "일반정규직", "무기계약직", "비정규직"}:
            employment_cols[c] = label
    if not employment_cols:
        employment_cols = {
            item_col + 1: "임원",
            item_col + 2: "일반정규직",
            item_col + 3: "무기계약직",
            item_col + 4: "비정규직",
        }
    current_group = ""
    for r in range(1, (ws.max_row or 0) + 1):
        group = flat_text(ws.cell(r, group_col).value)
        item = flat_text(ws.cell(r, item_col).value)
        if group:
            current_group = benefit_prefix(group)
        if not item or item in {"항목", "구분"}:
            continue
        item_key = normalize_key(f"{current_group}>{item}")
        for col, employment in employment_cols.items():
            records[(normalize_key(employment), item_key, "2025")] = (
                ws.cell(r, col).value,
                cell_addr(r, col),
            )
    return records


def read_submitted_13_5(ws: Any) -> dict[tuple, tuple[Any, str]]:
    records: dict[tuple, tuple[Any, str]] = {}
    header_hits = find_cells(ws, "항목")
    header_row = next((r for r, c in header_hits if any("구분" in flat_text(ws.cell(r, cc).value) for cc in range(1, (ws.max_column or 0) + 1))), 11)
    benefit_col = next((c for r, c in header_hits if r == header_row), 3)
    employment_col = max(1, benefit_col - 1)
    source_cols = {
        ("total_per", "총1인당"): benefit_col + 1,
        ("budget_total", "예산총합"): benefit_col + 2,
        ("budget_per", "예산"): benefit_col + 3,
        ("fund_total", "사내근로복지기금총합"): benefit_col + 4,
        ("fund_per", "사내근로복지기금"): benefit_col + 5,
        ("other_total", "기타총합"): benefit_col + 6,
        ("other_per", "기타"): benefit_col + 7,
    }
    current_employment = ""
    for r in range(1, (ws.max_row or 0) + 1):
        employment = flat_text(ws.cell(r, employment_col).value)
        benefit = flat_text(ws.cell(r, benefit_col).value)
        if employment:
            current_employment = employment
        if not current_employment or not benefit or benefit in {"항목", "구분"}:
            continue
        for (_kind, source), col in source_cols.items():
            records[(normalize_key(current_employment), normalize_key(benefit), normalize_key(source), "2025")] = (
                ws.cell(r, col).value,
                cell_addr(r, col),
            )
    return records


def read_submitted_6_2(ws: Any) -> dict[tuple, tuple[Any, str]]:
    records: dict[tuple, tuple[Any, str]] = {}
    title_hits = find_cells(ws, "제목", max_rows=30)
    title_col = next((c for r, c in title_hits if any("날짜" in flat_text(ws.cell(r, cc).value) for cc in range(1, c + 1))), 3)
    date_col = max(1, title_col - 1)
    data_start = next((r + 1 for r, c in title_hits if c == title_col), 1)
    for r in range(data_start, (ws.max_row or 0) + 1):
        row_text = " ".join(
            flat_text(ws.cell(r, c).value)
            for c in range(1, (ws.max_column or 0) + 1)
        )
        if "작성방법" in row_text or "검증 자료" in row_text or "제출해주시기" in row_text:
            break
        date_value = ws.cell(r, date_col).value
        title = flat_text(ws.cell(r, title_col).value)
        if not title or title.startswith("제목") or "작성방법" in title:
            continue
        title_key = normalize_title(title)
        if not title_key:
            continue
        source_cell = f"날짜 {cell_addr(r, date_col)} / 제목 {cell_addr(r, title_col)}"
        records[(title_key,)] = (date_value, source_cell)
    return records


def alio_13_1_item_key(submitted_item: str) -> str:
    return submitted_item.split(">", 1)[1] if ">" in submitted_item else submitted_item


def append_alio_missing(
    findings: list[Finding],
    item: str,
    org_name: str,
    key: str,
    value: Any,
    source_file: str,
    addr: str,
    note: str = "ALIO 키 미매칭 또는 원천 누락",
) -> None:
    findings.append(
        Finding(
            "확인",
            item,
            org_name,
            "기관/항목 ALIO 매칭 필요",
            key,
            value,
            "",
            "",
            source_file,
            addr,
            note,
        )
    )


def read_submitted_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=False, data_only=True)
    org_name = extract_institution_name(wb, path)
    parsed = {
        "path": path,
        "institution": org_name,
        "org_key": normalize_org_name(org_name),
        "5-1": {},
        "6-2": {},
        "10-1": {},
        "13-1": {},
        "13-5": {},
    }
    ws = find_sheet(wb, "5-1")
    if ws:
        parsed["5-1"] = read_submitted_5_1(ws)
    ws = find_sheet(wb, "6-2")
    if ws:
        parsed["6-2"] = read_submitted_6_2(ws)
    ws = find_sheet(wb, "10-1")
    if ws:
        parsed["10-1"] = read_submitted_10_1(ws)
    ws = find_sheet(wb, "13-1")
    if ws:
        parsed["13-1"] = read_submitted_13_1(ws)
    ws = find_sheet(wb, "13-5")
    if ws:
        parsed["13-5"] = read_submitted_13_5(ws)
    return parsed


def compare_submitted_to_alio(parsed: dict[str, Any], indexes: dict[str, dict[tuple, Any]]) -> list[Finding]:
    findings: list[Finding] = []
    org = parsed["org_key"]
    org_name = parsed["institution"]
    source_file = str(parsed["path"])

    for (cls, period), (value, addr) in parsed["5-1"].items():
        alio_value = indexes["5-1"].get((org, cls, period))
        if alio_value is None:
            findings.append(Finding("확인", "5-1", org_name, "기관/분류 ALIO 매칭 필요", f"{cls}/{period}", value, "", "", source_file, addr))
        elif not value_equal(value, alio_value):
            findings.append(Finding("주의", "5-1", org_name, "ALIO-제출 신규채용 인원 불일치", f"{cls}/{period}", value, alio_value, value_diff(value, alio_value), source_file, addr))

    submitted_titles = set(k[0] for k in parsed["6-2"].keys())
    alio_titles = {title for (alio_org, title), _record in indexes["6-2"].items() if alio_org == org}
    for title in sorted(alio_titles - submitted_titles)[:200]:
        record = indexes["6-2"][(org, title)]
        findings.append(Finding("확인", "6-2", org_name, "ALIO 채용공고가 제출 목록에 없음", flat_text(record.get("제목")), "", "", "", source_file, "", f"ALIO {record.get('source')}"))
    for title in sorted(submitted_titles - alio_titles)[:200]:
        value, addr = parsed["6-2"][(title,)]
        findings.append(Finding("확인", "6-2", org_name, "제출 채용공고가 ALIO 목록에 없음", title, value, "", "", source_file, addr))

    for (role, category, year), (value, addr) in parsed["10-1"].items():
        alio_value = indexes["10-1"].get((org, role, category, year))
        if alio_value is None:
            # The 10-1 "비고" cells and other nonnumeric labels are not ALIO numeric match targets.
            if category == "비고" or parse_number(value) is None:
                continue
            append_alio_missing(
                findings,
                "10-1",
                org_name,
                f"{role}/{category}/{year}",
                value,
                source_file,
                addr,
            )
            continue
        if not value_equal(value, alio_value):
            findings.append(Finding("주의", "10-1", org_name, "ALIO-제출 임원연봉 불일치", f"{role}/{category}/{year}", value, alio_value, value_diff(value, alio_value), source_file, addr))

    for (employment, item, year), (value, addr) in parsed["13-1"].items():
        alio_value = indexes["13-1"].get((org, employment, item, year))
        if alio_value is None:
            alio_value = indexes["13-1"].get((org, employment, alio_13_1_item_key(item), year))
        if alio_value is None:
            append_alio_missing(
                findings,
                "13-1",
                org_name,
                f"{employment}/{item}/{year}",
                value,
                source_file,
                addr,
            )
            continue
        if not value_equal(value, alio_value):
            findings.append(Finding("주의", "13-1", org_name, "ALIO-제출 예산상 복리후생비 불일치", f"{employment}/{item}/{year}", value, alio_value, value_diff(value, alio_value), source_file, addr))

    for (employment, benefit, source, year), (value, addr) in parsed["13-5"].items():
        if source.endswith("총합") or source == "총1인당":
            continue
        alio_value = indexes["13-5"].get((org, employment, benefit, source, year))
        if alio_value is None:
            append_alio_missing(
                findings,
                "13-5",
                org_name,
                f"{employment}/{benefit}/{source}/{year}",
                value,
                source_file,
                addr,
            )
            continue
        if not value_equal(value, alio_value):
            findings.append(Finding("주의", "13-5", org_name, "ALIO-제출 1인당 복리후생비 불일치", f"{employment}/{benefit}/{source}/{year}", value, alio_value, value_diff(value, alio_value), source_file, addr))

    return findings


def submitted_cross_checks(parsed: dict[str, Any]) -> list[Finding]:
    findings: list[Finding] = []
    org_name = parsed["institution"]
    source_file = str(parsed["path"])

    # 10-1: salary welfare for standing executives should match 13-1 officer salary welfare total.
    welfare_10 = 0.0
    welfare_cells = []
    for role in ("상임기관장", "상임감사", "상임이사"):
        record = parsed["10-1"].get((role, "급여성", "2025"))
        if record:
            num = parse_number(record[0])
            if num is not None:
                welfare_10 += num
                welfare_cells.append(record[1])
    welfare_13 = 0.0
    welfare_13_cells = []
    for (employment, item, year), (value, addr) in parsed["13-1"].items():
        if employment == "임원" and year == "2025" and item.startswith("급여성>"):
            num = parse_number(value)
            if num is not None:
                welfare_13 += num
                welfare_13_cells.append(addr)
    if welfare_cells and welfare_13_cells and abs(welfare_10 - welfare_13) > NUMERIC_TOLERANCE:
        findings.append(Finding(
            "주의", "10-1/13-1", org_name,
            "상임임원 급여성 복리후생비 합계 불일치",
            "10-1 급여성 복리후생비 합계 vs 13-1 임원 급여성 복리후생비",
            welfare_10, welfare_13, welfare_10 - welfare_13, source_file,
            ",".join(welfare_cells[:8]), "2024년 점검유형 반영",
        ))

    # 13-1 vs 13-5: budget amount should match 13-5 budget total by employment/benefit.
    for (employment, item, year), (value_13_1, addr_13_1) in parsed["13-1"].items():
        if year != "2025":
            continue
        if not item.startswith("급여성>") and not item.startswith("비급여성>"):
            continue
        benefit = item.split(">", 1)[1]
        record_13_5 = parsed["13-5"].get((employment, benefit, "예산총합", "2025"))
        if not record_13_5:
            continue
        value_13_5, addr_13_5 = record_13_5
        if not value_equal(value_13_1, value_13_5):
            findings.append(Finding(
                "주의", "13-1/13-5", org_name,
                "예산상 복리후생비와 1인당 복리후생비 예산총합 불일치",
                f"{employment}/{benefit}/2025",
                value_13_1, value_13_5, value_diff(value_13_1, value_13_5),
                source_file, f"{addr_13_1} vs {addr_13_5}", "2024년 점검유형 반영",
            ))

    return findings


def item_from_path(path: Path) -> str | None:
    for part in path.parts:
        for item in TARGET_ITEMS:
            if item in part:
                return item
    return None


def institution_folder_key(path: Path) -> str:
    for part in reversed(path.parts):
        lower = part.lower()
        if (
            re.match(r"^\d{3}\.", part)
            or "_노무" in part
            or "노무 " in part
            or "통합공시" in part
            or "disclosure" in lower
            or "review" in lower
            or "checklist" in lower
        ):
            return part
    return path.parent.name


def evidence_inventory(submitted_root: Path) -> list[dict[str, Any]]:
    counters: dict[tuple[str, str], Counter] = defaultdict(Counter)
    examples: dict[tuple[str, str], list[str]] = defaultdict(list)
    for path in iter_all_files(submitted_root):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        item = item_from_path(path)
        if not item:
            continue
        inst = institution_folder_key(path)
        key = (inst, item)
        suffix = path.suffix.lower() or "(no_ext)"
        counters[key][suffix] += 1
        if len(examples[key]) < 5:
            examples[key].append(str(path))
    rows = []
    for (inst, item), counter in sorted(counters.items()):
        rows.append({
            "institution_folder": inst,
            "item": item,
            "xlsx": counter.get(".xlsx", 0),
            "xls": counter.get(".xls", 0),
            "pdf": counter.get(".pdf", 0),
            "hwp": counter.get(".hwp", 0),
            "docx": counter.get(".docx", 0),
            "other": sum(v for k, v in counter.items() if k not in {".xlsx", ".xls", ".pdf", ".hwp", ".docx"}),
            "examples": "\n".join(examples[(inst, item)]),
        })
    return rows


def write_output(
    output_path: Path,
    findings: list[Finding],
    submitted: list[dict[str, Any]],
    inventory_rows: list[dict[str, Any]],
    alio_files: dict[str, Path],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["section", "key", "count_or_value"])
    ws.append(["run", "generated_at", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["input", "submitted_workbooks", len(submitted)])
    ws.append(["input", "alio_files", len(alio_files)])
    for item, count in Counter(f.item for f in findings).most_common():
        ws.append(["finding_by_item", item, count])
    for check_type, count in Counter(f.check_type for f in findings).most_common():
        ws.append(["finding_by_type", check_type, count])

    ws = wb.create_sheet("findings")
    headers = [
        "severity", "item", "institution", "check_type", "key",
        "submitted_value", "alio_value", "difference",
        "source_file", "source_cell", "note",
    ]
    ws.append(headers)
    for f in findings:
        ws.append([
            f.severity, f.item, f.institution, f.check_type, f.key,
            f.submitted_value, f.alio_value, f.difference,
            f.source_file, f.source_cell, f.note,
        ])

    ws = wb.create_sheet("submitted_workbooks")
    ws.append(["institution", "org_key", "path", "5-1_rows", "6-2_rows", "10-1_rows", "13-1_rows", "13-5_rows"])
    for parsed in submitted:
        ws.append([
            parsed["institution"],
            parsed["org_key"],
            str(parsed["path"]),
            len(parsed["5-1"]),
            len(parsed["6-2"]),
            len(parsed["10-1"]),
            len(parsed["13-1"]),
            len(parsed["13-5"]),
        ])

    ws = wb.create_sheet("evidence_inventory")
    inv_headers = ["institution_folder", "item", "xlsx", "xls", "pdf", "hwp", "docx", "other", "examples"]
    ws.append(inv_headers)
    for row in inventory_rows:
        ws.append([row.get(h, "") for h in inv_headers])

    ws = wb.create_sheet("alio_files")
    ws.append(["item", "path"])
    for item, path in sorted(alio_files.items()):
        ws.append([item, str(path)])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        for col in range(1, sheet.max_column + 1):
            width = 18
            if sheet.title == "findings" and col in {4, 5, 9, 11}:
                width = 42
            elif sheet.title in {"evidence_inventory", "submitted_workbooks"} and col in {3, 9}:
                width = 60
            sheet.column_dimensions[get_column_letter(col)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)


def parse_args(argv: Iterable[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Automated public-disclosure/submitted workbook review.")
    parser.add_argument("--root", default=".", help="Project root. Defaults to current directory.")
    parser.add_argument("--alio-dir", default=None, help="Directory containing public disclosure export xlsx files.")
    parser.add_argument("--submitted-root", default=None, help="Root containing institution submitted workbooks/materials.")
    parser.add_argument(
        "--submitted-name-keyword",
        action="append",
        default=None,
        help="Submitted workbook filename keyword. Repeat or comma-separate. Defaults to generic and Korean legacy keywords.",
    )
    parser.add_argument(
        "--disable-submitted-name-filter",
        action="store_true",
        help="Find submitted workbooks by sheet structure only, without filename keywords.",
    )
    parser.add_argument("--output", default=None, help="Output xlsx path.")
    parser.add_argument("--limit", type=int, default=None, help="Process only the first N submitted standard workbooks.")
    return parser.parse_args(list(argv))


def main(argv: Iterable[str] = sys.argv[1:]) -> int:
    args = parse_args(argv)
    root = Path(args.root).resolve()
    dirs = find_project_dirs(root)
    alio_dir = Path(args.alio_dir).resolve() if args.alio_dir else dirs["alio"]
    submitted_root = Path(args.submitted_root).resolve() if args.submitted_root else dirs["submitted"]
    if alio_dir is None or not alio_dir.exists():
        raise SystemExit("Could not find ALIO export directory. Pass --alio-dir.")
    if not submitted_root.exists():
        raise SystemExit("Could not find submitted root. Pass --submitted-root.")

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(args.output).resolve() if args.output else dirs["output"] / f"disclosure_auto_review_{stamp}.xlsx"

    alio_files = find_alio_files(alio_dir)
    indexes = load_alio_indexes(alio_files)
    name_keywords = () if args.disable_submitted_name_filter else parse_keyword_args(args.submitted_name_keyword)
    submitted_paths = find_submitted_workbooks(submitted_root, args.limit, name_keywords)

    submitted = []
    findings: list[Finding] = []
    seen_org_keys: set[str] = set()
    for i, path in enumerate(submitted_paths, start=1):
        print(f"[{i}/{len(submitted_paths)}] {path}", flush=True)
        try:
            parsed = read_submitted_workbook(path)
        except Exception as exc:
            findings.append(Finding("오류", "공통", path.stem, "제출 엑셀 파싱 실패", "", "", "", "", str(path), "", repr(exc)))
            continue
        if parsed["org_key"] in seen_org_keys:
            continue
        seen_org_keys.add(parsed["org_key"])
        submitted.append(parsed)
        findings.extend(compare_submitted_to_alio(parsed, indexes))
        findings.extend(submitted_cross_checks(parsed))

    inventory_rows = evidence_inventory(submitted_root)
    write_output(output_path, findings, submitted, inventory_rows, alio_files)
    print(f"Done: {output_path}")
    print(f"Submitted workbooks: {len(submitted)}")
    print(f"Findings: {len(findings)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
