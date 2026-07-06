# -*- coding: utf-8 -*-
"""Build result-table draft rows from the filtered review checklist.

The output is intentionally a draft. It maps high-priority candidates to the
result-table fields, but keeps the violation type and violation level as
reviewer-confirmed fields instead of auto-finalizing them.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SOURCE_SHEETS = [
    ("1차_필터링체크리스트", "결과표_기재후보"),
    ("자료보완_근거자료없음", "자료보완_요청후보"),
    ("2차_참고", "2차_참고후보"),
]

ITEM_MAP = {
    "5-1": ("5. 신규채용 현황\n(정기 기타)", "5-1. 신규채용 \n현황", "", "26~27"),
    "6-2": ("6. 임·직원 채용정보\n(수시)", "6-2. 직원 \n채용정보", "", "28~29"),
    "10-1": ("10. 임원 연봉\n(정기 주요)", "10-1. 임원연봉\n(상임임원)", "", "30~31"),
    "13-1": ("13. 복리후생비\n(정기 주요)", "13-1. 예산상\n복리후생내역", "", "32~33"),
    "13-5": ("13. 복리후생비\n(정기 주요)", "13-5. 1인당\n복리후생비", "", "34~35"),
    "13-1/13-5": ("13. 복리후생비\n(정기 주요)", "13-1/13-5 연계검증", "", "32~35"),
    "공통/본문": ("공통", "불성실공시 자진신고/본문확인", "", ""),
}

OUTPUT_HEADERS = [
    "상태",
    "결과표 반영 여부",
    "검토단계",
    "자동화 판정",
    "담당자",
    "기관번호",
    "기관명",
    "항목",
    "결과표 항목",
    "결과표 세항목",
    "결과표 세세항목",
    "결과표 세부내용 행",
    "불성실 유형 추정",
    "위반수준 추정",
    "불성실공시 세부내용 초안",
    "확인 필요사항",
    "대조군 값",
    "비교군 값",
    "차이/판정",
    "확인자료",
    "원천순번",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def flat(value: Any, limit: int = 400) -> str:
    value = " ".join(text(value).replace("\n", " ").split())
    return value if len(value) <= limit else value[: limit - 1] + "…"


def find_dir(parent: Path, prefix: str) -> Path:
    matches = [p for p in parent.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not matches:
        raise FileNotFoundError(f"Missing directory {prefix}* under {parent}")
    return sorted(matches, key=lambda p: p.name)[0]


def read_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    result: list[dict[str, Any]] = []
    for row in rows:
        if not any(text(v) for v in row):
            continue
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    wb.close()
    return result


def result_item_fields(item: str) -> tuple[str, str, str, str]:
    return ITEM_MAP.get(item, (item, "", "", ""))


def estimated_violation(row: dict[str, Any], source_sheet: str) -> tuple[str, str, str]:
    review_type = text(row.get("검토유형"))
    finding = text(row.get("발견 요약"))
    verdict = text(row.get("차이/판정"))

    if source_sheet == "자료보완_근거자료없음":
        return "자료보완 후 판단", "자동판정 불가", "자료보완 우선"
    if review_type == "불성실공시 키워드":
        return "자진신고/불성실공시 확인", "자동판정 불가", "원문 확인"
    if "1,000배 단위차" in verdict or "1000배 단위차" in verdict:
        return "공시오류 가능", "자동판정 불가", "단위보정 후 판단"
    if review_type in {"공시-제출 수치", "공시-제출 목록", "제출엑셀 내부"}:
        return "공시오류 가능", "자동판정 불가", "결과표 후보"
    if "근거자료 없음" in finding:
        return "자료보완 후 판단", "자동판정 불가", "자료보완 우선"
    return "검토 후 판단", "자동판정 불가", "수동확인"


def draft_detail(row: dict[str, Any], source_sheet: str) -> str:
    item = text(row.get("항목"))
    review_type = text(row.get("검토유형"))
    finding = text(row.get("발견 요약"))
    control = text(row.get("대조군 값"))
    compare = text(row.get("비교군 값"))
    verdict = text(row.get("차이/판정"))

    if source_sheet == "자료보완_근거자료없음":
        return flat(
            f"{item} 항목에 대해 기관 제출 점검표 값은 있으나 자동 탐색 기준상 근거자료 파일이 확인되지 않습니다. "
            f"{verdict} 자료 제출 누락, 압축/폴더명 오류, 별도 경로 제출 여부 확인이 필요합니다.",
            900,
        )
    if review_type == "공시-제출 목록":
        return flat(
            f"{item} 항목에서 ALIO 채용공고 목록과 기관 제출 점검표 목록이 일치하지 않습니다. "
            f"{control} / {compare}. {verdict}",
            900,
        )
    if review_type == "공시-제출 수치":
        return flat(
            f"{item} 항목에서 ALIO 공시자료와 기관 제출 점검표 수치가 불일치합니다. "
            f"{control} / {compare}. {verdict}",
            900,
        )
    if review_type == "제출엑셀 내부":
        return flat(
            f"{item} 항목 관련 제출 점검표 내부 교차검증에서 값이 불일치합니다. "
            f"{control} / {compare}. {verdict}",
            900,
        )
    if review_type == "불성실공시 키워드":
        return flat(
            f"제출자료 또는 증빙 본문에서 불성실공시 관련 키워드가 탐지되었습니다. "
            f"실제 자진신고/불성실공시 내역인지 원문 확인이 필요합니다. {verdict}",
            900,
        )
    return flat(f"{finding}. {control} / {compare}. {verdict}", 900)


def review_question(row: dict[str, Any]) -> str:
    pieces = [
        text(row.get("검토포인트")),
        text(row.get("검토 질문")),
        text(row.get("필터링 기준")),
        text(row.get("검산/보정 메모")),
    ]
    seen: list[str] = []
    for piece in pieces:
        if piece and piece not in seen:
            seen.append(piece)
    return "\n".join(seen)


def convert_rows(rows: list[dict[str, Any]], source_sheet: str) -> list[list[Any]]:
    converted: list[list[Any]] = []
    for row in rows:
        item = text(row.get("항목"))
        result_item, result_sub, result_subsub, result_rows = result_item_fields(item)
        violation, level, disposition = estimated_violation(row, source_sheet)
        converted.append(
            [
                "확인전",
                "미입력",
                text(row.get("검토단계")),
                disposition,
                text(row.get("담당자")),
                text(row.get("기관번호")),
                text(row.get("기관")),
                item,
                result_item,
                result_sub,
                result_subsub,
                result_rows,
                violation,
                level,
                draft_detail(row, source_sheet),
                review_question(row),
                text(row.get("대조군 값")),
                text(row.get("비교군 값")),
                text(row.get("차이/판정")),
                text(row.get("확인자료")),
                text(row.get("원천순번")),
            ]
        )
    return converted


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    widths = {
        "A": 12,
        "B": 16,
        "C": 12,
        "D": 18,
        "E": 10,
        "F": 10,
        "G": 28,
        "H": 12,
        "I": 24,
        "J": 24,
        "K": 18,
        "L": 14,
        "M": 20,
        "N": 18,
        "O": 70,
        "P": 48,
        "Q": 28,
        "R": 28,
        "S": 34,
        "T": 70,
        "U": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(size=9)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_dv = DataValidation(type="list", formula1='"확인전,이상없음,수정요청,판단보류,제외"', allow_blank=False)
    reflect_dv = DataValidation(type="list", formula1='"미입력,결과표반영,자료보완,설명요청,제외,이상없음"', allow_blank=False)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(reflect_dv)
    if ws.max_row >= 2:
        status_dv.add(f"A2:A{ws.max_row}")
        reflect_dv.add(f"B2:B{ws.max_row}")


def add_sheet(wb: Workbook, title: str, rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def add_summary(wb: Workbook, generated_at: str, sheet_rows: dict[str, list[list[Any]]]) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.append(["결과표 기재초안"])
    ws.append(["생성시각", generated_at])
    ws.append(["주의", "자동 확정 결과표가 아니라 검토자가 확인 후 결과표에 반영할 후보 목록입니다."])
    ws.append([])
    ws.append(["시트", "건수"])
    for title, rows in sheet_rows.items():
        ws.append([title, len(rows)])
    ws.append([])
    ws.append(["자동화 판정", "건수"])
    counter = Counter(row[3] for rows in sheet_rows.values() for row in rows)
    for key, count in counter.most_common():
        ws.append([key, count])
    ws.append([])
    ws.append(["항목", "건수"])
    item_counter = Counter(row[7] for rows in sheet_rows.values() for row in rows)
    for key, count in item_counter.most_common():
        ws.append([key, count])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_num in (1, 5, 10):
        for cell in ws[row_num]:
            if cell.value:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")


def read_coverage_counts(start_dir: Path) -> tuple[Counter[str], dict[str, Counter[str]]]:
    path = start_dir / "02_2024유형대비_필터커버리지.xlsx"
    if not path.exists():
        return Counter(), {}
    rows = read_rows(path, "유형별커버리지")
    total = Counter(text(row.get("판정")) for row in rows)
    by_item: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows:
        by_item[text(row.get("항목"))][text(row.get("판정"))] += 1
    return total, by_item


def read_text_status(root: Path) -> list[tuple[str, str]]:
    path = (
        find_dir(root, "30_")
        / "90_원천_자동검토"
        / "04_증빙본문_텍스트추출"
        / "텍스트추출_상태요약.csv"
    )
    if not path.exists():
        return []
    rows: list[tuple[str, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            rows.append((row.get("group", ""), row.get("count", "")))
    return rows


def write_system_report(
    root: Path,
    start_dir: Path,
    generated_at: str,
    sheet_rows: dict[str, list[list[Any]]],
) -> Path:
    coverage_counts, by_item = read_coverage_counts(start_dir)
    text_status = read_text_status(root)
    stage_lines = "\n".join(f"| {title} | {len(rows)} |" for title, rows in sheet_rows.items())
    disposition = Counter(row[3] for rows in sheet_rows.values() for row in rows)
    disposition_lines = "\n".join(f"| {key} | {count} |" for key, count in disposition.most_common())
    coverage_lines = "\n".join(f"| {key} | {coverage_counts.get(key, 0)} |" for key in ("직접커버", "부분커버", "미커버"))
    item_lines = "\n".join(
        f"| {item} | {counter.get('직접커버', 0)} | {counter.get('부분커버', 0)} | {counter.get('미커버', 0)} |"
        for item, counter in sorted(by_item.items())
    )
    text_lines = "\n".join(f"| {group} | {count} |" for group, count in text_status)

    report = f"""# 자동화 시스템 검토보고서

생성시각: {generated_at}

## 결론

현재 시스템은 결과표 확정 자동작성 시스템이 아니라, 사람이 볼 후보를 줄이는 `검토 큐 생성 시스템`으로 보는 것이 맞습니다.
ALIO-제출엑셀 수치/목록 대조와 제출엑셀 내부 검산은 바로 활용 가능하고, 기관별 근거자료 형식 차이는 기존 텍스트 추출 인덱스를 증빙 매칭에 연결하는 방식으로 자동화율을 더 올릴 수 있습니다.

## 결과표 작성 흐름

1. `01_관리자용_전체검토현황_필터링.xlsx`에서 후보를 확인합니다.
2. `04_결과표_기재초안.xlsx`의 `결과표_기재후보`에서 결과표 문안 초안을 검토합니다.
3. `확인자료`의 제출엑셀, ALIO, 근거자료를 열어 오류 여부를 확정합니다.
4. 확정된 건만 결과표 양식의 `#2 불성실공시 세부내용 총괄`에 옮기고, 기관 소명은 `#3 기관 소명내용 총괄`에 정리합니다.

## 현재 후보 물량

| 구분 | 건수 |
|---|---:|
{stage_lines}

## 자동화 판정

| 판정 | 건수 |
|---|---:|
{disposition_lines}

## 2024년 유형 대비 커버리지

| 판정 | 건수 |
|---|---:|
{coverage_lines}

| 항목 | 직접커버 | 부분커버 | 미커버 |
|---|---:|---:|---:|
{item_lines}

## 개발 우선순위

| 우선순위 | 개발 항목 | 기대효과 |
|---|---|---|
| 1 | 기존 PDF/HWP/엑셀 텍스트·숫자 추출 인덱스를 `evidence_match_review.py`에 결합 | 기관별 근거자료 양식 차이를 흡수하고 `근거자료 없음/숫자 매칭 낮음` 오탐 축소 |
| 2 | ALIO-제출 수치 단위 정규화 레이어 추가 | 원/천원 1,000배 단위차 후보를 2차로 밀지 않고 자동 제외 또는 보정 판정 |
| 3 | 6-2 채용공고 문서분류와 필수첨부 체크 | 2024년 미커버 유형인 근무분야, 공개일, 결과확정일, 직무기술서/입사지원서 누락 확인 자동화 |
| 4 | 5-1 개인별 로데이터 속성 파서 | 지역인재, 고졸, 청년, 직종/계약유형 판단 자동화 |
| 5 | 11-1 파서와 13-5 평균인원 연계 | 13-5 평균인원-상시종업원수 유형 검토 자동화 |

## 텍스트 추출 상태

| 상태 | 건수 |
|---|---:|
{text_lines}

## 운영상 주의

- `공시오류 가능`, `자료보완 후 판단`은 결과표 확정값이 아닙니다.
- 불성실 유형과 위반수준은 자동 추정치로 두고, 최종 벌점 판단 전에 원문과 기관 소명을 확인해야 합니다.
- 기관별 근거자료 폴더명과 압축 구조가 달라 경로 매칭 실패가 생길 수 있으므로, 근거자료 없음 후보는 먼저 제출 경로 누락 여부를 확인해야 합니다.
"""
    path = start_dir / "05_자동화시스템_검토보고서.md"
    path.write_text(report, encoding="utf-8-sig")
    return path


def build(root: Path) -> list[Path]:
    out_dir = find_dir(root, "30_")
    start_dir = out_dir / "00_00_먼저보기"
    source = start_dir / "01_관리자용_전체검토현황_필터링.xlsx"
    if not source.exists():
        raise FileNotFoundError(f"Missing filtered checklist: {source}")

    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    sheet_rows: dict[str, list[list[Any]]] = {}
    for source_sheet, target_sheet in SOURCE_SHEETS:
        sheet_rows[target_sheet] = convert_rows(read_rows(source, source_sheet), source_sheet)

    wb = Workbook()
    add_summary(wb, generated_at, sheet_rows)
    for title, rows in sheet_rows.items():
        add_sheet(wb, title, rows)

    xlsx_path = start_dir / "04_결과표_기재초안.xlsx"
    wb.save(xlsx_path)
    report_path = write_system_report(root, start_dir, generated_at, sheet_rows)
    return [xlsx_path, report_path]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    args = parser.parse_args()
    written = build(args.root.resolve())
    for path in written:
        print(path)


if __name__ == "__main__":
    main()
