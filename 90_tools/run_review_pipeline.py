# -*- coding: utf-8 -*-
r"""Run the disclosure review pipeline and leave only the lightweight package.

Final reviewer-facing output:

    30_검토산출물
      00_안내.md
      01_검토시작
        00_안내.md
        01_검토큐.xlsx
      90_실행로그
        01_파이프라인_점검.md
        90_재실행캐시

    40_전달패키지
      00_안내.md
      01_검토시작
        00_안내.md
        01_검토큐.xlsx

The detailed workbooks/CSVs are generated only as intermediate files, then
removed from the reviewer-facing output directory.
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


LOG_DIR_NAME = "90_실행로그"
CACHE_DIR_NAME = "90_재실행캐시"
AUTO_CACHE_DIR = "01_자동검토"
REFERENCE_CACHE_DIR = "02_기준입력"
LIGHT_DIR_NAME = "01_검토시작"
LIGHT_WORKBOOK = "01_검토큐.xlsx"
LIGHT_README = "00_안내.md"
ROOT_README = "00_안내.md"
AUDIT_REPORT = "01_파이프라인_점검.md"
DELIVERY_DIR_NAME = "40_전달패키지"
STALE_LIGHT_WORKBOOKS = {"01_수동검사_리스트.xlsx", "01_감사친화형_자동검토큐_양식.xlsx"}
STALE_TOP_LEVEL = {"00_검토시작", "99_보관_로그", "README_먼저읽기.md"}
REQUIRED_WORKBOOK_SHEETS = [
    "01_오늘검토",
    "02_자료보완",
    "03_2차참고",
    "04_결과표초안",
    "06_판정기준",
]
REQUIRED_QUEUE_HEADERS = [
    "상태",
    "판정",
    "재확인",
    "검토메모",
    "담당자",
    "기관번호",
    "기관명",
    "항목",
    "검토유형",
    "값 비교",
    "검토요지",
    "확인질문",
    "제출엑셀 링크",
    "셀",
    "ALIO 링크",
    "근거자료 링크",
    "결과표초안 링크",
]
REQUIRED_HIDDEN_HEADERS = ["원본 제출자료 폴더", "결과표 문안 초안", "원천순번", "확인자료 전체"]
REQUIRED_RESULT_HEADERS = [
    "상태",
    "결과표 반영 여부",
    "재확인",
    "검토메모",
    "담당자",
    "기관번호",
    "기관명",
    "항목",
    "결과표 항목",
    "결과표 세항목",
    "불성실 유형 추정",
    "위반수준 추정",
    "불성실공시 세부내용 초안",
    "확인 필요사항",
    "제출엑셀 링크",
    "셀",
    "ALIO 링크",
    "근거자료 링크",
]
REQUIRED_2024_SUPPLEMENT_HEADERS = [
    "상태",
    "판정",
    "재확인",
    "검토메모",
    "2024커버리지",
    "우선순위",
    "항목",
    "원본행",
    "2024 지적유형",
    "현행 커버리지",
    "보완 필요 작업",
    "적용 검토자료",
    "검토질문",
]
QUEUE_DEFAULTS = {
    "01_오늘검토": ("확인전", "미입력", "N"),
    "02_자료보완": ("자료보완", "자료보완", "Y"),
    "03_2차참고": ("판단보류", "미입력", "N"),
}
REQUIRED_ALIO_KEYS = {"5-1", "6-2-reg", "6-2-first", "10-1", "13-1", "13-5"}


@dataclass
class AuditItem:
    status: str
    name: str
    detail: str


def find_dir(parent: Path, prefix: str) -> Path:
    matches = [p for p in parent.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not matches:
        raise FileNotFoundError(f"Missing directory {prefix}* under {parent}")
    return sorted(matches, key=lambda p: p.name)[0]


def latest_file(parent: Path, prefix: str, suffix: str = ".xlsx") -> Path | None:
    if not parent.exists():
        return None
    matches = [
        p
        for p in parent.iterdir()
        if p.is_file() and p.name.startswith(prefix) and p.suffix.lower() == suffix
    ]
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def latest_recursive(parent: Path, prefix: str, suffix: str) -> Path | None:
    if not parent.exists():
        return None
    matches: list[Path] = []
    for dirpath, dirnames, filenames in os.walk(parent, onerror=lambda _exc: None):
        dirnames[:] = [name for name in dirnames if not name.startswith("~$")]
        for filename in filenames:
            if filename.startswith("~$"):
                continue
            if filename.startswith(prefix) and filename.endswith(suffix):
                path = Path(dirpath) / filename
                try:
                    if path.is_file():
                        matches.append(path)
                except OSError:
                    continue
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def required_root_file(root: Path, contains: str, suffix: str = ".xlsx") -> Path | None:
    matches = [
        p
        for p in root.iterdir()
        if p.is_file() and p.suffix.lower() == suffix and contains in p.name
    ]
    if not matches:
        return None
    return sorted(matches, key=lambda p: p.name)[0]


def run_step(cmd: list[str], cwd: Path) -> None:
    env = os.environ.copy()
    env.setdefault("PYTHONIOENCODING", "utf-8")
    print("\n> " + " ".join(cmd), flush=True)
    subprocess.run(cmd, cwd=cwd, env=env, check=True)


def copy2(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    return target


def safe_remove_path(path: Path, boundary: Path) -> None:
    if not path.exists():
        return
    boundary_resolved = boundary.resolve()
    path_resolved = path.resolve()
    if path_resolved != boundary_resolved and not path_resolved.is_relative_to(boundary_resolved):
        raise RuntimeError(f"Refusing to remove outside output directory: {path_resolved}")
    if path.is_dir():
        shutil.rmtree(path, ignore_errors=True)
        if path.exists():
            subprocess.run(
                ["cmd.exe", "/d", "/c", "rmdir", "/s", "/q", str(path)],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
            if path.exists():
                print(f"warning: could not remove directory: {path}")
    else:
        try:
            path.unlink()
        except (FileNotFoundError, PermissionError) as exc:
            print(f"warning: could not remove file: {path} ({exc})")


def out_dir(root: Path) -> Path:
    return find_dir(root, "30_")


def cache_dir(out: Path) -> Path:
    return out / LOG_DIR_NAME / CACHE_DIR_NAME


def auto_cache(out: Path) -> Path:
    return cache_dir(out) / AUTO_CACHE_DIR


def reference_cache(out: Path) -> Path:
    return cache_dir(out) / REFERENCE_CACHE_DIR


def current_auto_dir(out: Path) -> Path:
    return out / "90_원천_자동검토" / "06_auto_review"


def current_core_dir(out: Path) -> Path:
    return out / "00_검토용_핵심산출물"


def seed_cache_from_existing(out: Path) -> None:
    """Preserve the minimum inputs needed for fast lightweight rebuilds."""
    cache = cache_dir(out)
    cache.mkdir(parents=True, exist_ok=True)

    auto_sources = [
        ("disclosure_auto_review_", ".xlsx", auto_cache(out)),
        ("evidence_match_review_", ".xlsx", auto_cache(out)),
    ]
    for prefix, suffix, target_dir in auto_sources:
        if latest_file(target_dir, prefix, suffix):
            continue
        source = latest_recursive(out, prefix, suffix)
        if source:
            copy2(source, target_dir / source.name)

    reference_sources = [
        ("02_", ".csv", reference_cache(out)),
        ("06_", ".csv", reference_cache(out)),
    ]
    for prefix, suffix, target_dir in reference_sources:
        if latest_file(target_dir, prefix, suffix):
            continue
        source = latest_recursive(out, prefix, suffix)
        if source:
            copy2(source, target_dir / source.name)


def classify_alio_files(alio_dir: Path) -> dict[str, Path]:
    files: dict[str, Path] = {}
    for path in alio_dir.glob("*.xlsx"):
        name = path.name
        if name.startswith("5-1"):
            files["5-1"] = path
        elif name.startswith("6-2") and "최초" in name:
            files["6-2-first"] = path
        elif name.startswith("6-2"):
            files["6-2-reg"] = path
        elif name.startswith("10"):
            files["10-1"] = path
        elif name.startswith("13-1"):
            files["13-1"] = path
        elif name.startswith("13-5"):
            files["13-5"] = path
    return files


def alio_preflight_score(alio_dir: Path) -> tuple[int, int, float]:
    files = classify_alio_files(alio_dir)
    required_hits = len(REQUIRED_ALIO_KEYS & set(files))
    xlsx_count = len(list(alio_dir.glob("*.xlsx")))
    try:
        mtime = alio_dir.stat().st_mtime
    except OSError:
        mtime = 0.0
    return required_hits, xlsx_count, mtime


def preflight_required_inputs(root: Path, out: Path, *, use_existing_auto: bool) -> None:
    """Fail before deleting outputs when required cached/upstream inputs are missing."""
    errors: list[str] = []
    for prefix in ("02_", "06_"):
        if latest_file(reference_cache(out), prefix, ".csv") is None:
            errors.append(f"{prefix}*.csv cache")

    if use_existing_auto:
        for prefix in ("disclosure_auto_review_", "evidence_match_review_"):
            if latest_file(auto_cache(out), prefix, ".xlsx") is None:
                errors.append(f"{prefix}*.xlsx cache")
    else:
        alio_candidates = [
            path
            for path in root.iterdir()
            if path.is_dir() and list(path.glob("*.xlsx"))
        ]
        if not alio_candidates:
            errors.append("ALIO source directory")
        else:
            alio_dir = max(alio_candidates, key=alio_preflight_score)
            alio_files = classify_alio_files(alio_dir)
            missing = REQUIRED_ALIO_KEYS - set(alio_files)
            if missing:
                errors.append(
                    f"ALIO required files missing in {alio_dir.name}: "
                    + ", ".join(sorted(missing))
                )
        base20 = next(
            (
                path
                for path in root.iterdir()
                if path.is_dir() and path.name.startswith("20_") and path.name != "20_materials"
            ),
            None,
        )
        submitted_root = None
        if base20:
            submitted_root = next((path for path in base20.iterdir() if path.is_dir() and path.name.startswith("02_")), None)
        if submitted_root is None:
            errors.append("20_*\\02_* submitted workbook directory")

    if errors:
        joined = ", ".join(errors)
        raise FileNotFoundError(f"Preflight failed before deleting outputs. Missing: {joined}")


def remove_existing_outputs(out: Path) -> None:
    """Delete previous reviewer/detail outputs while keeping only pipeline cache."""
    out.mkdir(parents=True, exist_ok=True)
    log_dir = out / LOG_DIR_NAME
    cache = cache_dir(out)
    cache.mkdir(parents=True, exist_ok=True)

    for path in list(out.iterdir()):
        if path.name == LOG_DIR_NAME:
            continue
        safe_remove_path(path, out)

    log_dir.mkdir(parents=True, exist_ok=True)
    for path in list(log_dir.iterdir()):
        if path.resolve() == cache.resolve():
            continue
        safe_remove_path(path, out)


def prepare_intermediate_inputs(out: Path, *, use_existing_auto: bool) -> None:
    core_ref = current_core_dir(out) / "99_원천산출물_참고용"
    core_ref.mkdir(parents=True, exist_ok=True)

    for prefix in ("02_", "06_"):
        source = latest_file(reference_cache(out), prefix, ".csv")
        if source is None:
            raise FileNotFoundError(
                f"Missing cached reference input {prefix}*.csv. "
                "Run the full upstream shortlist extraction first or restore the previous outputs."
            )
        copy2(source, core_ref / source.name)

    if use_existing_auto:
        auto_dir = current_auto_dir(out)
        auto_dir.mkdir(parents=True, exist_ok=True)
        for prefix in ("disclosure_auto_review_", "evidence_match_review_"):
            source = latest_file(auto_cache(out), prefix, ".xlsx")
            if source is None:
                raise FileNotFoundError(
                    f"Missing cached auto-review workbook {prefix}*.xlsx. "
                    "Run without --use-existing-auto once."
                )
            copy2(source, auto_dir / source.name)


def update_auto_cache(out: Path) -> None:
    for prefix in ("disclosure_auto_review_", "evidence_match_review_"):
        source = latest_file(current_auto_dir(out), prefix, ".xlsx")
        if source is None:
            raise FileNotFoundError(f"Missing generated auto-review workbook {prefix}*.xlsx")
        copy2(source, auto_cache(out) / source.name)


def sync_latest_auto_outputs(out: Path) -> None:
    core_ref = current_core_dir(out) / "99_원천산출물_참고용"
    core_ref.mkdir(parents=True, exist_ok=True)
    disclosure = latest_file(current_auto_dir(out), "disclosure_auto_review_", ".xlsx")
    evidence = latest_file(current_auto_dir(out), "evidence_match_review_", ".xlsx")
    if disclosure is None:
        raise FileNotFoundError("No disclosure_auto_review_*.xlsx found.")
    if evidence is None:
        raise FileNotFoundError("No evidence_match_review_*.xlsx found.")
    copy2(disclosure, core_ref / "10_자동검토_공시대조_최신전체본.xlsx")
    copy2(evidence, core_ref / "11_자동검토_증빙매칭_최신전체본.xlsx")


def write_root_readme(out: Path) -> Path:
    path = out / ROOT_README
    text = """# 산출물 안내

검토자는 `01_검토시작` 폴더만 열면 됩니다.

```text
30_검토산출물\\01_검토시작
```

먼저 열 파일은 하나입니다.

```text
01_검토큐.xlsx
```

업무 흐름은 `01_오늘검토 확인 -> 값 비교 열에서 ALIO 값/기관 제출값/차이 확인 -> 제출엑셀 링크/ALIO 링크/근거자료 링크 대조 -> 재확인/메모 입력 -> 결과표초안 링크 확인 -> 확정 건만 결과표 반영`입니다.

2024년 1차 공시점검 유형은 별도 작업 탭이 아니라 `06_판정기준` 시트의 점검 가이드라인으로 통합되어 있습니다.

상세 자동화 산출물은 파이프라인 실행 중에만 생성되고 마지막에 삭제됩니다. 내부 재실행용 최소 캐시는 `90_실행로그\\90_재실행캐시`에만 남깁니다.
"""
    path.write_text(text.rstrip() + "\n", encoding="utf-8-sig")
    return path


def final_cleanup(out: Path) -> None:
    keep = {LIGHT_DIR_NAME, LOG_DIR_NAME, ROOT_README}
    for path in list(out.iterdir()):
        if path.name in keep:
            continue
        safe_remove_path(path, out)


def sync_delivery_package(root: Path, out: Path) -> Path:
    """Mirror the final lightweight review package into 40_전달패키지."""
    delivery = root / DELIVERY_DIR_NAME
    delivery.mkdir(parents=True, exist_ok=True)
    for path in list(delivery.iterdir()):
        safe_remove_path(path, delivery)

    source_readme = out / ROOT_README
    source_light = out / LIGHT_DIR_NAME
    if not source_readme.exists():
        raise FileNotFoundError(f"Missing root guide: {source_readme}")
    if not source_light.exists():
        raise FileNotFoundError(f"Missing review start directory: {source_light}")

    copy2(source_readme, delivery / ROOT_README)
    shutil.copytree(source_light, delivery / LIGHT_DIR_NAME)
    return delivery


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def audit_light_workbook(
    light_workbook: Path,
    ok,
    warn,
    fail,
) -> None:
    if not light_workbook.exists():
        return

    try:
        wb = load_workbook(light_workbook, data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - audit must report, not crash.
        fail("최종 엑셀 UX 검사", f"엑셀을 열 수 없음: {exc}")
        return

    try:
        if wb.sheetnames == REQUIRED_WORKBOOK_SHEETS:
            ok("최종 엑셀 시트 순서", " -> ".join(wb.sheetnames))
        else:
            missing = [name for name in REQUIRED_WORKBOOK_SHEETS if name not in wb.sheetnames]
            if missing:
                fail("최종 엑셀 시트 순서", "누락 시트: " + ", ".join(missing))
            else:
                warn("최종 엑셀 시트 순서", "현재 순서: " + " -> ".join(wb.sheetnames))

        if wb.active.title == "01_오늘검토":
            ok("최종 엑셀 활성 시트", "01_오늘검토")
        else:
            fail("최종 엑셀 활성 시트", f"현재 활성 시트: {wb.active.title}")

        if "01_오늘검토" not in wb.sheetnames:
            return

        ws = wb["01_오늘검토"]
        headers = [cell_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        if headers[: len(REQUIRED_QUEUE_HEADERS)] == REQUIRED_QUEUE_HEADERS:
            ok("검토 큐 필수 열", "앞쪽 입력/판단/링크 열 배치 확인")
        else:
            missing = [header for header in REQUIRED_QUEUE_HEADERS if header not in headers]
            if missing:
                fail("검토 큐 필수 열", "누락 열: " + ", ".join(missing))
            else:
                warn("검토 큐 필수 열", "필수 열은 있으나 앞쪽 순서가 다름")

        data_counts = {
            name: max(wb[name].max_row - 1, 0)
            for name in REQUIRED_WORKBOOK_SHEETS
            if name in wb.sheetnames and name != "06_판정기준"
        }
        if data_counts.get("01_오늘검토", 0) > 0:
            ok("검토 큐 행 수", ", ".join(f"{k} {v}건" for k, v in data_counts.items()))
        else:
            fail("검토 큐 행 수", "01_오늘검토 데이터 행이 없음")

        default_errors = []
        for sheet_name, expected in QUEUE_DEFAULTS.items():
            if sheet_name not in wb.sheetnames:
                continue
            target_ws = wb[sheet_name]
            if target_ws.max_row < 2:
                continue
            observed = {
                (
                    cell_text(target_ws.cell(row_idx, 1).value),
                    cell_text(target_ws.cell(row_idx, 2).value),
                    cell_text(target_ws.cell(row_idx, 3).value),
                )
                for row_idx in range(2, target_ws.max_row + 1)
            }
            if observed != {expected}:
                sample = ", ".join("/".join(values) for values in sorted(observed)[:3])
                default_errors.append(f"{sheet_name}: 기대 {'/'.join(expected)}, 현재 {sample}")
        if default_errors:
            fail("시트별 기본 상태", "; ".join(default_errors))
        else:
            ok("시트별 기본 상태", "01_오늘검토 확인전, 02_자료보완 Y, 03_2차참고 판단보류")

        workbook_dir = light_workbook.parent
        link_quality_cols: dict[str, list[int]] = {}
        for sheet_name in ("01_오늘검토", "02_자료보완", "03_2차참고"):
            if sheet_name not in wb.sheetnames:
                continue
            target_ws = wb[sheet_name]
            target_headers = [
                cell_text(target_ws.cell(1, col).value)
                for col in range(1, target_ws.max_column + 1)
            ]
            target_map = {header: idx + 1 for idx, header in enumerate(target_headers)}
            link_quality_cols[sheet_name] = [
                target_map[h]
                for h in ("제출엑셀 링크", "ALIO 링크", "근거자료 링크", "결과표초안 링크")
                if h in target_map
            ]
        if "04_결과표초안" in wb.sheetnames:
            result_ws_for_links = wb["04_결과표초안"]
            result_headers_for_links = [
                cell_text(result_ws_for_links.cell(1, col).value)
                for col in range(1, result_ws_for_links.max_column + 1)
            ]
            result_map_for_links = {
                header: idx + 1 for idx, header in enumerate(result_headers_for_links)
            }
            link_quality_cols["04_결과표초안"] = [
                result_map_for_links[h]
                for h in ("제출엑셀 링크", "ALIO 링크", "근거자료 링크")
                if h in result_map_for_links
            ]

        absolute_links = []
        multiline_links = []
        missing_targets = []
        for sheet_name, cols in link_quality_cols.items():
            target_ws = wb[sheet_name]
            for row_idx in range(2, target_ws.max_row + 1):
                for col_idx in cols:
                    link = target_ws.cell(row_idx, col_idx).hyperlink
                    if not link:
                        continue
                    target = link.target or link.location or ""
                    if not target:
                        continue
                    if "\n" in target or "\r" in target:
                        multiline_links.append((sheet_name, row_idx, col_idx))
                    if re.match(r"^[A-Za-z]:[\\/]", target):
                        absolute_links.append((sheet_name, row_idx, col_idx))
                    if target.startswith("#"):
                        continue
                    target_path = Path(target)
                    if not target_path.is_absolute():
                        target_path = workbook_dir / target
                    try:
                        exists = target_path.exists()
                    except OSError:
                        exists = False
                    if not exists:
                        missing_targets.append((sheet_name, row_idx, col_idx))
        if absolute_links or multiline_links or missing_targets:
            fail(
                "최종 엑셀 링크 대상",
                f"절대경로 {len(absolute_links)}개, 여러줄 {len(multiline_links)}개, 미존재 {len(missing_targets)}개",
            )
        else:
            ok("최종 엑셀 링크 대상", "상대경로/내부링크만 사용, 미존재 대상 없음")

        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        link_cols = [
            header_map[h]
            for h in ("제출엑셀 링크", "ALIO 링크", "근거자료 링크", "결과표초안 링크")
            if h in header_map
        ]
        link_count = sum(
            1
            for row_idx in range(2, ws.max_row + 1)
            for col_idx in link_cols
            if ws.cell(row_idx, col_idx).hyperlink
        )
        if link_count:
            ok("검토 큐 하이퍼링크", f"{link_count}개")
        else:
            fail("검토 큐 하이퍼링크", "제출엑셀/ALIO/근거자료/결과표초안 링크가 없음")

        draft_col = header_map.get("결과표초안 링크")
        draft_link_count = 0
        if draft_col:
            draft_link_count = sum(
                1
                for row_idx in range(2, ws.max_row + 1)
                if ws.cell(row_idx, draft_col).hyperlink
            )
        if draft_link_count >= data_counts.get("01_오늘검토", 0):
            ok("검토 큐 결과표초안 링크", f"{draft_link_count}개")
        else:
            fail(
                "검토 큐 결과표초안 링크",
                f"01_오늘검토 {data_counts.get('01_오늘검토', 0)}건 대비 {draft_link_count}개",
            )

        if "04_결과표초안" in wb.sheetnames:
            result_ws = wb["04_결과표초안"]
            result_headers = [
                cell_text(result_ws.cell(1, col).value)
                for col in range(1, result_ws.max_column + 1)
            ]
            if result_headers[: len(REQUIRED_RESULT_HEADERS)] == REQUIRED_RESULT_HEADERS:
                ok("결과표 초안 필수 열", "입력/문안/링크 열 배치 확인")
            else:
                missing = [header for header in REQUIRED_RESULT_HEADERS if header not in result_headers]
                if missing:
                    fail("결과표 초안 필수 열", "누락 열: " + ", ".join(missing))
                else:
                    warn("결과표 초안 필수 열", "필수 열은 있으나 앞쪽 순서가 다름")

            result_header_map = {header: idx + 1 for idx, header in enumerate(result_headers)}
            result_link_cols = [
                result_header_map[h]
                for h in ("제출엑셀 링크", "ALIO 링크", "근거자료 링크")
                if h in result_header_map
            ]
            result_link_count = sum(
                1
                for row_idx in range(2, result_ws.max_row + 1)
                for col_idx in result_link_cols
                if result_ws.cell(row_idx, col_idx).hyperlink
            )
            if result_link_count:
                ok("결과표 초안 하이퍼링크", f"{result_link_count}개")
            else:
                fail("결과표 초안 하이퍼링크", "제출엑셀/ALIO/근거자료 링크가 없음")

            result_hidden_missing = []
            for header in ("확인자료 전체", "원천순번"):
                col_idx = result_header_map.get(header)
                if not col_idx:
                    result_hidden_missing.append(header)
                    continue
                if not result_ws.column_dimensions[get_column_letter(col_idx)].hidden:
                    result_hidden_missing.append(header)
            if result_hidden_missing:
                warn("결과표 초안 원문 숨김", "숨김 아님/누락: " + ", ".join(result_hidden_missing))
            else:
                ok("결과표 초안 원문 숨김", "확인자료 전체, 원천순번")

        if "05_2024보완" in wb.sheetnames:
            fail("2024 가이드 통합", "05_2024보완 별도 작업 탭이 남아 있음")
        elif "06_판정기준" in wb.sheetnames:
            criteria_ws = wb["06_판정기준"]
            criteria_values = [
                cell_text(criteria_ws.cell(row_idx, col_idx).value)
                for row_idx in range(1, criteria_ws.max_row + 1)
                for col_idx in range(1, min(criteria_ws.max_column, 3) + 1)
            ]
            if any(value == "2024 점검유형 가이드" for value in criteria_values):
                guide_rows = sum(
                    1
                    for row_idx in range(1, criteria_ws.max_row + 1)
                    if "원본행" in cell_text(criteria_ws.cell(row_idx, 1).value)
                )
                ok("2024 가이드 통합", f"06_판정기준에 점검 확인 관점 {guide_rows}건 통합")
            else:
                fail("2024 가이드 통합", "06_판정기준에 2024 점검유형 가이드가 없음")

        hidden_ok = []
        hidden_missing = []
        for header in REQUIRED_HIDDEN_HEADERS:
            col_idx = header_map.get(header)
            if not col_idx:
                hidden_missing.append(header)
                continue
            if ws.column_dimensions[get_column_letter(col_idx)].hidden:
                hidden_ok.append(header)
            else:
                hidden_missing.append(header)
        if hidden_missing:
            warn("긴 상세 열 숨김", "숨김 아님/누락: " + ", ".join(hidden_missing))
        else:
            ok("긴 상세 열 숨김", ", ".join(hidden_ok))

        dv_count = len(ws.data_validations.dataValidation)
        if dv_count >= 3:
            ok("검토 입력 데이터검증", f"{dv_count}개")
        else:
            fail("검토 입력 데이터검증", "상태/판정/재확인 드롭다운이 부족함")

        oversized_rows = [
            row_idx
            for row_idx in range(2, min(ws.max_row, 80) + 1)
            if (ws.row_dimensions[row_idx].height or 0) > 60
        ]
        if oversized_rows:
            warn("검토 큐 행 높이", f"60pt 초과 행 샘플: {oversized_rows[:5]}")
        else:
            ok("검토 큐 행 높이", "본문 60pt 이하")
    finally:
        wb.close()


def audit(root: Path) -> list[AuditItem]:
    items: list[AuditItem] = []
    tools = root / "90_tools"
    out = out_dir(root)

    def ok(name: str, detail: str) -> None:
        items.append(AuditItem("OK", name, detail))

    def warn(name: str, detail: str) -> None:
        items.append(AuditItem("WARN", name, detail))

    def fail(name: str, detail: str) -> None:
        items.append(AuditItem("FAIL", name, detail))

    for script in [
        "disclosure_auto_review.py",
        "evidence_match_review.py",
        "audit_2024_type_coverage.py",
        "build_one_page_review_report.py",
        "build_review_start_package.py",
        "filter_first_pass_checklist.py",
        "build_result_table_draft.py",
        "build_light_review_package.py",
        "run_review_pipeline.py",
    ]:
        path = tools / script
        ok(script, str(path.relative_to(root))) if path.exists() else fail(script, "script is missing")

    type_book = required_root_file(root, "24")
    ok("2024 점검유형 원본", type_book.name) if type_book else fail("2024 점검유형 원본", "root에서 '*24*.xlsx' 파일을 찾지 못했습니다.")

    light_dir = out / LIGHT_DIR_NAME
    light_readme = light_dir / LIGHT_README
    light_workbook = light_dir / LIGHT_WORKBOOK
    root_readme = out / ROOT_README
    for path in (root_readme, light_readme, light_workbook):
        if path.exists():
            ok(path.name, str(path.relative_to(root)))
        else:
            fail(path.name, "경량 검토 산출물이 없습니다.")

    stale = sorted(path.name for path in light_dir.iterdir() if path.name in STALE_LIGHT_WORKBOOKS) if light_dir.exists() else []
    if stale:
        fail("구형 최종 엑셀 제거", "잔존 파일: " + ", ".join(stale))
    else:
        ok("구형 최종 엑셀 제거", "긴 파일명/구 파일명 없음")

    if light_dir.exists():
        allowed_light = {LIGHT_README, LIGHT_WORKBOOK}
        extra_light = sorted(path.name for path in light_dir.iterdir() if path.name not in allowed_light)
        if extra_light:
            warn("검토시작 폴더 경량화", "추가 항목: " + ", ".join(extra_light))
        else:
            ok("검토시작 폴더 경량화", "안내와 최종 엑셀만 남음")

    audit_light_workbook(light_workbook, ok, warn, fail)

    for prefix in ("disclosure_auto_review_", "evidence_match_review_"):
        path = latest_file(auto_cache(out), prefix, ".xlsx")
        ok(f"캐시 {prefix}*.xlsx", path.name) if path else warn(f"캐시 {prefix}*.xlsx", "빠른 재생성 캐시 없음")

    for prefix in ("02_", "06_"):
        path = latest_file(reference_cache(out), prefix, ".csv")
        ok(f"캐시 {prefix}*.csv", path.name) if path else fail(f"캐시 {prefix}*.csv", "경량 리스트 재생성 입력 없음")

    coverage_xlsx = out / LOG_DIR_NAME / "02_2024유형_커버리지.xlsx"
    coverage_md = out / LOG_DIR_NAME / "02_2024유형_커버리지.md"
    for path in (coverage_xlsx, coverage_md):
        if path.exists():
            ok(path.name, str(path.relative_to(root)))
        else:
            fail(path.name, "2024 유형 커버리지 점검 산출물이 없습니다.")

    visible = [p.name for p in out.iterdir()]
    old_visible = sorted(set(visible) & STALE_TOP_LEVEL)
    if old_visible:
        fail("구형 최상위 이름 제거", "잔존 항목: " + ", ".join(old_visible))
    else:
        ok("구형 최상위 이름 제거", "00_검토시작, 99_보관_로그, README_먼저읽기.md 없음")

    expected_visible = {LIGHT_DIR_NAME, LOG_DIR_NAME, ROOT_README}
    extra = sorted(set(visible) - expected_visible)
    if extra:
        warn("최상위 경량화", "추가 항목: " + ", ".join(extra))
    else:
        ok("최상위 경량화", "00_안내.md, 01_검토시작, 90_실행로그만 남음")

    delivery = root / DELIVERY_DIR_NAME
    delivery_workbook = delivery / LIGHT_DIR_NAME / LIGHT_WORKBOOK
    delivery_readme = delivery / ROOT_README
    extra_delivery = []
    if delivery.exists():
        allowed_delivery = {ROOT_README, LIGHT_DIR_NAME}
        extra_delivery = sorted(path.name for path in delivery.iterdir() if path.name not in allowed_delivery)
    if delivery_workbook.exists() and delivery_readme.exists() and not extra_delivery:
        ok("전달패키지 경량 동기화", f"{delivery_workbook.relative_to(root)}")
    else:
        problems = []
        if not delivery_workbook.exists():
            problems.append("01_검토큐.xlsx 없음")
        if not delivery_readme.exists():
            problems.append("00_안내.md 없음")
        if extra_delivery:
            problems.append("구형 항목 잔존: " + ", ".join(extra_delivery))
        fail("전달패키지 경량 동기화", "; ".join(problems) if problems else "40_전달패키지 없음")

    return items


def print_audit(items: Iterable[AuditItem]) -> None:
    print("\nPipeline audit")
    for item in items:
        print(f"[{item.status}] {item.name}: {item.detail}")


def write_audit_report(root: Path, items: list[AuditItem]) -> Path:
    out = out_dir(root)
    path = out / LOG_DIR_NAME / AUDIT_REPORT
    path.parent.mkdir(parents=True, exist_ok=True)
    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    rows = "\n".join(f"| {item.status} | {item.name} | {item.detail} |" for item in items)
    fail_count = sum(1 for item in items if item.status == "FAIL")
    warn_count = sum(1 for item in items if item.status == "WARN")
    text = f"""# 파이프라인 점검

생성시각: {generated_at}

## 결론

- FAIL: {fail_count}건
- WARN: {warn_count}건
- 최종 검토 산출물은 `30_검토산출물\\01_검토시작\\01_검토큐.xlsx`입니다.

## 점검 결과

| 상태 | 항목 | 내용 |
|---|---|---|
{rows}

## 표준 실행 명령

캐시된 자동검토 결과를 기준으로 빠르게 재생성:

```powershell
python .\\90_tools\\run_review_pipeline.py --root . --use-existing-auto
```

공시대조와 증빙매칭 자동검토부터 새로 실행:

```powershell
python .\\90_tools\\run_review_pipeline.py --root .
```
"""
    path.write_text(text, encoding="utf-8-sig")
    return path


def main(argv: Iterable[str] = sys.argv[1:]) -> int:
    parser = argparse.ArgumentParser(description="Run/audit the lightweight review pipeline.")
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--use-existing-auto", action="store_true", help="Use cached auto-review workbooks.")
    parser.add_argument("--audit-only", action="store_true", help="Only audit final lightweight outputs.")
    parser.add_argument("--limit", type=int, default=None, help="Pass --limit to expensive auto-review steps for smoke tests.")
    parser.add_argument("--scan-excel-text-for-titles", action="store_true", help="Pass through to evidence_match_review.py.")
    parser.add_argument(
        "--submitted-name-keyword",
        action="append",
        default=None,
        help="Submitted workbook filename keyword for auto-review steps. Repeat or comma-separate.",
    )
    parser.add_argument(
        "--disable-submitted-name-filter",
        action="store_true",
        help="Find submitted workbooks by sheet structure only in auto-review steps.",
    )
    args = parser.parse_args(list(argv))

    root = args.root.resolve()
    out = out_dir(root)

    if args.audit_only:
        items = audit(root)
        print_audit(items)
        report = write_audit_report(root, items)
        print(f"audit report: {report}")
        return 0

    seed_cache_from_existing(out)
    preflight_required_inputs(root, out, use_existing_auto=args.use_existing_auto)
    remove_existing_outputs(out)
    prepare_intermediate_inputs(out, use_existing_auto=args.use_existing_auto)

    py = sys.executable
    if args.use_existing_auto:
        print("using cached auto-review workbooks")
    else:
        disclosure_cmd = [py, str(root / "90_tools" / "disclosure_auto_review.py"), "--root", str(root)]
        evidence_cmd = [py, str(root / "90_tools" / "evidence_match_review.py"), "--root", str(root)]
        if args.limit is not None:
            disclosure_cmd += ["--limit", str(args.limit)]
            evidence_cmd += ["--limit", str(args.limit)]
        if args.scan_excel_text_for_titles:
            evidence_cmd.append("--scan-excel-text-for-titles")
        if args.disable_submitted_name_filter:
            disclosure_cmd.append("--disable-submitted-name-filter")
            evidence_cmd.append("--disable-submitted-name-filter")
        for keyword in args.submitted_name_keyword or []:
            disclosure_cmd += ["--submitted-name-keyword", keyword]
            evidence_cmd += ["--submitted-name-keyword", keyword]
        run_step(disclosure_cmd, root)
        run_step(evidence_cmd, root)
        update_auto_cache(out)

    sync_latest_auto_outputs(out)
    run_step([py, str(root / "90_tools" / "build_one_page_review_report.py"), "--root", str(root)], root)
    run_step([py, str(root / "90_tools" / "build_review_start_package.py"), "--root", str(root), "--no-zip"], root)
    run_step([py, str(root / "90_tools" / "filter_first_pass_checklist.py"), "--root", str(root)], root)
    run_step([py, str(root / "90_tools" / "build_result_table_draft.py"), "--root", str(root)], root)
    run_step([py, str(root / "90_tools" / "build_light_review_package.py"), "--root", str(root)], root)

    write_root_readme(out)
    final_cleanup(out)
    sync_delivery_package(root, out)
    run_step([py, str(root / "90_tools" / "audit_2024_type_coverage.py"), "--root", str(root)], root)
    items = audit(root)
    print_audit(items)
    report = write_audit_report(root, items)
    print(f"audit report: {report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
