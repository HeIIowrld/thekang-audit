# -*- coding: utf-8 -*-
"""Audit current pipeline coverage against the 2024 first-round check types."""

from __future__ import annotations

import argparse
import datetime as dt
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.append(str(Path(__file__).resolve().parent))
from filter_first_pass_checklist import (  # noqa: E402
    classify_2024_coverage,
    extract_2024_first_round_rules,
    find_2024_type_book,
)


LOG_DIR_NAME = "90_실행로그"
OUTPUT_XLSX = "02_2024유형_커버리지.xlsx"
OUTPUT_MD = "02_2024유형_커버리지.md"


def find_dir(parent: Path, prefix: str) -> Path:
    matches = [p for p in parent.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not matches:
        raise FileNotFoundError(f"Missing directory {prefix}* under {parent}")
    return sorted(matches, key=lambda p: p.name)[0]


def style_sheet(ws, widths: dict[str, int]) -> None:
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=9)
            cell.border = border
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def coverage_rows(root: Path) -> tuple[list[dict[str, str]], Counter[str], dict[str, Counter[str]]]:
    rows: list[dict[str, str]] = []
    status_counts: Counter[str] = Counter()
    by_item: dict[str, Counter[str]] = defaultdict(Counter)

    for rule in extract_2024_first_round_rules(root):
        status, coverage, action = classify_2024_coverage(rule)
        status_counts[status] += 1
        by_item[rule["sheet"]][status] += 1
        rows.append(
            {
                "판정": status,
                "항목": rule["sheet"],
                "원본행": rule["row"],
                "2024 지적유형": rule["issue"],
                "2024 비고": rule["note"],
                "현행 커버리지": coverage,
                "보완 필요 작업": action,
            }
        )

    return rows, status_counts, by_item


def write_xlsx(
    path: Path,
    source_name: str,
    generated_at: str,
    rows: list[dict[str, str]],
    status_counts: Counter[str],
    by_item: dict[str, Counter[str]],
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "01_요약"
    ws.append(["구분", "내용"])
    ws.append(["생성시각", generated_at])
    ws.append(["기준파일", source_name])
    ws.append(["결론", "현행 파이프라인은 2024년 1차 지적유형 전체를 자동 포착하지 못함"])
    ws.append(["전체 유형", len(rows)])
    ws.append(["직접커버", status_counts.get("직접커버", 0)])
    ws.append(["부분커버", status_counts.get("부분커버", 0)])
    ws.append(["미커버", status_counts.get("미커버", 0)])
    style_sheet(ws, {"A": 20, "B": 88})

    ws_item = wb.create_sheet("02_항목별")
    ws_item.append(["항목", "직접커버", "부분커버", "미커버", "전체"])
    for item in sorted(by_item):
        counter = by_item[item]
        ws_item.append(
            [
                item,
                counter.get("직접커버", 0),
                counter.get("부분커버", 0),
                counter.get("미커버", 0),
                sum(counter.values()),
            ]
        )
    style_sheet(ws_item, {"A": 34, "B": 12, "C": 12, "D": 12, "E": 10})

    ws_detail = wb.create_sheet("03_유형별")
    headers = ["판정", "항목", "원본행", "2024 지적유형", "2024 비고", "현행 커버리지", "보완 필요 작업"]
    ws_detail.append(headers)
    for row in rows:
        ws_detail.append([row.get(header, "") for header in headers])
    style_sheet(ws_detail, {"A": 12, "B": 30, "C": 10, "D": 90, "E": 24, "F": 52, "G": 58})

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_md(
    path: Path,
    source_name: str,
    generated_at: str,
    rows: list[dict[str, str]],
    status_counts: Counter[str],
    by_item: dict[str, Counter[str]],
) -> Path:
    item_lines = "\n".join(
        f"| {item} | {counter.get('직접커버', 0)} | {counter.get('부분커버', 0)} | {counter.get('미커버', 0)} | {sum(counter.values())} |"
        for item, counter in sorted(by_item.items())
    )
    uncovered = [row for row in rows if row["판정"] == "미커버"]
    uncovered_lines = "\n".join(
        f"| {row['항목']} | {row['원본행']} | {row['2024 지적유형']} | {row['보완 필요 작업']} |"
        for row in uncovered
    )
    text = f"""# 2024 유형 커버리지 점검

생성시각: {generated_at}

기준파일: `{source_name}`

## 결론

현행 파이프라인은 2024년 1차 점검 지적유형 전체를 자동으로 잡아내지 못합니다.
42개 지적유형 중 직접커버는 {status_counts.get('직접커버', 0)}건, 부분커버는 {status_counts.get('부분커버', 0)}건, 미커버는 {status_counts.get('미커버', 0)}건입니다.

## 전체 판정

| 판정 | 건수 |
|---|---:|
| 직접커버 | {status_counts.get('직접커버', 0)} |
| 부분커버 | {status_counts.get('부분커버', 0)} |
| 미커버 | {status_counts.get('미커버', 0)} |
| 전체 | {len(rows)} |

## 항목별 커버리지

| 항목 | 직접커버 | 부분커버 | 미커버 | 전체 |
|---|---:|---:|---:|---:|
{item_lines}

## 미커버 유형

| 항목 | 원본행 | 2024 지적유형 | 보완 필요 작업 |
|---|---:|---|---|
{uncovered_lines}
"""
    path.write_text(text, encoding="utf-8-sig")
    return path


def build(root: Path) -> list[Path]:
    source = find_2024_type_book(root)
    if source is None:
        raise FileNotFoundError("2024 점검유형 원본 파일을 찾지 못했습니다.")

    rows, status_counts, by_item = coverage_rows(root)
    if not rows:
        raise RuntimeError("2024년 1차 점검 지적유형을 추출하지 못했습니다.")

    out = find_dir(root, "30_")
    log_dir = out / LOG_DIR_NAME
    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    xlsx = write_xlsx(log_dir / OUTPUT_XLSX, source.name, generated_at, rows, status_counts, by_item)
    md = write_md(log_dir / OUTPUT_MD, source.name, generated_at, rows, status_counts, by_item)
    return [xlsx, md]


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit 2024 check-type coverage.")
    parser.add_argument("--root", type=Path, default=Path.cwd())
    args = parser.parse_args()
    for path in build(args.root.resolve()):
        print(path)


if __name__ == "__main__":
    main()
