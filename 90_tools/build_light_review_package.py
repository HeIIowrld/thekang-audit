# -*- coding: utf-8 -*-
"""Build the reviewer-facing lightweight audit queue workbook.

The final workbook is intentionally a review queue, not an auto-finalized
result table. It keeps the first screen focused on fields a reviewer edits and
links that open the submitted workbook, ALIO scrape, and evidence files.
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.append(str(Path(__file__).resolve().parent))
from filter_first_pass_checklist import (  # noqa: E402
    classify_2024_coverage,
    extract_2024_first_round_rules,
)


WORKBOOK_NAME = "01_검토큐.xlsx"
START_DIR_NAME = "01_검토시작"
START_README_NAME = "00_안내.md"
STALE_WORKBOOK_NAMES = [
    "01_수동검사_리스트.xlsx",
    "01_감사친화형_자동검토큐_양식.xlsx",
]

QUEUE_SHEETS = ["01_오늘검토", "02_자료보완", "03_2차참고"]
REQUIRED_SHEETS = [*QUEUE_SHEETS, "04_결과표초안", "06_판정기준"]
QUEUE_DEFAULTS = {
    "01_오늘검토": ("확인전", "미입력", "N"),
    "02_자료보완": ("자료보완", "자료보완", "Y"),
    "03_2차참고": ("판단보류", "미입력", "N"),
}

VISIBLE_HEADERS = [
    "상태",
    "판정",
    "재확인",
    "검토메모",
    "담당자",
    "기관번호",
    "기관명",
    "항목",
    "검토유형",
    "값 비교",
    "검토요지",
    "확인질문",
    "제출엑셀 링크",
    "셀",
    "ALIO 링크",
    "근거자료 링크",
    "결과표초안 링크",
]

DETAIL_HEADERS = [
    "원본 제출자료 폴더",
    "결과표 문안 초안",
    "차이/판정",
    "원천순번",
    "확인자료 전체",
]

MAIN_HEADERS = [*VISIBLE_HEADERS, *DETAIL_HEADERS]

RESULT_HEADERS = [
    "상태",
    "결과표 반영 여부",
    "재확인",
    "검토메모",
    "담당자",
    "기관번호",
    "기관명",
    "항목",
    "결과표 항목",
    "결과표 세항목",
    "불성실 유형 추정",
    "위반수준 추정",
    "불성실공시 세부내용 초안",
    "확인 필요사항",
    "제출엑셀 링크",
    "셀",
    "ALIO 링크",
    "근거자료 링크",
    "확인자료 전체",
    "원천순번",
]

SUPPLEMENT_2024_HEADERS = [
    "상태",
    "판정",
    "재확인",
    "검토메모",
    "2024커버리지",
    "우선순위",
    "항목",
    "원본행",
    "2024 지적유형",
    "현행 커버리지",
    "보완 필요 작업",
    "적용 검토자료",
    "검토질문",
]

STATUS_VALUES = ["확인전", "이상없음", "수정요청", "자료보완", "판단보류", "제외"]
JUDGMENT_VALUES = ["미입력", "결과표반영", "자료보완", "설명요청", "제외", "이상없음"]
RECHECK_VALUES = ["N", "Y"]
RESULT_DECISION_VALUES = ["미입력", "반영", "수정후반영", "자료보완", "미반영", "판단보류"]


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def flat(value: Any, limit: int = 500) -> str:
    value = re.sub(r"\s+", " ", text(value).replace("\n", " ")).strip()
    return value if len(value) <= limit else value[: limit - 1] + "…"


def pick(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = text(row.get(key))
        if value:
            return value
    return ""


def find_dir(parent: Path, prefix: str) -> Path:
    matches = [p for p in parent.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not matches:
        raise FileNotFoundError(f"Missing directory {prefix}* under {parent}")
    return sorted(matches, key=lambda p: p.name)[0]


def find_source_dir(out_dir: Path) -> Path:
    current = out_dir / "00_00_먼저보기"
    if current.exists():
        return current
    matches = []
    for archive_root in (out_dir / "90_실행로그", out_dir / "99_보관_로그"):
        matches.extend(
            path / "00_00_먼저보기"
            for path in archive_root.glob("90_상세산출물_보관_*")
            if (path / "00_00_먼저보기").exists()
        )
    if matches:
        return sorted(matches, key=lambda p: str(p.parent.name))[-1]
    raise FileNotFoundError("Missing 00_00_먼저보기 in current output or archived detail outputs")


def read_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Missing sheet {sheet_name!r} in {path}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    result: list[dict[str, Any]] = []
    for row in rows:
        if not any(text(v) for v in row):
            continue
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    wb.close()
    return result


def clean_material_path(raw: str) -> str:
    raw = raw.strip().strip('"')
    raw = re.sub(r"^(ALIO 공시자료|제출엑셀|근거자료)\s*:\s*", "", raw).strip()
    return raw


def path_candidates(root: Path, raw: str) -> list[Path]:
    raw = clean_material_path(raw)
    if not raw:
        return []
    candidate = Path(raw)
    candidates = [candidate]
    if not candidate.is_absolute():
        candidates.extend(
            [
                root / candidate,
                root.parent / candidate,
                root / "20_기초자료" / candidate,
                root / "20_기초자료" / "02_배정완료_원본" / candidate,
                root / "20_기준자료" / candidate,
                root / "20_기준자료" / "02_배정완료_원본" / candidate,
                root / "20_materials" / candidate,
            ]
        )
        if candidate.parts and candidate.parts[0] == root.name:
            candidates.append(root.parent / candidate)
    return candidates


def resolve_path(root: Path, raw: str) -> Path | None:
    for path in path_candidates(root, raw):
        try:
            if path.exists():
                return path.resolve()
        except OSError:
            continue
    return None


def split_materials(root: Path, raw: Any) -> dict[str, Any]:
    lines = [line.strip() for line in text(raw).splitlines() if line.strip()]
    submitted = ""
    alio = ""
    cell = ""
    evidence: list[str] = []
    for line in lines:
        if line.startswith("셀/범위"):
            cell = line.replace("셀/범위:", "").strip()
        elif line.startswith("ALIO 공시자료"):
            alio = clean_material_path(line)
        elif line == "근거자료 파일 없음":
            evidence.append(line)
        elif not submitted:
            submitted = clean_material_path(line)
        else:
            evidence.append(clean_material_path(line))

    submitted_path = resolve_path(root, submitted)
    alio_path = resolve_path(root, alio)
    evidence_paths = [resolve_path(root, item) for item in evidence if item != "근거자료 파일 없음"]
    evidence_paths = [path for path in evidence_paths if path is not None]
    submitted_folder = submitted_path.parent if submitted_path and submitted_path.is_file() else submitted_path

    return {
        "submitted": submitted,
        "submitted_path": submitted_path,
        "submitted_folder": submitted_folder,
        "alio": alio,
        "alio_path": alio_path,
        "cell": cell,
        "evidence": "\n".join(evidence),
        "evidence_path": evidence_paths[0] if evidence_paths else None,
    }


def compact_key(value: str) -> str:
    return re.sub(r"[\s_()（）\[\]{}·.,/-]+", "", text(value)).lower()


def submission_aliases(value: str) -> list[str]:
    value = text(value)
    if not value:
        return []
    aliases = [value]
    no_number = re.sub(r"^\d+\.\s*", "", value).strip()
    aliases.append(no_number)
    no_category = re.sub(r"\((?:공기업|준정부기관|기타공공기관|부설기관)[^)]*\)", "", no_number).strip()
    aliases.append(no_category)
    result: list[str] = []
    seen: set[str] = set()
    for alias in aliases:
        key = compact_key(alias)
        if len(key) < 4 or key in seen:
            continue
        seen.add(key)
        result.append(alias)
    return result


def find_submission_folder(root: Path, org_name: str, raw_path: str = "") -> Path | None:
    base = root / "20_기준자료" / "02_배정완료_원본"
    if not base.exists():
        return None
    keys = submission_aliases(org_name)
    raw = text(raw_path)
    if raw:
        for part in Path(raw).parts:
            if part and not part.lower().endswith((".xlsx", ".xls", ".pdf", ".hwp", ".hwpx")):
                keys.extend(submission_aliases(part))
    if not keys:
        return None
    try:
        folders = [path for path in base.iterdir() if path.is_dir()]
    except OSError:
        return None
    scored: list[tuple[int, int, Path]] = []
    for folder in folders:
        folder_key = compact_key(folder.name)
        for key in keys:
            key_compact = compact_key(key)
            if key and key in folder.name:
                scored.append((0, len(folder.name), folder))
                break
            if key_compact and key_compact in folder_key:
                scored.append((1, len(folder.name), folder))
                break
    if not scored:
        return None
    return sorted(scored, key=lambda item: (item[0], item[1], item[2].name))[0][2].resolve()


def load_result_drafts(path: Path) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    if not path.exists():
        return result
    wb = load_workbook(path, read_only=True, data_only=True)
    sheetnames = set(wb.sheetnames)
    wb.close()
    for sheet_name in ("결과표_기재후보", "자료보완_요청후보", "2차_참고후보"):
        if sheet_name not in sheetnames:
            continue
        for row in read_rows(path, sheet_name):
            key = text(row.get("원천순번"))
            if key and key not in result:
                result[key] = row
    return result


def clean_value_text(value: Any, limit: int = 180) -> str:
    value_text = flat(value, limit)
    return re.sub(
        r"^(ALIO 공시값|ALIO 값|제출엑셀 값|기관 제출값|제출값|값1|값2)\s*:\s*",
        "",
        value_text,
    ).strip()


def value_compare_line(row: dict[str, Any]) -> str:
    review_type = text(row.get("검토유형"))
    control = text(row.get("대조군 값"))
    compare = text(row.get("비교군 값"))
    finding = text(row.get("차이/판정"))
    if not (control or compare or finding):
        return ""

    haystack = "\n".join([review_type, control, compare])
    if "ALIO" in haystack and ("제출" in haystack or "기관" in haystack):
        alio_value = ""
        submitted_value = ""
        for value in (control, compare):
            if "ALIO" in value and not alio_value:
                alio_value = clean_value_text(value)
            if "제출" in value and not submitted_value:
                submitted_value = clean_value_text(value)

        parts = []
        if alio_value:
            parts.append(f"ALIO: {alio_value}")
        if submitted_value:
            parts.append(f"기관제출: {submitted_value}")
        if finding:
            parts.append(f"차이/판정: {clean_value_text(finding, 220)}")
        return flat(" / ".join(parts), 520)

    parts = []
    if control:
        parts.append(f"대조값: {clean_value_text(control)}")
    if compare:
        parts.append(f"비교값: {clean_value_text(compare)}")
    if finding:
        parts.append(f"차이/판정: {clean_value_text(finding, 220)}")
    return flat(" / ".join(parts), 520)


def manual_row(
    root: Path,
    row: dict[str, Any],
    draft: dict[str, Any] | None,
    *,
    status: str = "확인전",
    judgment: str = "미입력",
    recheck: str = "N",
) -> list[Any]:
    materials = split_materials(root, row.get("확인자료"))
    source_no = text(row.get("원천순번"))
    org_name = pick(row, "기관명", "기관")
    fallback_folder = find_submission_folder(root, org_name, str(materials["submitted"]))
    submitted_link = str(materials["submitted_path"] or "")
    folder_link = str(materials["submitted_folder"] or fallback_folder or "")
    draft_text = text((draft or {}).get("불성실공시 세부내용 초안"))
    value_line = value_compare_line(row)
    question = "\n".join(
        part
        for part in [
            text(row.get("검토포인트")),
            text(row.get("검증 질문")),
            text(row.get("필터링 기준")),
            text(row.get("검증 보정 메모")),
        ]
        if part
    )
    review_summary = text(row.get("수집요약")) or flat(f"{row.get('발견 요약')} {row.get('차이/판정')}", 700)

    return [
        status,
        judgment,
        recheck,
        "",
        pick(row, "담당자"),
        pick(row, "기관번호"),
        org_name,
        pick(row, "항목"),
        pick(row, "검토유형"),
        value_line,
        review_summary,
        question,
        submitted_link,
        materials["cell"],
        str(materials["alio_path"] or materials["alio"]),
        str(materials["evidence_path"] or "근거자료 파일 없음"),
        "",
        folder_link,
        draft_text,
        text(row.get("차이/판정")),
        source_no,
        text(row.get("확인자료")),
    ]


def result_row(root: Path, row: dict[str, Any]) -> list[Any]:
    materials = split_materials(root, row.get("확인자료"))
    org_name = pick(row, "기관명", "기관")
    fallback_folder = find_submission_folder(root, org_name, str(materials["submitted"]))
    submitted_link = str(materials["submitted_path"] or materials["submitted_folder"] or fallback_folder or "")
    return [
        "확인전",
        "미입력",
        "N",
        "",
        pick(row, "담당자"),
        pick(row, "기관번호"),
        org_name,
        pick(row, "항목"),
        pick(row, "결과표 항목"),
        pick(row, "결과표 세항목"),
        pick(row, "불성실 유형 추정"),
        pick(row, "위반수준 추정"),
        pick(row, "불성실공시 세부내용 초안"),
        pick(row, "확인 필요사항"),
        submitted_link,
        materials["cell"],
        str(materials["alio_path"] or materials["alio"]),
        str(materials["evidence_path"] or "근거자료 파일 없음"),
        pick(row, "확인자료"),
        pick(row, "원천순번"),
    ]


def supplement_priority(status: str, item: str, issue: str) -> int:
    if status == "부분커버":
        return 3
    if item.startswith("11-1."):
        return 1
    if item.startswith("6-2."):
        return 1
    if item.startswith("10-1."):
        return 2
    if "수정공시" in issue or "정오표" in issue:
        return 2
    return 3


def supplement_source_hint(item: str) -> str:
    if item.startswith("6-2."):
        return "ALIO 6-2 등록/최초등록 자료, 제출엑셀 6-2, 채용공고 첨부파일"
    if item.startswith("11-1."):
        return "제출엑셀 11-1, 직원 평균보수 세부수당 첨부, 급여지급/근속 산정 근거"
    if item.startswith("10-1."):
        return "ALIO 10-1, 제출엑셀 10-1, 임원보수규정, 성과상여금/연봉제 근거"
    if item.startswith("13-1."):
        return "ALIO 13-1, 제출엑셀 13-1, 수정공시/정오표 이력, 복리후생 예산 근거"
    if item.startswith("5-1."):
        return "ALIO 5-1, 제출엑셀 5-1, 개인별 채용 로데이터"
    if item.startswith("13-5."):
        return "ALIO 13-5, 제출엑셀 13-5, 11-1 상시종업원수, 13-1 총액"
    return "해당 항목 제출엑셀 및 근거자료"


def supplement_question(status: str, item: str, issue: str) -> str:
    if item.startswith("6-2."):
        return "공고문/첨부파일/내부결재일을 열어 근무분야, 공개일, 결과확정일, 필수첨부 누락 여부를 확인"
    if item.startswith("11-1."):
        return "상시종업원수, 평균근속연수, 성별 평균임금, 세부수당 첨부 금액이 매뉴얼 산식과 맞는지 확인"
    if item.startswith("10-1."):
        return "연봉제 비고, 성과상여금 공시, 상임현원 존재 시 10-1 공시 여부, 만근 환산 기준을 확인"
    if item.startswith("13-1.") and ("수정공시" in issue or "정오표" in issue):
        return "정기공시 이후 수정공시/정오표 이력과 ALIO 반영 여부를 확인"
    if status == "부분커버":
        return "자동 후보가 같은 사안을 포착했는지 확인하고, 미포착 판단요소는 수동으로 보완 확인"
    return "2024 지적유형과 동일한 사실관계가 있는지 제출자료와 근거자료로 확인"


def supplement_2024_rows(root: Path) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for rule in extract_2024_first_round_rules(root):
        status, coverage, action = classify_2024_coverage(rule)
        if status == "직접커버":
            continue
        item = text(rule.get("sheet"))
        issue = text(rule.get("issue"))
        rows.append(
            [
                "확인전",
                "미입력",
                "Y",
                "",
                status,
                supplement_priority(status, item, issue),
                item,
                text(rule.get("row")),
                issue,
                coverage,
                action,
                supplement_source_hint(item),
                supplement_question(status, item, issue),
            ]
        )
    return sorted(rows, key=lambda row: (row[5], row[6], int(row[7]) if str(row[7]).isdigit() else 9999))


def hyperlink_target(target: str, workbook_dir: Path) -> str:
    target = text(target)
    if not target or target == "근거자료 파일 없음":
        return ""
    path = Path(target)
    try:
        if path.exists():
            return os.path.relpath(path.resolve(), workbook_dir.resolve())
    except (OSError, ValueError):
        return target
    return target


def short_label(value: str, limit: int = 34) -> str:
    value = text(value)
    if len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def path_name(target: str) -> str:
    target = text(target)
    if not target:
        return ""
    try:
        name = Path(target).name
    except (OSError, ValueError):
        name = ""
    return name or target


def parent_folder(target: str) -> str:
    target = text(target)
    if not target:
        return ""
    try:
        path = Path(target)
        return str(path.parent if path.suffix else path)
    except (OSError, ValueError):
        return ""


def link_label(prefix: str, target: str, limit: int = 34) -> str:
    name = short_label(path_name(target), limit)
    return f"{prefix}: {name}" if name else prefix


def add_hyperlink_with_dir(cell, target: str, label: str, workbook_dir: Path) -> None:
    link = hyperlink_target(target, workbook_dir)
    if not link:
        return
    cell.value = label
    cell.hyperlink = link
    cell.style = "Hyperlink"


def list_formula(values: list[str]) -> str:
    return '"' + ",".join(values) + '"'


def add_queue_validations(ws) -> None:
    status_dv = DataValidation(type="list", formula1=list_formula(STATUS_VALUES), allow_blank=False)
    judgment_dv = DataValidation(type="list", formula1=list_formula(JUDGMENT_VALUES), allow_blank=False)
    recheck_dv = DataValidation(type="list", formula1=list_formula(RECHECK_VALUES), allow_blank=False)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(judgment_dv)
    ws.add_data_validation(recheck_dv)
    if ws.max_row >= 2:
        status_dv.add(f"A2:A{ws.max_row}")
        judgment_dv.add(f"B2:B{ws.max_row}")
        recheck_dv.add(f"C2:C{ws.max_row}")


def add_result_validations(ws) -> None:
    status_dv = DataValidation(type="list", formula1=list_formula(STATUS_VALUES), allow_blank=False)
    decision_dv = DataValidation(type="list", formula1=list_formula(RESULT_DECISION_VALUES), allow_blank=False)
    recheck_dv = DataValidation(type="list", formula1=list_formula(RECHECK_VALUES), allow_blank=False)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(decision_dv)
    ws.add_data_validation(recheck_dv)
    if ws.max_row >= 2:
        status_dv.add(f"A2:A{ws.max_row}")
        decision_dv.add(f"B2:B{ws.max_row}")
        recheck_dv.add(f"C2:C{ws.max_row}")


def add_supplement_validations(ws) -> None:
    status_dv = DataValidation(type="list", formula1=list_formula(STATUS_VALUES), allow_blank=False)
    judgment_dv = DataValidation(type="list", formula1=list_formula(JUDGMENT_VALUES), allow_blank=False)
    recheck_dv = DataValidation(type="list", formula1=list_formula(RECHECK_VALUES), allow_blank=False)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(judgment_dv)
    ws.add_data_validation(recheck_dv)
    if ws.max_row >= 2:
        status_dv.add(f"A2:A{ws.max_row}")
        judgment_dv.add(f"B2:B{ws.max_row}")
        recheck_dv.add(f"C2:C{ws.max_row}")


def style_queue_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions
    tab_colors = {
        "01_오늘검토": "1F4E78",
        "02_자료보완": "C65911",
        "03_2차참고": "5B9BD5",
    }
    if ws.title in tab_colors:
        ws.sheet_properties.tabColor = tab_colors[ws.title]

    widths = {
        "A": 10,
        "B": 14,
        "C": 8,
        "D": 24,
        "E": 10,
        "F": 11,
        "G": 26,
        "H": 12,
        "I": 18,
        "J": 38,
        "K": 44,
        "L": 38,
        "M": 24,
        "N": 10,
        "O": 20,
        "P": 26,
        "Q": 26,
        "R": 28,
        "S": 60,
        "T": 32,
        "U": 10,
        "V": 70,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_headers = {"검토메모", "값 비교", "검토요지", "확인질문", "결과표 문안 초안", "확인자료 전체"}
    input_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.row_dimensions[1].height = 26
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 52

    for row in ws.iter_rows():
        for cell in row:
            header = text(ws.cell(1, cell.column).value)
            cell.alignment = Alignment(vertical="top", wrap_text=header in wrap_headers)
            cell.font = Font(size=9)
            cell.border = border
            if cell.row >= 2 and cell.column <= 4:
                cell.fill = input_fill

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for header in DETAIL_HEADERS:
        col_idx = MAIN_HEADERS.index(header) + 1
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    add_queue_validations(ws)


def style_supplement_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = "A64D79"
    widths = {
        "A": 10,
        "B": 12,
        "C": 8,
        "D": 24,
        "E": 12,
        "F": 10,
        "G": 30,
        "H": 9,
        "I": 78,
        "J": 46,
        "K": 54,
        "L": 44,
        "M": 48,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_headers = {"검토메모", "2024 지적유형", "현행 커버리지", "보완 필요 작업", "적용 검토자료", "검토질문"}
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    priority_fill = PatternFill("solid", fgColor="FCE4D6")
    ws.row_dimensions[1].height = 26
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 58

    for row in ws.iter_rows():
        for cell in row:
            header = text(ws.cell(1, cell.column).value)
            cell.alignment = Alignment(vertical="top", wrap_text=header in wrap_headers)
            cell.font = Font(size=9)
            cell.border = border
            if cell.row >= 2 and cell.column <= 4:
                cell.fill = input_fill
            if cell.row >= 2 and cell.column == 6 and cell.value == 1:
                cell.fill = priority_fill

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    add_supplement_validations(ws)


def style_result_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = "70AD47"
    widths = {
        "A": 10,
        "B": 16,
        "C": 8,
        "D": 24,
        "E": 10,
        "F": 11,
        "G": 26,
        "H": 12,
        "I": 16,
        "J": 18,
        "K": 18,
        "L": 16,
        "M": 72,
        "N": 48,
        "O": 24,
        "P": 10,
        "Q": 20,
        "R": 26,
        "S": 60,
        "T": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_headers = {"검토메모", "불성실공시 세부내용 초안", "확인 필요사항", "확인자료 전체"}
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    ws.row_dimensions[1].height = 26

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 58

    for row in ws.iter_rows():
        for cell in row:
            header = text(ws.cell(1, cell.column).value)
            cell.alignment = Alignment(vertical="top", wrap_text=header in wrap_headers)
            cell.font = Font(size=9)
            cell.border = border
            if cell.row >= 2 and cell.column <= 4:
                cell.fill = input_fill

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for header in ("확인자료 전체", "원천순번"):
        col_idx = RESULT_HEADERS.index(header) + 1
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    add_result_validations(ws)


def style_criteria_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for col, width in {"A": 24, "B": 44, "C": 64}.items():
        ws.column_dimensions[col].width = width
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=9)
            cell.border = border
    section_headers = {
        "구분",
        "시트",
        "입력 열",
        "주의",
        "공시-ALIO 매칭확인",
        "2024 점검유형 가이드",
        "우선순위/항목",
    }
    for row_idx in range(1, ws.max_row + 1):
        if text(ws.cell(row_idx, 1).value) in section_headers:
            for cell in ws[row_idx]:
                if cell.value:
                    cell.font = Font(bold=True, color="FFFFFF", size=9)
                    cell.fill = PatternFill("solid", fgColor="1F4E78")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def append_queue_rows(ws, rows: list[list[Any]], workbook_dir: Path, result_link_map: dict[str, int]) -> None:
    ws.append(MAIN_HEADERS)
    for row in rows:
        ws.append(row)
        row_idx = ws.max_row
        submitted = text(row[12])
        folder = text(row[17])
        alio = text(row[14])
        evidence = text(row[15])
        source_no = text(row[20])

        if submitted:
            add_hyperlink_with_dir(ws.cell(row_idx, 13), submitted, link_label("제출", submitted), workbook_dir)
        elif folder:
            add_hyperlink_with_dir(ws.cell(row_idx, 13), folder, link_label("제출 없음/원본폴더", folder), workbook_dir)

        if alio:
            add_hyperlink_with_dir(ws.cell(row_idx, 15), alio, link_label("ALIO", alio), workbook_dir)

        if evidence and evidence != "근거자료 파일 없음":
            add_hyperlink_with_dir(ws.cell(row_idx, 16), evidence, link_label("근거", evidence), workbook_dir)
        elif folder:
            add_hyperlink_with_dir(ws.cell(row_idx, 16), folder, link_label("근거 없음/원본폴더", folder), workbook_dir)
        else:
            ws.cell(row_idx, 16).value = "근거자료 없음"

        result_row_idx = result_link_map.get(source_no)
        if result_row_idx:
            add_hyperlink_with_dir(
                ws.cell(row_idx, 17),
                f"#'04_결과표초안'!A{result_row_idx}",
                f"초안: 04_결과표초안 {result_row_idx}행",
                workbook_dir,
            )
        else:
            ws.cell(row_idx, 17).value = "초안 없음"

        if folder:
            add_hyperlink_with_dir(ws.cell(row_idx, 18), folder, link_label("원본폴더", folder), workbook_dir)

    style_queue_sheet(ws)


def append_result_rows(ws, rows: list[list[Any]], workbook_dir: Path) -> None:
    ws.append(RESULT_HEADERS)
    for row in rows:
        ws.append(row)
        row_idx = ws.max_row
        submitted = text(row[14])
        alio = text(row[16])
        evidence = text(row[17])
        folder = parent_folder(submitted)

        if submitted:
            submitted_prefix = "제출" if Path(submitted).suffix else "제출폴더"
            add_hyperlink_with_dir(ws.cell(row_idx, 15), submitted, link_label(submitted_prefix, submitted), workbook_dir)

        if alio:
            add_hyperlink_with_dir(ws.cell(row_idx, 17), alio, link_label("ALIO", alio), workbook_dir)

        if evidence and evidence != "근거자료 파일 없음":
            add_hyperlink_with_dir(ws.cell(row_idx, 18), evidence, link_label("근거", evidence), workbook_dir)
        elif folder:
            add_hyperlink_with_dir(ws.cell(row_idx, 18), folder, link_label("근거 없음/제출폴더", folder), workbook_dir)
        else:
            ws.cell(row_idx, 18).value = "근거자료 없음"

    style_result_sheet(ws)


def append_supplement_rows(ws, rows: list[list[Any]]) -> None:
    ws.append(SUPPLEMENT_2024_HEADERS)
    for row in rows:
        ws.append(row)
    style_supplement_sheet(ws)


def append_criteria(ws, generated_at: str, counts: Counter[str], guideline_rows: list[list[Any]]) -> None:
    ws.append(["구분", "내용", "운영 기준"])
    ws.append(["생성시각", generated_at, "파이프라인 실행 시점 기준"])
    ws.append(["목적", "검토 보고서 자동 확정이 아니라 수동 검토 큐 생성", "검토자가 확인한 건만 결과표에 반영"])
    ws.append(["첫 화면", "01_오늘검토", "상태/판정/재확인/검토메모 입력 후 링크 자료 확인"])
    ws.append(["링크 흐름", "제출엑셀 링크 -> ALIO 링크 -> 근거자료 링크 -> 결과표초안 링크", "검토행에서 초안 링크로 같은 건의 결과표 문안까지 이동"])
    ws.append(["숨김 열", "원본 제출자료 폴더, 결과표 문안 초안, 차이/판정, 원천순번, 확인자료 전체", "필요 시 열 숨김 해제로 원문 확인"])
    ws.append([])
    ws.append(["시트", "건수", "검토 기준"])
    ws.append(["01_오늘검토", counts.get("01_오늘검토", 0), "기본값 확인전/미입력/N. 공시-제출 수치, 6-2 목록, 제출엑셀 내부 검증 등 즉시 확인 후보"])
    ws.append(["02_자료보완", counts.get("02_자료보완", 0), "기본값 자료보완/자료보완/Y. 근거자료 없음 또는 제출자료 불충분으로 기관 재확인이 필요한 후보"])
    ws.append(["03_2차참고", counts.get("03_2차참고", 0), "기본값 판단보류/미입력/N. 1차 확정 전 참고 또는 후순위 검토 후보"])
    ws.append(["04_결과표초안", counts.get("04_결과표초안", 0), "자동 확정본이 아니라 문안 후보. 링크 자료 대조 후 반영 여부를 입력"])
    ws.append([])
    ws.append(["입력 열", "허용값", "의미"])
    ws.append(["상태", ", ".join(STATUS_VALUES), "현재 처리 상태"])
    ws.append(["판정", ", ".join(JUDGMENT_VALUES), "수동 검토 후 결과표 또는 보완 흐름으로 넘길 판단"])
    ws.append(["재확인", ", ".join(RECHECK_VALUES), "기관 또는 내부 재확인이 필요하면 Y"])
    ws.append(["검토메모", "자유 입력", "판단 근거와 추후 추적할 내용을 짧게 기록"])
    ws.append([])
    ws.append(["주의", "자동화 결과는 후보 축소용", "2024 점검유형은 별도 작업 탭이 아니라 점검 시 확인 관점입니다. 실제 결과표 반영은 검토자가 확정합니다."])
    ws.append([])
    ws.append(["공시-ALIO 매칭확인", "03_2차참고", "ALIO 원천 파일은 있으나 기관명·고용형태·항목·연도 키가 제출엑셀과 매칭되지 않은 후보"])
    ws.append(["확인 포인트", "ALIO 원천 파일, 기관명 정규화, 항목명 prefix, 고용형태 표기 차이", "원천 자동검토에는 셀 단위 전체 후보가 남고 최종 큐에는 기관·항목 단위 대표 후보가 표시됨"])
    ws.append([])
    ws.append(["2024 점검유형 가이드", "별도 작업 탭 아님", "아래 항목은 점검 시 놓치지 말아야 할 확인 관점입니다. 직접 자동검토 후보가 있으면 해당 큐에서 보고, 없으면 기관별 원본 제출자료와 근거자료를 수동 확인합니다."])
    ws.append(["우선순위/항목", "2024 지적유형", "확인자료 및 질문"])
    for row in guideline_rows:
        priority = row[5]
        item = row[6]
        source_row = row[7]
        issue = row[8]
        action = row[10]
        source_hint = row[11]
        question = row[12]
        ws.append([
            f"{priority} / {item} / 원본행 {source_row}",
            issue,
            "\n".join(part for part in [action, source_hint, question] if text(part)),
        ])
    style_criteria_sheet(ws)


def make_workbook(
    root: Path,
    output_path: Path,
    first_rows: list[dict[str, Any]],
    supplement_rows: list[dict[str, Any]],
    second_rows: list[dict[str, Any]],
    result_rows: list[dict[str, Any]],
    drafts: dict[str, dict[str, Any]],
    generated_at: str,
) -> None:
    workbook_dir = output_path.parent.resolve()
    row_groups = {
        "01_오늘검토": [
            manual_row(
                root,
                row,
                drafts.get(text(row.get("원천순번"))),
                status=QUEUE_DEFAULTS["01_오늘검토"][0],
                judgment=QUEUE_DEFAULTS["01_오늘검토"][1],
                recheck=QUEUE_DEFAULTS["01_오늘검토"][2],
            )
            for row in first_rows
        ],
        "02_자료보완": [
            manual_row(
                root,
                row,
                drafts.get(text(row.get("원천순번"))),
                status=QUEUE_DEFAULTS["02_자료보완"][0],
                judgment=QUEUE_DEFAULTS["02_자료보완"][1],
                recheck=QUEUE_DEFAULTS["02_자료보완"][2],
            )
            for row in supplement_rows
        ],
        "03_2차참고": [
            manual_row(
                root,
                row,
                drafts.get(text(row.get("원천순번"))),
                status=QUEUE_DEFAULTS["03_2차참고"][0],
                judgment=QUEUE_DEFAULTS["03_2차참고"][1],
                recheck=QUEUE_DEFAULTS["03_2차참고"][2],
            )
            for row in second_rows
        ],
    }
    result_draft_rows = [result_row(root, row) for row in result_rows]
    result_link_map = {
        text(row[-1]): idx + 2
        for idx, row in enumerate(result_draft_rows)
        if text(row[-1])
    }
    supplement_rows_2024 = supplement_2024_rows(root)
    counts = Counter({name: len(rows) for name, rows in row_groups.items()})
    counts["04_결과표초안"] = len(result_draft_rows)

    wb = Workbook()
    ws_first = wb.active
    ws_first.title = "01_오늘검토"
    append_queue_rows(ws_first, row_groups["01_오늘검토"], workbook_dir, result_link_map)

    for sheet_name in ("02_자료보완", "03_2차참고"):
        ws = wb.create_sheet(sheet_name)
        append_queue_rows(ws, row_groups[sheet_name], workbook_dir, result_link_map)

    ws_result = wb.create_sheet("04_결과표초안")
    append_result_rows(ws_result, result_draft_rows, workbook_dir)

    ws_criteria = wb.create_sheet("06_판정기준")
    append_criteria(ws_criteria, generated_at, counts, supplement_rows_2024)

    wb.active = wb.index(ws_first)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    for stale in STALE_WORKBOOK_NAMES:
        stale_path = output_path.parent / stale
        if stale_path.exists() and stale_path.resolve() != output_path.resolve():
            stale_path.unlink()
    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise PermissionError(
            f"{output_path} 저장 실패: 파일이 Excel에서 열려 있으면 닫고 다시 실행하세요."
        ) from exc


def write_readme(output_dir: Path, generated_at: str, workbook_name: str = WORKBOOK_NAME) -> Path:
    text_value = f"""# 검토 시작 안내

생성시각: {generated_at}

이 폴더에서는 아래 엑셀 하나만 열고 검토를 시작하면 됩니다.

```text
{workbook_name}
```

## 사용 흐름

1. `01_오늘검토` 시트에서 상태, 판정, 재확인, 검토메모를 먼저 입력합니다.
2. `값 비교` 열에서 ALIO 값, 기관 제출값, 차이/판정을 먼저 확인한 뒤 `제출엑셀 링크`, `ALIO 링크`, `근거자료 링크`를 열어 수치와 문안을 대조합니다. `근거자료 링크`가 `근거 없음/원본폴더`로 표시되면 원본 제출 폴더에서 누락 여부를 먼저 확인합니다.
3. `02_자료보완`은 기본값이 `자료보완 / 자료보완 / Y`이므로 기관 재확인 대상으로 먼저 처리합니다.
4. `03_2차참고`는 기본값이 `판단보류 / 미입력 / N`인 후순위 검토 또는 참고자료로 사용합니다.
5. `결과표초안 링크` 열의 초안 링크를 눌러 같은 건의 `04_결과표초안` 문안으로 이동합니다.
6. 실제 위반으로 확정한 건만 `04_결과표초안`의 링크 자료와 문안을 검토 후 결과표 양식에 반영합니다.
7. 2024년 1차 지적유형은 별도 작업 탭이 아니라 `06_판정기준`의 점검 가이드라인입니다. 검토 시 확인 관점으로 참고하되, 실제 반영은 원본 제출자료와 근거자료 확인 후 판단합니다.

긴 원천 텍스트와 결과표 문안 초안은 각 작업 시트의 숨김 열에 남겨 두었습니다. 필요하면 열 숨김을 해제해 원문을 확인하면 됩니다.
"""
    path = output_dir / START_README_NAME
    path.write_text(text_value, encoding="utf-8-sig")
    return path


def build(root: Path, output_path: Path | None = None) -> list[Path]:
    out_dir = find_dir(root, "30_")
    source_dir = find_source_dir(out_dir)
    checklist = source_dir / "01_관리자용_전체검토현황_필터링.xlsx"
    result_draft = source_dir / "04_결과표_기재초안.xlsx"
    if not checklist.exists():
        raise FileNotFoundError(f"Missing checklist workbook: {checklist}")
    if not result_draft.exists():
        raise FileNotFoundError(f"Missing result draft workbook: {result_draft}")

    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    first_rows = read_rows(checklist, "1차_필터링체크리스트")
    supplement_rows = read_rows(checklist, "자료보완_근거자료없음")
    second_rows = read_rows(checklist, "2차_참고")
    result_rows = read_rows(result_draft, "결과표_기재후보")
    drafts = load_result_drafts(result_draft)

    output_path = output_path or out_dir / START_DIR_NAME / WORKBOOK_NAME
    output_dir = output_path.parent
    make_workbook(
        root,
        output_path,
        first_rows,
        supplement_rows,
        second_rows,
        result_rows,
        drafts,
        generated_at,
    )
    readme = write_readme(output_dir, generated_at, output_path.name)
    return [readme, output_path]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--output-path", type=Path, default=None)
    args = parser.parse_args()
    output_path = args.output_path.resolve() if args.output_path else None
    for path in build(args.root.resolve(), output_path=output_path):
        print(path)


if __name__ == "__main__":
    main()
