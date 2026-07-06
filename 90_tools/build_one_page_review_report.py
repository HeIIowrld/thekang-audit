# -*- coding: utf-8 -*-
"""
Build a reviewer-facing one-page summary from the review pipeline outputs.

The first sheet and markdown are intentionally written as a review checklist:
which institution, which submitted/public/evidence document, what was detected,
and what the reviewer should confirm next.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import math
import os
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from build_review_start_package import REVIEWERS, build_records, counts_by_reviewer

TARGET_PREFIXES = ("5-1.", "6-2.", "10-1.", "11-1.", "13-1.", "13-5.")
ALIO_MISSING_CHECK_TYPE = "기관/항목 ALIO 매칭 필요"
KEYWORD_UNFAIR_DISCLOSURE = "불성실공시"
PRIMARY_REPORT_XLSX = "00_여기부터_검토시작.xlsx"
LEGACY_REPORT_XLSX = "00_한장_요약보고서.xlsx"
STATUS_VALUES = "확인전,이상없음,수정요청,판단보류,제외"
JUDGMENT_VALUES = "미입력,수치수정,증빙보완,설명보완,제외,이상없음"
RECHECK_VALUES = "N,Y"
REPORT_HEADERS = [
    "순번",
    "상태",
    "판단 결과",
    "재확인 필요",
    "항목",
    "기관",
    "검토유형",
    "검토 질문",
    "발견 요약",
    "대조군 값",
    "비교군 값",
    "차이/판정",
    "확인자료",
    "검토 메모",
    "대조군 자료",
    "비교군 자료",
    "상세키/미매칭",
    "자동점수/원천근거",
]


def find_dir(parent: Path, prefix: str) -> Path:
    matches = [p for p in parent.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not matches:
        raise FileNotFoundError(f"Missing directory {prefix}* under {parent}")
    return sorted(matches, key=lambda p: p.name)[0]


def find_core_dir(out_dir: Path) -> Path:
    matches = [
        p
        for p in out_dir.iterdir()
        if p.is_dir() and p.name.startswith("00_") and "핵심산출물" in p.name
    ]
    if matches:
        return sorted(matches, key=lambda p: p.name)[0]
    legacy = [
        p
        for p in out_dir.iterdir()
        if p.is_dir() and p.name.startswith("00_") and "먼저보기" not in p.name
    ]
    if legacy:
        return sorted(legacy, key=lambda p: p.name)[0]
    raise FileNotFoundError(f"Missing core output directory under {out_dir}")


def find_auto_dir(out_dir: Path) -> Path:
    preferred = out_dir / "90_원천_자동검토" / "06_auto_review"
    if preferred.exists():
        return preferred
    legacy = out_dir / "06_auto_review"
    if legacy.exists():
        return legacy
    return preferred


def latest_file(parent: Path, prefix: str, suffix: str) -> Path:
    matches = [
        p
        for p in parent.iterdir()
        if p.is_file() and p.name.startswith(prefix) and p.suffix.lower() == suffix
    ]
    if not matches:
        raise FileNotFoundError(f"Missing file {prefix}*{suffix} under {parent}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def latest_file_any(parents: list[Path], prefix: str, suffix: str) -> Path:
    matches: list[Path] = []
    for parent in parents:
        if parent.exists():
            matches.extend(
                p
                for p in parent.iterdir()
                if p.is_file() and p.name.startswith(prefix) and p.suffix.lower() == suffix
            )
    if not matches:
        labels = ", ".join(str(parent) for parent in parents)
        raise FileNotFoundError(f"Missing file {prefix}*{suffix} under {labels}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def flat(value: Any, limit: int = 240) -> str:
    value = re.sub(r"\s+", " ", text(value).replace("\n", " ")).strip()
    return value if len(value) <= limit else value[: limit - 1] + "..."


def multiline(value: Any, limit: int = 420) -> str:
    value = text(value)
    value = "\n".join(re.sub(r"[ \t]+", " ", line).strip() for line in value.splitlines())
    value = re.sub(r"\n{3,}", "\n\n", value).strip()
    return value if len(value) <= limit else value[: limit - 1] + "..."


def number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isfinite(float(value)):
            return float(value)
        return None
    raw = str(value).strip().replace(",", "")
    raw = raw.replace("(", "-").replace(")", "")
    m = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def display_number(value: Any) -> str:
    num = number(value)
    if num is None:
        return flat(value, 80)
    if abs(num - round(num)) < 0.000001:
        return f"{int(round(num)):,}"
    return f"{num:,.2f}"


def display_percent(value: Any) -> str:
    num = number(value)
    if num is None:
        return ""
    if num <= 1:
        num *= 100
    return f"{num:.1f}%"


def split_path_parts(value: Any) -> list[str]:
    return [part for part in re.split(r"[\\/]+", text(value)) if part]


def short_path(value: Any, keep: int = 4, limit: int = 180) -> str:
    parts = split_path_parts(value)
    if not parts:
        return ""
    label = "\\".join(parts[-keep:])
    return flat(label, limit)


def short_file_list(value: Any, limit: int = 3, keep: int = 4) -> str:
    raw = text(value)
    if not raw:
        return ""
    pieces = [part.strip().strip('"') for part in re.split(r"\s*\|\|\s*|\n+", raw) if part.strip()]
    labels = [short_path(piece, keep=keep) for piece in pieces[:limit]]
    if len(pieces) > limit:
        labels.append(f"외 {len(pieces) - limit}건")
    return "\n".join(labels)


def first_file_path(value: Any) -> str:
    raw = text(value)
    if not raw:
        return ""
    pieces = [part.strip().strip('"') for part in re.split(r"\s*\|\|\s*|\n+", raw) if part.strip()]
    return pieces[0] if pieces else ""


def submitted_location(row: dict[str, Any]) -> str:
    location = short_path(row.get("source_file") or row.get("submitted_file"), keep=4)
    cell = text(row.get("source_cell"))
    if cell:
        location = f"{location}\n셀/범위: {cell}"
    return location


def submitted_location_with_cell(row: dict[str, Any], cell: str | None = None) -> str:
    location = short_path(row.get("source_file") or row.get("submitted_file"), keep=4)
    cell_text = text(cell) or text(row.get("source_cell"))
    if cell_text:
        location = f"{location}\n셀/범위: {cell_text}"
    return location


def split_source_cells(value: Any) -> tuple[str, str]:
    raw = text(value)
    if " vs " in raw:
        left, right = raw.split(" vs ", 1)
        return left.strip(), right.strip()
    return raw, raw


def evidence_bundle(*parts: str) -> str:
    seen: list[str] = []
    for part in parts:
        for line in text(part).splitlines():
            line = line.strip()
            if line and line not in seen:
                seen.append(line)
    return "\n".join(seen)


def hyperlink_target(raw: Any, root: Path, workbook_dir: Path) -> str:
    lines = text(raw).splitlines()
    if not lines:
        return ""
    value = lines[0].strip().strip('"')
    if not value:
        return ""
    value = re.sub(r"^(ALIO 공시자료|제출엑셀|근거자료)\s*:\s*", "", value).strip()
    candidate = Path(value)
    candidates = [candidate]
    if not candidate.is_absolute():
        candidates.extend([root / candidate, root.parent / candidate])
    for path in candidates:
        try:
            if path.exists():
                return os.path.relpath(path.resolve(), workbook_dir.resolve())
        except OSError:
            continue
    return ""


def read_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    result = []
    for row in rows:
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    wb.close()
    return result


def read_body_shortlist(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_alio_files(path: Path) -> dict[str, str]:
    try:
        rows = read_sheet_rows(path, "alio_files")
    except Exception:
        return {}
    return {text(row.get("item")): text(row.get("path")) for row in rows if text(row.get("item"))}


def extract_2024_rules(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    rules: list[dict[str, str]] = []
    for ws in wb.worksheets:
        if not ws.title.startswith(TARGET_PREFIXES):
            continue
        group_starts = []
        for row in range(1, min(ws.max_row, 5) + 1):
            for col in range(1, ws.max_column + 1):
                value = text(ws.cell(row, col).value)
                if "24년도 1차 점검" in value:
                    group_starts.append(col)
        if not group_starts:
            continue
        header_row = None
        for row in range(1, min(ws.max_row, 8) + 1):
            vals = [text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
            if any("지적 유형" in v for v in vals):
                header_row = row
                break
        if header_row is None:
            continue
        for start in sorted(set(group_starts)):
            columns = {"type": start, "opinion": start + 1, "note": start + 2}
            marker_seen = False
            for row in range(header_row + 1, ws.max_row + 1):
                issue = text(ws.cell(row, columns["type"]).value)
                issue_flat = flat(issue, 500)
                if (
                    "2차 점검" in issue_flat
                    or "3차 점검" in issue_flat
                    or "주요 이의신청 유형" in issue_flat
                ):
                    marker_seen = True
                    continue
                if marker_seen:
                    continue
                if (
                    not issue_flat
                    or issue_flat in {" ", "-"}
                    or issue_flat.startswith("1차 점검 _")
                    or issue_flat.startswith("점검항목")
                ):
                    continue
                opinion = text(ws.cell(row, columns["opinion"]).value)
                note = text(ws.cell(row, columns["note"]).value)
                rules.append(
                    {
                        "sheet": ws.title,
                        "issue": issue_flat,
                        "opinion": flat(opinion, 280),
                        "note": flat(note, 120),
                    }
                )
    wb.close()
    return rules


def is_unit_scale_suspect(left: Any, right: Any) -> bool:
    a = number(left)
    b = number(right)
    if a is None or b is None or abs(a) < 1 or abs(b) < 1:
        return False
    ratio = abs(a / b)
    return 990 <= ratio <= 1010 or 0.00099 <= ratio <= 0.00101


def unit_scale_suspect_count(auto_rows: list[dict[str, Any]]) -> int:
    return sum(
        1
        for row in auto_rows
        if "ALIO-제출" in text(row.get("check_type"))
        and is_unit_scale_suspect(row.get("submitted_value"), row.get("alio_value"))
    )


def alio_source(row: dict[str, Any], alio_files: dict[str, str]) -> str:
    path = alio_source_path(row, alio_files)
    return f"ALIO 공시자료: {short_path(path, keep=3)}" if path else "ALIO 공시자료"


def alio_source_path(row: dict[str, Any], alio_files: dict[str, str]) -> str:
    item = text(row.get("item"))
    note = text(row.get("note"))
    key = item
    if item == "6-2":
        key = "6-2-last" if "last" in note else "6-2-first"
    return alio_files.get(key) or alio_files.get(item) or ""


def row_base(
    review_type: str,
    institution: Any,
    item: Any,
    issue: str,
    control_doc: str,
    control_value: str,
    compare_doc: str,
    compare_value: str,
    finding: str,
    evidence_doc: str,
    action: str,
    detail: str = "",
    control_link: Any = "",
    compare_link: Any = "",
    evidence_link: Any = "",
    reviewer: Any = "",
    priority: Any = "",
    source_score: Any = "",
) -> dict[str, str]:
    return {
        "상태": "확인전",
        "판단 결과": "미입력",
        "재확인 필요": "N",
        "항목": text(item),
        "기관": flat(institution, 60),
        "검토유형": review_type,
        "검토 질문": flat(action, 340),
        "발견 요약": flat(issue, 260),
        "대조군 값": multiline(control_value, 220),
        "비교군 값": multiline(compare_value, 220),
        "차이/판정": multiline(finding, 360),
        "확인자료": multiline(evidence_doc, 520),
        "검토 메모": "",
        "대조군 자료": multiline(control_doc, 420),
        "비교군 자료": multiline(compare_doc, 420),
        "상세키/미매칭": flat(detail, 260),
        "자동점수/원천근거": flat(
            " / ".join(
                part
                for part in [
                    f"담당자 {flat(reviewer, 20)}" if text(reviewer) else "",
                    f"우선순위 {flat(priority, 20)}" if text(priority) else "",
                    flat(source_score, 180),
                ]
                if part
            ),
            260,
        ),
        "_대조군 링크": text(control_link),
        "_비교군 링크": text(compare_link),
        "_확인 링크": text(evidence_link) or text(control_link) or text(compare_link),
    }


def alio_number_examples(
    auto_rows: list[dict[str, Any]], alio_files: dict[str, str], limit: int = 3
) -> list[dict[str, str]]:
    candidates = []
    for row in auto_rows:
        if "ALIO-제출" not in text(row.get("check_type")):
            continue
        diff = number(row.get("difference"))
        if diff is None or abs(diff) < 0.000001:
            continue
        candidates.append((abs(diff), row))
    candidates.sort(key=lambda x: x[0], reverse=True)

    result = []
    seen: set[tuple[str, str]] = set()
    for _, row in candidates:
        key = (text(row.get("institution")), text(row.get("item")))
        if key in seen:
            continue
        seen.add(key)
        finding = f"차이 {display_number(row.get('difference'))}"
        action = "ALIO 공시값과 제출엑셀 셀값 중 어느 쪽이 맞는지 확인"
        if is_unit_scale_suspect(row.get("submitted_value"), row.get("alio_value")):
            finding += " / 1,000배 단위차 의심"
            action = "원/천원 단위 보정 후에도 실제 불일치인지 확인"
        submitted_doc = submitted_location(row)
        alio_doc = alio_source(row, alio_files)
        alio_path = alio_source_path(row, alio_files)
        result.append(
            row_base(
                "공시-제출 수치",
                row.get("institution"),
                row.get("item"),
                text(row.get("check_type")),
                submitted_doc,
                f"제출엑셀 값: {display_number(row.get('submitted_value'))}",
                alio_doc,
                f"ALIO 공시값: {display_number(row.get('alio_value'))}",
                finding,
                evidence_bundle(submitted_doc, alio_doc),
                action,
                text(row.get("key")),
                control_link=row.get("source_file"),
                compare_link=alio_path,
            )
        )
        if len(result) >= limit:
            break
    return result


def alio_missing_examples(
    auto_rows: list[dict[str, Any]], alio_files: dict[str, str], limit: int = 30
) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    counts: Counter[tuple[str, str]] = Counter()
    for row in auto_rows:
        if text(row.get("check_type")) != ALIO_MISSING_CHECK_TYPE:
            continue
        key = (text(row.get("institution")), text(row.get("item")))
        counts[key] += 1
        grouped.setdefault(key, row)

    result: list[dict[str, str]] = []
    for key in sorted(grouped, key=lambda k: (k[1], k[0])):
        row = grouped[key]
        count = counts[key]
        submitted_doc = submitted_location(row)
        alio_doc = alio_source(row, alio_files)
        alio_path = alio_source_path(row, alio_files)
        sample_key = text(row.get("key"))
        finding = f"ALIO 키 미매칭 또는 원천 누락 후보 {count:,}건"
        detail = sample_key if count == 1 else f"{sample_key} 외 {count - 1:,}건"
        result.append(
            row_base(
                "공시-ALIO 매칭확인",
                row.get("institution"),
                row.get("item"),
                ALIO_MISSING_CHECK_TYPE,
                submitted_doc,
                f"제출값: {display_number(row.get('submitted_value'))}",
                alio_doc,
                "ALIO 매칭값 미발견",
                finding,
                evidence_bundle(submitted_doc, alio_doc),
                "기관명, 고용형태, 항목명, 연도 키가 ALIO 원천 파일과 맞는지 확인",
                detail,
                control_link=row.get("source_file"),
                compare_link=alio_path,
                source_score=f"원천 자동검토 전체 미매칭 {count:,}건을 기관·항목 단위로 요약",
            )
        )
        if len(result) >= limit:
            break
    return result


def alio_list_examples(
    auto_rows: list[dict[str, Any]], alio_files: dict[str, str], limit: int = 2
) -> list[dict[str, str]]:
    candidates = [
        row
        for row in auto_rows
        if "채용공고" in text(row.get("check_type")) and text(row.get("item")) == "6-2"
    ]
    result = []
    seen: set[str] = set()
    for row in candidates:
        institution = text(row.get("institution"))
        if institution in seen:
            continue
        seen.add(institution)
        check_type = text(row.get("check_type"))
        submitted_doc = submitted_location(row)
        alio_doc = alio_source(row, alio_files)
        alio_path = alio_source_path(row, alio_files)
        title = flat(row.get("key"), 180)
        if check_type.startswith("ALIO"):
            control_doc = alio_doc
            control_value = f"ALIO 채용공고: {title}"
            compare_doc = submitted_doc
            compare_value = "제출엑셀 목록에서 같은 채용공고 미발견"
            finding = "ALIO에는 있으나 제출엑셀 6-2 목록에는 없음"
        else:
            control_doc = submitted_doc
            control_value = f"제출 채용공고: {title}"
            if row.get("submitted_value"):
                control_value += f"\n제출일자/값: {display_number(row.get('submitted_value'))}"
            compare_doc = alio_doc
            compare_value = "ALIO 채용공고 목록에서 미발견"
            finding = "제출엑셀에는 있으나 ALIO 6-2 목록에는 없음"
        action = "ALIO 채용공고 목록과 제출 6-2 목록 중 누락/초과 포함 여부 확인"
        result.append(
            row_base(
                "공시-제출 목록",
                institution,
                row.get("item"),
                check_type,
                control_doc,
                control_value,
                compare_doc,
                compare_value,
                finding,
                evidence_bundle(submitted_doc, alio_doc),
                action,
                text(row.get("source_cell")) or text(row.get("note")),
                control_link=alio_path if check_type.startswith("ALIO") else row.get("source_file"),
                compare_link=row.get("source_file") if check_type.startswith("ALIO") else alio_path,
            )
        )
        if len(result) >= limit:
            break
    return result


def cross_examples(auto_rows: list[dict[str, Any]], limit: int = 2) -> list[dict[str, str]]:
    candidates = []
    for row in auto_rows:
        check_type = text(row.get("check_type"))
        if "불일치" not in check_type or "ALIO-제출" in check_type:
            continue
        diff = number(row.get("difference"))
        if diff is None or abs(diff) < 0.000001:
            continue
        candidates.append((abs(diff), row))
    candidates.sort(key=lambda x: x[0], reverse=True)

    result = []
    for _, row in candidates[:limit]:
        left_cell, right_cell = split_source_cells(row.get("source_cell"))
        control_doc = submitted_location_with_cell(row, left_cell)
        compare_doc = submitted_location_with_cell(row, right_cell)
        result.append(
            row_base(
                "제출엑셀 내부",
                row.get("institution"),
                row.get("item"),
                text(row.get("check_type")),
                control_doc,
                f"값1: {display_number(row.get('submitted_value'))}",
                compare_doc,
                f"값2: {display_number(row.get('alio_value'))}",
                f"차이 {display_number(row.get('difference'))}",
                evidence_bundle(control_doc, compare_doc),
                "두 시트가 같은 기준연도/고용유형/복리후생 항목을 기준으로 작성됐는지 확인",
                text(row.get("key")),
                control_link=row.get("source_file"),
                compare_link=row.get("source_file"),
            )
        )
    return result


def evidence_examples(evidence_rows: list[dict[str, Any]], limit: int = 3) -> list[dict[str, str]]:
    status_rank = {
        "근거자료 없음": 0,
        "숫자 매칭 낮음": 1,
        "6-2 제목 매칭 낮음": 2,
        "부분 매칭": 3,
        "매칭 양호": 9,
    }

    def sort_key(row: dict[str, Any]) -> tuple:
        status = text(row.get("status"))
        numeric_total = int(number(row.get("numeric_submitted_distinct")) or 0)
        title_total = int(number(row.get("title_submitted_distinct")) or 0)
        numeric_coverage = number(row.get("numeric_coverage"))
        title_coverage = number(row.get("title_coverage"))
        return (
            status_rank.get(status, 5),
            -(numeric_total + title_total),
            numeric_coverage if numeric_coverage is not None else 9,
            title_coverage if title_coverage is not None else 9,
        )

    candidates = [
        row
        for row in evidence_rows
        if text(row.get("status")) in {"근거자료 없음", "숫자 매칭 낮음", "6-2 제목 매칭 낮음", "부분 매칭"}
    ]
    candidates.sort(key=sort_key)

    result = []
    seen: set[tuple[str, str, str]] = set()
    for row in candidates:
        key = (text(row.get("status")), text(row.get("institution")), text(row.get("item")))
        if key in seen:
            continue
        seen.add(key)
        status = text(row.get("status"))
        evidence_doc = short_file_list(row.get("evidence_files_sample"), limit=3, keep=5)
        if not evidence_doc:
            evidence_doc = "근거자료 파일 없음"
        evidence_link = first_file_path(row.get("evidence_files_sample"))
        control_bits = [f"제출숫자 {display_number(row.get('numeric_submitted_distinct'))}개"]
        compare_bits = [
            f"근거파일 {display_number(row.get('evidence_file_count'))}건",
            f"숫자매칭 {display_number(row.get('numeric_matched_distinct'))}개",
        ]
        finding_bits = [f"숫자커버 {display_percent(row.get('numeric_coverage'))}"]
        if text(row.get("item")) == "6-2":
            control_bits.append(f"제출제목 {display_number(row.get('title_submitted_distinct'))}개")
            compare_bits.append(f"제목매칭 {display_number(row.get('title_matched_distinct'))}개")
            finding_bits.append(f"제목커버 {display_percent(row.get('title_coverage'))}")
        unmatched = flat(row.get("unmatched_values_sample"), 160)
        if unmatched:
            finding_bits.append(f"미매칭 예시: {unmatched}")
        submitted_doc = short_path(row.get("submitted_file"), keep=4)
        action = "근거자료가 실제 누락됐는지, 다른 폴더명/파일명으로 제출됐는지 확인"
        if status != "근거자료 없음":
            action = "제출 숫자/채용공고 제목이 근거자료에 있는지, 단위/파일명 차이인지 확인"
        result.append(
            row_base(
                "제출-근거자료",
                row.get("institution"),
                row.get("item"),
                status,
                submitted_doc,
                " / ".join(control_bits),
                evidence_doc,
                " / ".join(compare_bits),
                " / ".join(finding_bits),
                evidence_bundle(submitted_doc, evidence_doc),
                action,
                row.get("unmatched_values_sample"),
                control_link=row.get("submitted_file"),
                compare_link=evidence_link,
                evidence_link=evidence_link or row.get("submitted_file"),
            )
        )
        if len(result) >= limit:
            break
    return result


def keyword_summary(reasons: str, max_keywords: int = 8) -> str:
    keywords: list[str] = []
    for part in [p.strip() for p in reasons.split(";") if p.strip()]:
        if "본문 키워드:" not in part:
            continue
        raw = part.split(":", 1)[1]
        for keyword in re.split(r"[,/]", raw):
            keyword = keyword.strip()
            if keyword and keyword not in keywords:
                keywords.append(keyword)
    if not keywords:
        return flat(reasons, 160)
    shown = ", ".join(keywords[:max_keywords])
    if len(keywords) > max_keywords:
        shown += f" 외 {len(keywords) - max_keywords}개"
    return shown


def unfair_disclosure_examples(body_rows: list[dict[str, str]], limit: int = 2) -> list[dict[str, str]]:
    candidates = []
    for row in body_rows:
        reasons = row.get("reasons") or ""
        if KEYWORD_UNFAIR_DISCLOSURE not in reasons:
            continue
        try:
            score = int(float(row.get("score") or 0))
        except Exception:
            score = 0
        candidates.append((score, row))
    candidates.sort(key=lambda x: x[0], reverse=True)

    result = []
    for _, row in candidates[:limit]:
        agency = re.sub(r"^\d{3}\.\s*", "", text(row.get("agency_folder")))
        evidence_doc = short_file_list(row.get("files"), limit=3, keep=5)
        evidence_link = first_file_path(row.get("files"))
        keyword_value = f"점수 {display_number(row.get('score'))} / 근거 {display_number(row.get('source_count'))}건"
        keyword_value += f"\n키워드: {keyword_summary(row.get('reasons') or '')}"
        result.append(
            row_base(
                "불성실공시 키워드",
                agency,
                text(row.get("category")) or "공통/본문",
                "제출/증빙본문에 불성실공시 관련 키워드 포함",
                "2024년 점검유형 원본",
                "불성실공시 내역 포함 여부",
                evidence_doc,
                keyword_value,
                "제출/증빙본문에서 불성실공시 관련 키워드 탐지",
                evidence_doc,
                "원문에 불성실공시 내역을 실제 포함했는지, 단순 안내문구인지 확인",
                flat(row.get("examples"), 220),
                compare_link=evidence_link,
                evidence_link=evidence_link,
                reviewer=row.get("reviewer"),
                priority=row.get("priority"),
                source_score=f"본문점수 {display_number(row.get('score'))} / 근거파일 {display_number(row.get('source_count'))}건",
            )
        )
    return result


def pipeline_audit_text(
    rules: list[dict[str, str]],
    auto_rows: list[dict[str, Any]],
    evidence_rows: list[dict[str, Any]],
    body_rows: list[dict[str, str]],
) -> list[str]:
    auto_types = Counter(text(r.get("check_type")) for r in auto_rows)
    evidence_status = Counter(text(r.get("status")) for r in evidence_rows)
    unfair_count = sum(1 for r in body_rows if KEYWORD_UNFAIR_DISCLOSURE in (r.get("reasons") or ""))
    return [
        f"2024년 1차 유형 파일에서 노무 주요 항목 지적유형 {len(rules)}건을 추출",
        f"공시자료-제출엑셀 비교 후보 {sum(c for k, c in auto_types.items() if 'ALIO' in k)}건",
        f"제출엑셀 내부 교차검증 후보 {auto_types.get('예산상 복리후생비와 1인당 복리후생비 예산총합 불일치', 0) + auto_types.get('상임임원 급여성 복리후생비 합계 불일치', 0)}건",
        f"증빙-제출 매칭 {len(evidence_rows)}개 기관-항목 중 근거자료 없음 {evidence_status.get('근거자료 없음', 0)}건, 숫자 매칭 낮음 {evidence_status.get('숫자 매칭 낮음', 0)}건",
        f"불성실공시 키워드 후보 {unfair_count}건",
        f"공시-제출 수치 중 1,000배 단위차 의심 {unit_scale_suspect_count(auto_rows)}건",
    ]


def coverage_gap_rows() -> list[tuple[str, str, str]]:
    return [
        ("직접 자동검토", "공시자료-제출엑셀 숫자 불일치", "5-1, 10-1, 13-1, 13-5 ALIO 수치 대조"),
        ("직접 자동검토", "6-2 채용공고 목록 누락/초과", "ALIO 채용공고 제목과 제출 목록 비교"),
        ("직접 자동검토", "10-1/13-1, 13-1/13-5 항목간 불일치", "2024년 점검유형 중 복리후생 연계 항목 반영"),
        ("부분 자동검토", "제출엑셀-근거자료 차이", "기관별 근거 폴더 존재, 증빙 엑셀 숫자 커버리지, 6-2 제목 커버리지 확인"),
        ("부분 자동검토", "불성실공시 내역 포함", "본문 키워드로 후보 추출. 실제 포함 여부는 원문 확인 필요"),
        ("보정 필요", "공시-제출 수치 단위 차이", "10-1 등에서 원/천원 1,000배 차이가 상위 후보를 밀어올릴 수 있음"),
        ("수동/추가개발 필요", "지역인재/청년/고졸/직종/임용일 판단", "개인별 로데이터 속성 파싱과 2024 유형별 규칙화 필요"),
        ("수동/추가개발 필요", "6-2 근무분야, 전형단계, 결과확정일, 첨부파일 3종", "공고문/내부결재 본문 구조화 또는 LLM/OCR 후 규칙 적용 필요"),
        ("수동/추가개발 필요", "13-5 평균인원-11-1 상시종업원수 연계", "11-1 파서와 13-5 평균인원 비교 로직 추가 필요"),
    ]


def build_examples(
    auto_rows: list[dict[str, Any]],
    alio_files: dict[str, str],
    evidence_rows: list[dict[str, Any]],
    body_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    examples = (
        alio_number_examples(auto_rows, alio_files, 10)
        + alio_missing_examples(auto_rows, alio_files, 5)
        + alio_list_examples(auto_rows, alio_files, 5)
        + evidence_examples(evidence_rows, 10)
        + cross_examples(auto_rows, 5)
        + unfair_disclosure_examples(body_rows, 5)
    )
    return examples[:30]


def build_detail_rows(
    auto_rows: list[dict[str, Any]],
    alio_files: dict[str, str],
    evidence_rows: list[dict[str, Any]],
    body_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    rows = (
        alio_number_examples(auto_rows, alio_files, 80)
        + alio_missing_examples(auto_rows, alio_files, 240)
        + alio_list_examples(auto_rows, alio_files, 80)
        + evidence_examples(evidence_rows, 140)
        + cross_examples(auto_rows, 120)
        + unfair_disclosure_examples(body_rows, 80)
    )
    return rows[:500]


def build_markdown(
    generated_at: str,
    audit_lines: list[str],
    examples: list[dict[str, str]],
    gaps: list[tuple[str, str, str]],
    xlsx_name: str,
) -> str:
    audit_bullets = "\n".join(f"- {line}" for line in audit_lines)
    cards = []
    for idx, e in enumerate(examples, start=1):
        cards.append(
            "\n".join(
                [
                    f"### {idx}. [{e['검토유형']}] {e['기관']} / {e['항목']}",
                    f"- 발견 요약: {e['발견 요약']}",
                    f"- 대조군: {e['대조군 값']} (`{e['대조군 자료']}`)",
                    f"- 비교군: {e['비교군 값']} (`{e['비교군 자료']}`)",
                    f"- 차이/판정: {e['차이/판정']}",
                    f"- 확인자료: {e['확인자료']}",
                    f"- 검토 질문: {e['검토 질문']}",
                ]
            )
        )
    gap_rows = "\n".join(f"| {a} | {b} | {c} |" for a, b, c in gaps)
    return f"""# 한 장 요약보고서

생성시각: {generated_at}

엑셀판: `{xlsx_name}`

읽는 순서: `대조군 값`과 `비교군 값`을 먼저 비교하고, `확인자료`에 적힌 제출엑셀/공시자료/근거자료를 열어 `검토 질문`에 답합니다.

## 먼저 볼 대표 후보

{chr(10).join(cards)}

## 파이프라인 점검 결과

{audit_bullets}

## 2024년 점검유형 기준 커버리지

| 상태 | 검토유형 | 현재 판단 |
|---|---|---|
{gap_rows}

엑셀판의 `검토시작`을 먼저 보고, `대표후보_바로보기`와 `상세후보_값대조`에서 후보를 확인합니다. 담당자 배분과 상태 관리는 `01_관리자용_전체검토현황.xlsx` 또는 담당자별 큐 파일에서 진행합니다.
"""


def style_report(ws, header_row: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"E{header_row + 1}"
    last_col = get_column_letter(len(REPORT_HEADERS))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

    widths = [6, 12, 14, 12, 10, 28, 18, 36, 36, 24, 24, 34, 42, 34, 46, 46, 30, 30]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for col in ("Q", "R"):
        ws.column_dimensions[col].hidden = True

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=9)
            cell.border = border

    if header_row > 1:
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    type_fills = {
        "공시-제출 수치": "FCE4D6",
        "공시-제출 목록": "FFF2CC",
        "제출-근거자료": "E2F0D9",
        "제출엑셀 내부": "D9EAF7",
        "불성실공시 키워드": "EADCF8",
    }
    status_fills = {
        "수정요청": "F4CCCC",
        "판단보류": "D9D9D9",
        "이상없음": "D9EAD3",
        "제외": "E7E6E6",
    }
    for row in range(header_row + 1, ws.max_row + 1):
        fill = PatternFill("solid", fgColor=type_fills.get(text(ws.cell(row, 7).value), "FFFFFF"))
        ws.cell(row, 7).fill = fill
        ws.cell(row, 12).fill = fill
        status = text(ws.cell(row, 2).value)
        if status in status_fills:
            ws.cell(row, 2).fill = PatternFill("solid", fgColor=status_fills[status])
        for col in (2, 3, 4, 14):
            ws.cell(row, col).protection = Protection(locked=False)

    status_dv = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=False)
    ws.add_data_validation(status_dv)
    status_dv.add(f"B{header_row + 1}:B{max(ws.max_row, header_row + 1)}")
    judgment_dv = DataValidation(type="list", formula1=f'"{JUDGMENT_VALUES}"', allow_blank=False)
    ws.add_data_validation(judgment_dv)
    judgment_dv.add(f"C{header_row + 1}:C{max(ws.max_row, header_row + 1)}")
    recheck_dv = DataValidation(type="list", formula1=f'"{RECHECK_VALUES}"', allow_blank=False)
    ws.add_data_validation(recheck_dv)
    recheck_dv.add(f"D{header_row + 1}:D{max(ws.max_row, header_row + 1)}")
    ws.protection.sheet = True
    ws.protection.enable()
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def add_summary_sheet(wb: Workbook, audit_lines: list[str], gaps: list[tuple[str, str, str]]) -> None:
    ws = wb.create_sheet("유형별_요약")
    ws.append(["구분", "내용"])
    for line in audit_lines:
        ws.append(["점검결과", line])
    ws.append([])
    ws.append(["상태", "검토유형", "현재 판단"])
    for row in gaps:
        ws.append(list(row))

    for col, width in enumerate([20, 80, 90], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_num in (1, len(audit_lines) + 3):
        for cell in ws[row_num]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"


def save_workbook_with_fallback(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        stale_fallback = path.with_name(f"{path.stem}_가독성개선본{path.suffix}")
        if stale_fallback.exists():
            try:
                stale_fallback.unlink()
            except PermissionError:
                pass
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_가독성개선본{path.suffix}")
        try:
            wb.save(fallback)
        except PermissionError:
            fallback = path.with_name(
                f"{path.stem}_가독성개선본_{dt.datetime.now().strftime('%H%M%S')}{path.suffix}"
            )
            wb.save(fallback)
        print(f"warning: {path.name} is open or locked; wrote {fallback.name} instead")
        return fallback


def append_report_rows(ws, rows: list[dict[str, str]], root: Path, workbook_dir: Path) -> None:
    link_columns = {
        "대조군 자료": "_대조군 링크",
        "비교군 자료": "_비교군 링크",
        "확인자료": "_확인 링크",
    }
    for idx, row_data in enumerate(rows, start=1):
        ws.append([idx] + [row_data[h] for h in REPORT_HEADERS[1:]])
        row_num = ws.max_row
        for header, key in link_columns.items():
            target = hyperlink_target(row_data.get(key), root, workbook_dir)
            if not target:
                continue
            col_num = REPORT_HEADERS.index(header) + 1
            cell = ws.cell(row_num, col_num)
            cell.hyperlink = target
            cell.style = "Hyperlink"


def add_dashboard_sheet(
    wb: Workbook,
    generated_at: str,
    queue_records: list[dict[str, object]],
    examples: list[dict[str, str]],
    detail_rows: list[dict[str, str]],
    report_name: str,
) -> None:
    ws = wb.active
    ws.title = "검토시작"
    first_total = sum(1 for rec in queue_records if rec.get("bucket") == "1차")
    second_total = sum(1 for rec in queue_records if rec.get("bucket") == "2차")
    high_total = sum(1 for rec in queue_records if rec.get("priority") == "HIGH")
    counts = counts_by_reviewer(queue_records)
    type_counts = Counter(row.get("검토유형", "") for row in detail_rows)

    rows = [
        ["검토 시작 대시보드", ""],
        ["1. 이 파일에서 시작하세요.", "대표후보_바로보기 → 상세후보_값대조 순서로 봅니다."],
        ["2. 값부터 비교하세요.", "대조군 값 / 비교군 값 / 차이·판정 / 확인자료를 한 행에서 확인합니다."],
        ["3. 이것만 입력하세요.", "상태 / 판단 결과 / 재확인 필요 / 검토 메모만 입력합니다."],
        ["4. 원천은 처음에 열지 않습니다.", "CSV, OCR, 로그, 백업은 근거가 부족할 때만 확인합니다."],
        ["생성시각", generated_at],
        ["오늘 먼저 볼 파일", report_name],
        ["처음 열 시트", "대표후보_바로보기"],
        ["다음 시트", "상세후보_값대조"],
        ["상태 기록", "상태 / 판단 결과 / 재확인 필요 / 검토 메모만 입력"],
        [],
        ["오늘 볼 것", "건수"],
        ["대표 후보", len(examples)],
        ["상세 값대조 후보", len(detail_rows)],
        ["담당자 큐 1차", first_total],
        ["담당자 큐 2차", second_total],
        ["HIGH 후보", high_total],
        [],
        ["작업 순서", "내용"],
        ["1", "대표후보_바로보기에서 대조군 값과 비교군 값, 차이/판정, 확인자료를 확인"],
        ["2", "상세후보_값대조에서 같은 형식으로 추가 후보 확인"],
        ["3", "담당자별 큐 파일에서 상태와 판단 결과 기록"],
        ["4", "근거가 부족한 경우에만 99_원천산출물_참고용 또는 상세 CSV 확인"],
        [],
        ["열지 말 파일", "기본적으로 직접 열지 않음"],
        ["원천 자동검토", "99_원천산출물_참고용"],
        ["대용량 OCR/로그/백업", "재현 또는 점검이 필요할 때만 확인"],
        [],
        ["상태값", STATUS_VALUES.replace(",", " / ")],
        ["판단 결과", JUDGMENT_VALUES.replace(",", " / ")],
    ]
    for row in rows:
        ws.append(row)

    start = ws.max_row + 2
    ws.cell(start, 1, "담당자")
    ws.cell(start, 2, "1차")
    ws.cell(start, 3, "2차")
    ws.cell(start, 4, "전체")
    for row_idx, reviewer in enumerate(REVIEWERS, start=start + 1):
        ws.cell(row_idx, 1, reviewer)
        ws.cell(row_idx, 2, counts[reviewer]["1차"])
        ws.cell(row_idx, 3, counts[reviewer]["2차"])
        ws.cell(row_idx, 4, counts[reviewer]["전체"])

    type_start = start
    ws.cell(type_start, 6, "검토유형")
    ws.cell(type_start, 7, "상세후보")
    for offset, (review_type, count) in enumerate(type_counts.most_common(), start=1):
        ws.cell(type_start + offset, 6, review_type)
        ws.cell(type_start + offset, 7, count)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for col, width in {"A": 22, "B": 72, "C": 12, "D": 12, "F": 24, "G": 12}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)
    for row_num in (1, 12, 19, 25, 29, start, type_start):
        for cell in ws[row_num]:
            if cell.value:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    for row_num in range(2, 6):
        ws.cell(row_num, 1).font = Font(size=12, bold=True, color="9C0006")
        ws.cell(row_num, 1).fill = PatternFill("solid", fgColor="FCE4D6")
        ws.cell(row_num, 2).fill = PatternFill("solid", fgColor="FFF2CC")


def build_workbook(
    root: Path,
    path: Path,
    generated_at: str,
    audit_lines: list[str],
    queue_records: list[dict[str, object]],
    examples: list[dict[str, str]],
    detail_rows: list[dict[str, str]],
    gaps: list[tuple[str, str, str]],
    rules: list[dict[str, str]],
) -> Path:
    wb = Workbook()
    add_dashboard_sheet(wb, generated_at, queue_records, examples, detail_rows, path.name)

    ws = wb.create_sheet("대표후보_바로보기")
    ws.append(["한 장 요약보고서"] + [""] * (len(REPORT_HEADERS) - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_HEADERS))
    ws.append(["생성시각", generated_at])
    ws.append(["목적", "대조군 값과 비교군 값, 차이, 확인자료를 한 줄에서 확인"])
    ws.append(["읽는법", "대조군 값과 비교군 값을 먼저 비교하고, 확인자료를 열어 판정 질문에 답한 뒤 상태를 바꿉니다."])
    ws.append([])
    header_row = ws.max_row + 1
    ws.append(REPORT_HEADERS)
    append_report_rows(ws, examples, root, path.parent)
    style_report(ws, header_row)

    ws_detail = wb.create_sheet("상세후보_값대조")
    ws_detail.append(REPORT_HEADERS)
    append_report_rows(ws_detail, detail_rows, root, path.parent)
    style_report(ws_detail, 1)

    add_summary_sheet(wb, audit_lines, gaps)

    ws2 = wb.create_sheet("2024유형_추출")
    ws2.append(["시트", "지적유형", "정평의견", "비고"])
    for r in rules:
        ws2.append([r["sheet"], r["issue"], r["opinion"], r["note"]])
    for col, width in enumerate([28, 80, 60, 24], start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    ws2.sheet_state = "hidden"

    return save_workbook_with_fallback(wb, path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    args = parser.parse_args()

    root = args.root.resolve()
    out_dir = find_dir(root, "30_")
    core_dir = find_core_dir(out_dir)
    auto_dir = find_auto_dir(out_dir)
    reference_dirs = [core_dir, core_dir / "99_원천산출물_참고용"]
    type_workbook = [
        p for p in root.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx" and "24" in p.name
    ][0]
    auto_path = latest_file(auto_dir, "disclosure_auto_review_", ".xlsx")
    evidence_path = latest_file(auto_dir, "evidence_match_review_", ".xlsx")
    body_csv = latest_file_any(reference_dirs, "06_", ".csv")
    number_csv = latest_file_any(reference_dirs, "02_", ".csv")

    rules = extract_2024_rules(type_workbook)
    auto_rows = read_sheet_rows(auto_path, "findings")
    alio_files = read_alio_files(auto_path)
    evidence_rows = read_sheet_rows(evidence_path, "evidence_match")
    body_rows = read_body_shortlist(body_csv)
    number_rows = read_body_shortlist(number_csv)
    queue_records = build_records(body_rows, number_rows)
    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    audit_lines = pipeline_audit_text(rules, auto_rows, evidence_rows, body_rows)
    gaps = coverage_gap_rows()
    examples = build_examples(auto_rows, alio_files, evidence_rows, body_rows)
    detail_rows = build_detail_rows(auto_rows, alio_files, evidence_rows, body_rows)

    md_path = core_dir / "00_한장_요약보고서.md"
    xlsx_path = core_dir / PRIMARY_REPORT_XLSX
    written_xlsx_path = build_workbook(root, xlsx_path, generated_at, audit_lines, queue_records, examples, detail_rows, gaps, rules)
    md = build_markdown(generated_at, audit_lines, examples, gaps, written_xlsx_path.name)
    md_path.write_text(md, encoding="utf-8-sig")
    for stale in [
        core_dir / LEGACY_REPORT_XLSX,
        core_dir / "00_한장_요약보고서_가독성개선본.xlsx",
    ]:
        if stale.exists() and stale != written_xlsx_path:
            try:
                stale.unlink()
            except PermissionError:
                pass

    print(f"one-page markdown: {md_path}")
    print(f"one-page workbook: {written_xlsx_path}")
    print(f"examples: {len(examples)}")
    print(f"detail rows: {len(detail_rows)}")
    print(f"2024 rules: {len(rules)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
