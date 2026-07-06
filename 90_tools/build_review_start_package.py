# -*- coding: utf-8 -*-
r"""
Build a reviewer-friendly start package from existing review outputs.

The original review outputs intentionally keep detailed CSV/XLSX evidence.
This script adds a smaller human workflow layer:
- one workbook with reviewer queues;
- one top-level start document;
- short README files in each reviewer delivery folder.

Usage:
    python .\90_tools\build_review_start_package.py --root .
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import shutil
import zipfile
from collections import Counter, OrderedDict, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def reviewer_defaults() -> list[str]:
    raw = os.getenv("DISCLOSURE_REVIEWERS", "A,B,C,D")
    reviewers = [part.strip() for part in raw.split(",") if part.strip()]
    return reviewers or ["A", "B", "C", "D"]


REVIEWERS = reviewer_defaults()


def include_reviewers(values: Iterable[object]) -> None:
    for value in values:
        reviewer = str(value or "").strip()
        if reviewer and reviewer not in REVIEWERS:
            REVIEWERS.append(reviewer)


BODY_FIRST_PASS_SCORE = 1000
BODY_SECOND_PASS_SCORE = 500
STATUS_VALUES = "확인전,이상없음,수정요청,판단보류,제외"
PRIMARY_REPORT_XLSX = "00_여기부터_검토시작.xlsx"
ADMIN_QUEUE_XLSX = "01_관리자용_전체검토현황.xlsx"


def preferred_report_xlsx(core_dir: Path) -> Path:
    primary_stem = Path(PRIMARY_REPORT_XLSX).stem
    candidates = []
    for pattern in [
        PRIMARY_REPORT_XLSX,
        f"{primary_stem}_가독성개선본*.xlsx",
        "00_한장_요약보고서.xlsx",
        "00_한장_요약보고서_가독성개선본*.xlsx",
    ]:
        candidates.extend(core_dir.glob(pattern))
    candidates = [path for path in candidates if path.exists() and path.is_file()]
    if candidates:
        return max(candidates, key=lambda path: path.stat().st_mtime)
    return core_dir / PRIMARY_REPORT_XLSX


def preferred_xlsx(core_dir: Path, filename: str) -> Path:
    path = core_dir / filename
    candidates = [path] if path.exists() else []
    candidates.extend(core_dir.glob(f"{path.stem}_가독성개선본*.xlsx"))
    candidates = [candidate for candidate in candidates if candidate.exists() and candidate.is_file()]
    if candidates:
        return max(candidates, key=lambda candidate: candidate.stat().st_mtime)
    return path


def find_child(
    parent: Path,
    prefix: str,
    suffix: str | None = None,
    *,
    want_dir: bool | None = None,
) -> Path:
    matches = [
        p
        for p in parent.iterdir()
        if p.name.startswith(prefix) and (suffix is None or p.name.endswith(suffix))
        and (want_dir is None or p.is_dir() == want_dir)
    ]
    if not matches:
        raise FileNotFoundError(f"Cannot find {prefix}* under {parent}")
    return sorted(matches, key=lambda p: p.name)[0]


def find_core_dir(out_dir: Path) -> Path:
    matches = [
        p
        for p in out_dir.iterdir()
        if p.is_dir() and p.name.startswith("00_") and "핵심산출물" in p.name
    ]
    if matches:
        return sorted(matches, key=lambda p: p.name)[0]
    legacy = [
        p
        for p in out_dir.iterdir()
        if p.is_dir() and p.name.startswith("00_") and "먼저보기" not in p.name
    ]
    if legacy:
        return sorted(legacy, key=lambda p: p.name)[0]
    raise FileNotFoundError(f"Cannot find core output directory under {out_dir}")


def find_reference_file(core_dir: Path, prefix: str, suffix: str) -> Path:
    parents = [core_dir, core_dir / "99_원천산출물_참고용"]
    matches: list[Path] = []
    for parent in parents:
        if not parent.exists():
            continue
        matches.extend(
            p
            for p in parent.iterdir()
            if p.is_file() and p.name.startswith(prefix) and p.name.endswith(suffix)
        )
    if not matches:
        raise FileNotFoundError(f"Cannot find {prefix}*{suffix} under core/reference folders")
    return sorted(matches, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def rel_reference(core_dir: Path, root: Path, prefix: str, suffix: str = ".csv") -> str:
    try:
        return str(find_reference_file(core_dir, prefix, suffix).relative_to(root))
    except FileNotFoundError:
        return str((core_dir / "99_원천산출물_참고용" / f"{prefix}...{suffix}").relative_to(root))


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8-sig")


def copy_with_fallback(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(source, target)
        for stale in target.parent.glob(f"{target.stem}_가독성개선본*.xlsx"):
            try:
                stale.unlink()
            except PermissionError:
                pass
        return target
    except PermissionError:
        fallback = target.with_name(f"{target.stem}_가독성개선본{target.suffix}")
        try:
            shutil.copy2(source, fallback)
        except PermissionError:
            fallback = target.with_name(
                f"{target.stem}_가독성개선본_{dt.datetime.now().strftime('%H%M%S')}{target.suffix}"
            )
            shutil.copy2(source, fallback)
        print(f"warning: {target.name} is open or locked; copied {fallback.name} instead")
        return fallback


def as_int(value: object) -> int:
    try:
        return int(float(str(value).replace(",", "").strip()))
    except Exception:
        return 0


def clean_space(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def split_semicolon(value: str) -> list[str]:
    return [part.strip() for part in str(value or "").split(";") if part.strip()]


def agency_name_from_folder(value: str) -> str:
    name = clean_space(value)
    return re.sub(r"^\d{3}\.\s*", "", name)


def short_path_label(value: str, keep: int = 4) -> str:
    parts = [p for p in re.split(r"[\\/]+", value or "") if p]
    if not parts:
        return ""
    return "\\".join(parts[-keep:])


def short_file_list(value: str, limit: int = 3) -> str:
    files = [part.strip().strip('"') for part in str(value or "").split("||") if part.strip()]
    labels = [short_path_label(file, keep=4) for file in files[:limit]]
    if len(files) > limit:
        labels.append(f"외 {len(files) - limit}건")
    return "\n".join(labels)


def compact_reason(value: str, max_keywords: int = 8) -> str:
    parts = split_semicolon(value)
    keywords: list[str] = []
    plain: list[str] = []

    for part in parts:
        if part.startswith("본문 키워드:"):
            raw = part.split(":", 1)[1]
            for keyword in re.split(r"[,/]", raw):
                keyword = keyword.strip()
                if keyword and keyword not in keywords:
                    keywords.append(keyword)
        elif part not in plain:
            plain.append(part)

    if keywords:
        shown = ", ".join(keywords[:max_keywords])
        if len(keywords) > max_keywords:
            shown += f" 외 {len(keywords) - max_keywords}개"
        plain.append(f"본문 키워드({shown})")

    return "; ".join(plain[:8])


def merge_priority(left: str, right: str) -> str:
    rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "": 9}
    return min((left or "", right or ""), key=lambda v: rank.get(v, 9))


def queue_bucket(rec: dict[str, object]) -> str:
    if rec.get("has_number"):
        return "1차"
    if as_int(rec.get("body_score")) >= BODY_FIRST_PASS_SCORE:
        return "1차"
    if as_int(rec.get("body_score")) >= BODY_SECOND_PASS_SCORE:
        return "2차"
    return "참고"


def recommended_action(rec: dict[str, object]) -> str:
    has_body = bool(rec.get("has_body"))
    has_number = bool(rec.get("has_number"))
    if has_body and has_number:
        return "한장 요약의 대조군/비교군 값 차이와 증빙본문 징후를 같은 기관/항목에서 함께 확인"
    if has_number:
        return "한장 요약 또는 숫자대조 원본에서 대조군 값, 비교군 값, 차이, 확인자료 확인"
    return "증빙 예시 파일에서 키워드, 누락, 오류, 보호/추출실패 여부 확인"


def build_records(body_rows: list[dict[str, str]], number_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    records: OrderedDict[tuple[str, str, str], dict[str, object]] = OrderedDict()

    def get_rec(reviewer: str, check_no: str, category: str) -> dict[str, object]:
        key = (reviewer, check_no, category)
        if key not in records:
            records[key] = {
                "status": "확인전",
                "reviewer": reviewer,
                "check_no": check_no,
                "agency": "",
                "category": category,
                "priority": "",
                "sources": [],
                "body_score": 0,
                "number_score": 0,
                "number_diff_count": 0,
                "source_count": 0,
                "reasons": [],
                "examples": [],
                "files": [],
                "has_body": False,
                "has_number": False,
            }
        return records[key]

    for row in body_rows:
        reviewer = clean_space(row.get("reviewer"))
        rec = get_rec(reviewer, clean_space(row.get("check_no")), clean_space(row.get("category")))
        rec["agency"] = rec["agency"] or agency_name_from_folder(row.get("agency_folder", ""))
        rec["priority"] = merge_priority(str(rec.get("priority", "")), row.get("priority", ""))
        rec["body_score"] = max(as_int(rec.get("body_score")), as_int(row.get("score")))
        rec["source_count"] = max(as_int(rec.get("source_count")), as_int(row.get("source_count")))
        rec["has_body"] = True
        if "본문/OCR" not in rec["sources"]:
            rec["sources"].append("본문/OCR")
        reason = compact_reason(row.get("reasons", ""))
        if reason:
            rec["reasons"].append(reason)
        if row.get("examples"):
            rec["examples"].append(clean_space(row.get("examples")))
        if row.get("files"):
            rec["files"].append(short_file_list(row.get("files", "")))

    for row in number_rows:
        reviewer = clean_space(row.get("reviewer"))
        rec = get_rec(reviewer, clean_space(row.get("check_no")), clean_space(row.get("category")))
        rec["agency"] = rec["agency"] or clean_space(row.get("agency_name"))
        rec["priority"] = merge_priority(str(rec.get("priority", "")), row.get("priority", ""))
        rec["number_score"] = max(as_int(rec.get("number_score")), as_int(row.get("score")))
        rec["number_diff_count"] = max(
            as_int(rec.get("number_diff_count")), as_int(row.get("number_diff_count"))
        )
        rec["has_number"] = True
        if "숫자대조" not in rec["sources"]:
            rec["sources"].append("숫자대조")
        reason = compact_reason(row.get("reasons", ""))
        if row.get("number_diff_count"):
            reason = f"{reason}; 숫자 차이 후보 {row.get('number_diff_count')}건".strip("; ")
        if reason:
            rec["reasons"].append(reason)
        if row.get("number_diff_examples"):
            rec["examples"].append(clean_space(row.get("number_diff_examples")))
        if row.get("files"):
            rec["files"].append(short_file_list(row.get("files", "")))

    output: list[dict[str, object]] = []
    for rec in records.values():
        body_score = as_int(rec.get("body_score"))
        number_score = as_int(rec.get("number_score"))
        diff_count = as_int(rec.get("number_diff_count"))
        rec["bucket"] = queue_bucket(rec)
        rec["total_score"] = body_score + number_score + min(diff_count, 500)
        rec["source_type"] = " + ".join(rec.get("sources", []))
        rec["action"] = recommended_action(rec)
        rec["reason_summary"] = "\n".join(dict.fromkeys(rec.get("reasons", [])))
        rec["evidence_hint"] = "\n".join(dict.fromkeys([x for x in rec.get("files", []) if x]))
        rec["detail_source"] = detail_source_text(rec)
        output.append(rec)

    include_reviewers(rec.get("reviewer") for rec in output)

    bucket_rank = {"1차": 0, "2차": 1, "참고": 2}
    priority_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "": 9}
    return sorted(
        output,
        key=lambda r: (
            bucket_rank.get(str(r.get("bucket")), 9),
            REVIEWERS.index(str(r.get("reviewer"))) if r.get("reviewer") in REVIEWERS else 9,
            priority_rank.get(str(r.get("priority")), 9),
            -as_int(r.get("total_score")),
            str(r.get("check_no")),
            str(r.get("category")),
        ),
    )


def detail_source_text(rec: dict[str, object]) -> str:
    sources = []
    reviewer = rec.get("reviewer") or "<담당자>"
    if rec.get("has_body"):
        sources.append(
            f"40_전달패키지\\01_개인별_직첨부\\{reviewer}\\03_검토숏리스트\\본문검토_숏리스트.csv"
        )
    if rec.get("has_number"):
        sources.append(
            f"40_전달패키지\\01_개인별_직첨부\\{reviewer}\\03_검토숏리스트\\숫자대조_숏리스트.csv"
        )
    return "\n".join(sources)


def record_to_row(rec: dict[str, object]) -> list[object]:
    return [
        rec.get("status", "확인전"),
        "미입력",
        "N",
        "",
        rec.get("bucket", ""),
        rec.get("reviewer", ""),
        rec.get("priority", ""),
        rec.get("total_score", 0),
        rec.get("check_no", ""),
        rec.get("agency", ""),
        rec.get("category", ""),
        rec.get("source_type", ""),
        rec.get("action", ""),
        rec.get("reason_summary", ""),
        rec.get("body_score", 0),
        rec.get("number_diff_count", 0),
        rec.get("source_count", 0),
        rec.get("evidence_hint", ""),
        rec.get("detail_source", ""),
    ]


def add_table_sheet(wb: Workbook, title: str, rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(title)
    headers = [
        "상태",
        "판단 결과",
        "재확인 필요",
        "검토 메모",
        "구분",
        "담당자",
        "우선순위",
        "정렬점수",
        "기관번호",
        "기관명",
        "항목",
        "확인유형",
        "권장 확인",
        "이유 요약",
        "본문점수",
        "숫자차이건수",
        "본문근거건수",
        "증빙 예시",
        "세부 원본 위치",
    ]
    ws.append(headers)
    for rec in rows:
        ws.append(record_to_row(rec))
    style_table(ws)
    add_status_dropdown(ws, 2, max(ws.max_row, 2))


def style_table(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.column <= 4:
                cell.protection = Protection(locked=False)
    widths = {
        "A": 12,
        "B": 14,
        "C": 12,
        "D": 34,
        "E": 8,
        "F": 8,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 34,
        "K": 10,
        "L": 18,
        "M": 38,
        "N": 46,
        "O": 10,
        "P": 12,
        "Q": 12,
        "R": 52,
        "S": 58,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in ("H", "O", "P", "Q"):
        ws.column_dimensions[col].hidden = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.conditional_formatting.add(
        f"E2:E{max(ws.max_row, 2)}",
        CellIsRule(operator="equal", formula=['"1차"'], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    ws.conditional_formatting.add(
        f"G2:G{max(ws.max_row, 2)}",
        CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill("solid", fgColor="F8CBAD")),
    )
    ws.protection.sheet = True
    ws.protection.enable()
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def add_status_dropdown(ws, start_row: int, end_row: int) -> None:
    dv = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"A{start_row}:A{end_row}")
    judgment = DataValidation(type="list", formula1='"미입력,수치수정,증빙보완,설명보완,제외,이상없음"', allow_blank=False)
    ws.add_data_validation(judgment)
    judgment.add(f"B{start_row}:B{end_row}")
    recheck = DataValidation(type="list", formula1='"N,Y"', allow_blank=False)
    ws.add_data_validation(recheck)
    recheck.add(f"C{start_row}:C{end_row}")


def add_intro_sheet(
    wb: Workbook,
    records: list[dict[str, object]],
    generated_at: str,
    report_xlsx_name: str,
) -> None:
    ws = wb.active
    ws.title = "읽는법"
    ws["A1"] = "검토 시작용 요약"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"생성시각: {generated_at}"
    ws["A4"] = "사용 순서"
    ws["A4"].font = Font(bold=True)
    steps = [
        ("1", f"`{report_xlsx_name}`의 `대표후보_바로보기`에서 대조군 값, 비교군 값, 차이, 확인자료를 먼저 봅니다."),
        ("2", f"같은 파일의 `상세후보_값대조`에서 같은 형식으로 추가 후보를 이어서 확인합니다."),
        ("3", "담당자별 시트 또는 `1차_검토큐`에서 상태를 `이상없음`, `수정요청`, `판단보류` 등으로 바꿉니다."),
        ("4", "`세부 원본 위치`는 판단 근거가 부족할 때만 엽니다."),
        ("5", "`2차_후보`는 1차 확인이 끝난 뒤 시간이 남을 때 봅니다."),
    ]
    for idx, (no, text) in enumerate(steps, start=5):
        ws[f"A{idx}"] = no
        ws[f"B{idx}"] = text

    ws["A11"] = "담당자별 건수"
    ws["A11"].font = Font(bold=True)
    headers = ["담당자", "1차", "2차", "전체"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=12, column=col, value=header)
    by_reviewer: dict[str, Counter[str]] = defaultdict(Counter)
    for rec in records:
        by_reviewer[str(rec.get("reviewer"))][str(rec.get("bucket"))] += 1
        by_reviewer[str(rec.get("reviewer"))]["전체"] += 1
    for row_idx, reviewer in enumerate(REVIEWERS, start=13):
        ws.cell(row=row_idx, column=1, value=reviewer)
        ws.cell(row=row_idx, column=2, value=by_reviewer[reviewer]["1차"])
        ws.cell(row=row_idx, column=3, value=by_reviewer[reviewer]["2차"])
        ws.cell(row=row_idx, column=4, value=by_reviewer[reviewer]["전체"])

    ws["A20"] = "1차 기준"
    ws["B20"] = f"본문검토 점수 {BODY_FIRST_PASS_SCORE} 이상 또는 숫자대조 후보 전체"
    ws["A21"] = "2차 기준"
    ws["B21"] = f"본문검토 점수 {BODY_SECOND_PASS_SCORE} 이상 {BODY_FIRST_PASS_SCORE - 1} 이하"
    ws["A22"] = "상태값"
    ws["B22"] = STATUS_VALUES.replace(",", " / ")

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = [14, 70, 12, 12][col - 1]
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[12]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)


def save_workbook_with_fallback(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        for stale in path.parent.glob(f"{path.stem}_가독성개선본*.xlsx"):
            try:
                stale.unlink()
            except PermissionError:
                pass
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_가독성개선본{path.suffix}")
        try:
            wb.save(fallback)
        except PermissionError:
            fallback = path.with_name(
                f"{path.stem}_가독성개선본_{dt.datetime.now().strftime('%H%M%S')}{path.suffix}"
            )
            wb.save(fallback)
        print(f"warning: {path.name} is open or locked; wrote {fallback.name} instead")
        return fallback


def build_workbook(path: Path, records: list[dict[str, object]], generated_at: str) -> Path:
    wb = Workbook()
    add_intro_sheet(wb, records, generated_at, preferred_report_xlsx(path.parent).name)

    first_pass = [rec for rec in records if rec.get("bucket") == "1차"]
    second_pass = [rec for rec in records if rec.get("bucket") == "2차"]
    add_summary_sheet(wb, records)
    add_table_sheet(wb, "1차_검토큐", first_pass)
    add_table_sheet(wb, "2차_후보", second_pass)
    for reviewer in REVIEWERS:
        add_table_sheet(wb, f"{reviewer}_1차", [rec for rec in first_pass if rec.get("reviewer") == reviewer])
    add_table_sheet(wb, "전체후보", records)
    return save_workbook_with_fallback(wb, path)


def build_reviewer_workbook(path: Path, reviewer: str, records: list[dict[str, object]], generated_at: str) -> Path:
    subset = [rec for rec in records if rec.get("reviewer") == reviewer]
    first_pass = [rec for rec in subset if rec.get("bucket") == "1차"]
    second_pass = [rec for rec in subset if rec.get("bucket") == "2차"]
    wb = Workbook()
    add_intro_sheet(wb, subset, generated_at, preferred_report_xlsx(path.parent).name)
    add_table_sheet(wb, "검토시작_1차", first_pass)
    add_table_sheet(wb, "2차_후보", second_pass)
    add_table_sheet(wb, "전체후보", subset)
    return save_workbook_with_fallback(wb, path)


def add_summary_sheet(wb: Workbook, records: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("담당자_요약")
    headers = ["담당자", "1차", "2차", "전체", "숫자대조 포함", "본문/OCR 포함", "1차 상위 기관"]
    ws.append(headers)
    for reviewer in REVIEWERS:
        subset = [rec for rec in records if rec.get("reviewer") == reviewer]
        first = [rec for rec in subset if rec.get("bucket") == "1차"]
        second = [rec for rec in subset if rec.get("bucket") == "2차"]
        top_agencies = ", ".join(
            f"{rec.get('check_no')} {rec.get('agency')}" for rec in first[:5]
        )
        ws.append(
            [
                reviewer,
                len(first),
                len(second),
                len(subset),
                sum(1 for rec in subset if rec.get("has_number")),
                sum(1 for rec in subset if rec.get("has_body")),
                top_agencies,
            ]
        )
    style_simple_sheet(ws)


def style_simple_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [10, 10, 10, 10, 14, 14, 90]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def counts_by_reviewer(records: list[dict[str, object]]) -> dict[str, Counter[str]]:
    result: dict[str, Counter[str]] = defaultdict(Counter)
    for rec in records:
        reviewer = str(rec.get("reviewer"))
        result[reviewer][str(rec.get("bucket"))] += 1
        result[reviewer]["전체"] += 1
    return result


def build_start_markdown(
    root: Path,
    core_dir: Path,
    queue_path: Path,
    records: list[dict[str, object]],
    generated_at: str,
) -> str:
    rel_queue = queue_path.relative_to(root)
    counts = counts_by_reviewer(records)
    first_total = sum(1 for rec in records if rec.get("bucket") == "1차")
    second_total = sum(1 for rec in records if rec.get("bucket") == "2차")
    all_total = len(records)
    reviewer_rows = "\n".join(
        f"| {reviewer} | {counts[reviewer]['1차']} | {counts[reviewer]['2차']} | {counts[reviewer]['전체']} |"
        for reviewer in REVIEWERS
    )
    rel_report = core_dir.relative_to(root) / "00_한장_요약보고서.md"
    rel_report_xlsx = preferred_report_xlsx(core_dir).relative_to(root)
    reviewer_files = ", ".join(f"`{core_dir.relative_to(root) / f'01_{reviewer}_검토큐.xlsx'}`" for reviewer in REVIEWERS)
    start_dir = core_dir.parent / "00_00_먼저보기"
    rel_start_readme = start_dir.relative_to(root) / "00_읽는순서.md"
    filtered_admin = start_dir.relative_to(root) / "01_관리자용_전체검토현황_필터링.xlsx"
    filtered_reviewer_files = ", ".join(
        f"`{start_dir.relative_to(root) / f'01_{reviewer}_검토큐_필터링.xlsx'}`" for reviewer in REVIEWERS
    )
    keyword_reference = rel_reference(core_dir, root, "09_", ".csv")
    return f"""
# 검토 시작 안내

생성시각: {generated_at}

이번 검토는 파일을 많이 열지 말고 `30_검토산출물\\00_00_먼저보기`의 필터링 체크리스트에서 시작합니다.

## 먼저 열 파일

1. `{rel_start_readme}`
   - 담당자별로 어떤 필터링 큐를 먼저 열지 확인합니다.

2. 담당자별 필터링 큐: {filtered_reviewer_files}
   - 담당자는 자기 파일의 `검토시작_1차` 시트만 먼저 봅니다.
   - 검토 결과는 앞쪽 `상태`, `판단 결과`, `재확인 필요`, `검토 메모`에 기록합니다.

3. `{filtered_admin}`
   - 관리자는 이 파일에서 전체 배분, 1차/2차/제외 현황을 봅니다.

4. `{rel_report_xlsx}`
   - 필터링 큐의 근거를 더 추적해야 할 때만 엽니다.
   - `대표후보_바로보기`와 `상세후보_값대조`는 `대조군 값`, `비교군 값`, `차이/판정`, `확인자료`, `검토 질문` 순서로 읽습니다.

5. 원본 담당자별 큐: {reviewer_files}
   - 필터링 전 후보 전체를 확인할 때만 엽니다.

6. `{rel_queue}`
   - 필터링 전 관리자용 전체 현황입니다.

7. `{rel_report}`
   - 엑셀을 열 수 없을 때 보는 요약 Markdown입니다.

8. `40_전달패키지\\01_개인별_직첨부\\<담당자>\\01_결과표_양식`
   - 큐에서 찍힌 기관의 실제 결과표를 확인할 때만 엽니다.

9. `40_전달패키지\\01_개인별_직첨부\\<담당자>\\02_배정기준자료`
   - 증빙 확인이 필요한 행에서만 엽니다.

## 검토 범위

| 구분 | 건수 | 기준 |
|---|---:|---|
| 1차 검토큐 | {first_total} | 본문검토 점수 {BODY_FIRST_PASS_SCORE} 이상 또는 숫자대조 후보 전체 |
| 2차 후보 | {second_total} | 본문검토 점수 {BODY_SECOND_PASS_SCORE} 이상 {BODY_FIRST_PASS_SCORE - 1} 이하 |
| 전체 후보 | {all_total} | 기존 본문/숫자 숏리스트 전체를 기관-항목 단위로 병합 |

| 담당자 | 1차 | 2차 | 전체 |
|---|---:|---:|---:|
{reviewer_rows}

## 검토 규칙

- `확인유형`이 `숫자대조`이면 대조군 값과 비교군 값, 차이/판정, 확인자료를 먼저 봅니다.
- `확인유형`이 `본문/OCR`이면 증빙 예시 파일과 키워드/누락/오류 징후를 봅니다.
- `확인유형`이 `본문/OCR + 숫자대조`이면 같은 기관/항목에서 값 차이와 본문 징후가 겹친 것이므로 우선 처리합니다.
- 판단이 끝나면 `상태`와 `판단 결과`를 바꾸고, 필요한 경우 `재확인 필요`를 `Y`로 둡니다.

## 일반 검토자가 바로 열지 않아도 되는 파일

- `30_검토산출물\\90_원천_자동검토`: 원천 PDF, 숫자대조/본문검토 산출물, 자동검토 원본입니다.
- `30_검토산출물\\99_보관_로그`: 과거 백업, 기타 작업목록, 실행 로그입니다.
- `30_검토산출물\\80_압축본`: 전달·보관용 ZIP입니다.
- `{core_dir.relative_to(root)}\\99_원천산출물_참고용`: 자동검토 원본입니다.
- `{keyword_reference}`: 상세 키워드 근거가 필요할 때만 엽니다.

## 파일 링크 안내

엑셀의 확인자료 링크는 프로젝트 폴더 구조가 유지된 상태에서 가장 안정적으로 동작합니다.
`30_검토산출물\\80_압축본\\00_검토용_핵심산출물.zip`만 별도 위치에 풀면 일부 원천/증빙 링크가 열리지 않을 수 있습니다.
그 경우 `30_검토산출물\\80_압축본\\99_원천산출물_참고용.zip`을 같은 상위 폴더에 풀거나 공유폴더 원본 위치에서 확인합니다.

## 다시 생성

이미 생성된 자동검토 엑셀을 기준으로 한 장 요약보고서, 검토큐, 시작 안내, 핵심 ZIP을 다시 만듭니다.

```powershell
python .\\90_tools\\run_review_pipeline.py --root . --use-existing-auto
```

공시대조와 증빙매칭 자동검토부터 새로 돌릴 때는 `--use-existing-auto`를 빼고 실행합니다.
"""


def build_core_readme(root: Path, queue_path: Path, generated_at: str) -> str:
    rel_queue = queue_path.relative_to(root)
    rel_report = queue_path.parent.relative_to(root) / "00_한장_요약보고서.md"
    core_dir = queue_path.parent
    rel_report_xlsx = preferred_report_xlsx(core_dir).relative_to(root)
    number_shortlist = rel_reference(core_dir, root, "02_", ".csv")
    number_diff = rel_reference(core_dir, root, "04_", ".csv")
    body_shortlist = rel_reference(core_dir, root, "06_", ".csv")
    body_manual = rel_reference(core_dir, root, "07_", ".csv")
    number_agency = rel_reference(core_dir, root, "03_", ".csv")
    body_agency = rel_reference(core_dir, root, "08_", ".csv")
    return f"""
# 핵심산출물 먼저열기

생성시각: {generated_at}

이 폴더에는 상세 산출물이 함께 들어 있지만, 검토자는 먼저 아래 두 파일만 보면 됩니다.

1. `{rel_report_xlsx}`
   - 여기서 시작합니다. 첫 시트 `검토시작`은 작업 지시 화면입니다.
   - `대표후보_바로보기`: 검토 시작용 대표 후보입니다.
   - `상세후보_값대조`: 같은 형식으로 더 많은 후보를 확인하는 시트입니다.
   - 행마다 `상태`, `판단 결과`, `재확인 필요`, `기관`, `항목`, `검토 질문`, `대조군 값`, `비교군 값`, `차이/판정`, `확인자료`를 왼쪽부터 봅니다.
   - `상태`, `판단 결과`, `재확인 필요`, `검토 메모`만 입력합니다.

2. `{rel_queue}`
   - 관리자용 전체 현황입니다.
   - 담당자는 `01_강_검토큐.xlsx`, `01_정_검토큐.xlsx`, `01_황_검토큐.xlsx`, `01_실_검토큐.xlsx` 중 자기 파일을 먼저 엽니다.
   - `1차_검토큐`: 오늘 바로 볼 기관/항목
   - `2차_후보`: 1차 완료 뒤 추가 확인할 후보
   - `전체후보`: 기존 숏리스트 전체를 병합한 참고용

3. `{rel_report}`
   - 엑셀을 열 수 없을 때 보는 Markdown 요약입니다.

## 상세 파일을 여는 기준

- 숫자 차이 판단이 필요할 때: 먼저 `{rel_report_xlsx}`의 값대조 시트를 보고, 부족할 때만 `{number_shortlist}`, `{number_diff}`를 엽니다.
- 본문/OCR 판단이 필요할 때: 먼저 검토큐의 `증빙 예시`를 보고, 부족할 때만 `{body_shortlist}`, `{body_manual}`를 엽니다.
- 기관별 묶음으로 보고 싶을 때: `{number_agency}`, `{body_agency}`
- 최신 자동검토 원본이 필요할 때: `99_원천산출물_참고용` 폴더

## 파일 링크 안내

엑셀의 확인자료 링크는 프로젝트 폴더 구조가 유지된 상태에서 가장 안정적으로 동작합니다.
핵심 ZIP만 별도 위치에 풀면 일부 원천/증빙 링크가 열리지 않을 수 있으므로, 필요한 경우 `30_검토산출물\\80_압축본\\99_원천산출물_참고용.zip` 또는 공유폴더 원본 위치를 함께 확인합니다.

대용량 상세 근거 파일은 반복 검토를 줄이기 위한 보조 자료입니다. 처음부터 전체 CSV를 열지 말고 검토큐의 `세부 원본 위치` 컬럼에서 필요한 경우에만 들어가면 됩니다.
"""


def build_reviewer_readme(
    reviewer: str,
    root: Path,
    queue_path: Path,
    records: list[dict[str, object]],
    generated_at: str,
) -> str:
    counts = counts_by_reviewer(records)[reviewer]
    rel_queue = queue_path.relative_to(root)
    return f"""
# {reviewer} 담당 검토 시작

생성시각: {generated_at}

먼저 `{rel_queue}` 파일에서 `검토시작_1차` 시트를 엽니다.
상태, 판단 결과, 재확인 필요, 검토 메모만 입력하고 나머지 원천 값은 수정하지 않습니다.

## 이 폴더에서 쓰는 위치

- `01_결과표_양식`: 기관별 결과표 확인
- `02_배정기준자료`: 증빙 원문 확인
- `03_검토숏리스트`: 세부 CSV가 필요할 때만 확인
- `99_목록_누락확인`: 기준자료 누락 여부 확인

## 담당 범위

| 구분 | 건수 |
|---|---:|
| 1차 검토큐 | {counts['1차']} |
| 2차 후보 | {counts['2차']} |
| 전체 후보 | {counts['전체']} |

`상태` 컬럼을 채우는 방식으로 검토 이력을 남기면 같은 기관을 반복해서 다시 훑는 일을 줄일 수 있습니다.

파일 링크는 프로젝트 폴더 구조가 유지된 상태에서 가장 안정적으로 동작합니다.
"""


def move_reference_outputs(core_dir: Path) -> list[Path]:
    reference_dir = core_dir / "99_원천산출물_참고용"
    reference_dir.mkdir(parents=True, exist_ok=True)
    moved: list[Path] = []
    prefixes = (
        "01_숫자대조_",
        "02_숫자대조_",
        "03_숫자대조_",
        "04_숫자대조_",
        "05_본문검토_",
        "06_본문검토_",
        "07_본문검토_",
        "08_본문검토_",
        "09_본문검토_",
        "12_숫자대조_",
    )
    for path in sorted(core_dir.iterdir()):
        if not path.is_file() or not path.name.startswith(prefixes):
            continue
        target = reference_dir / path.name
        try:
            if target.exists():
                target.unlink()
            shutil.move(str(path), str(target))
            moved.append(target)
        except PermissionError:
            print(f"warning: could not move locked reference file: {path.name}")
    stale_zip = core_dir / "00_검토용_핵심산출물.zip"
    if stale_zip.exists():
        try:
            stale_zip.unlink()
        except PermissionError:
            print(f"warning: could not remove stale nested zip: {stale_zip.name}")
    return moved


def write_core_zip(core_dir: Path, zip_path: Path) -> list[Path]:
    skipped: list[Path] = []
    include_paths = [
        core_dir / "00_README_먼저열기.md",
        preferred_report_xlsx(core_dir),
        core_dir / "00_한장_요약보고서.md",
        preferred_xlsx(core_dir, ADMIN_QUEUE_XLSX),
        *[preferred_xlsx(core_dir, f"01_{reviewer}_검토큐.xlsx") for reviewer in REVIEWERS],
    ]
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(dict.fromkeys(include_paths), key=lambda item: item.name):
            if not path.exists() or not path.is_file():
                continue
            try:
                zf.write(path, path.name)
            except PermissionError:
                skipped.append(path)
    return skipped


def write_reference_zip(reference_dir: Path, zip_path: Path) -> list[Path]:
    skipped: list[Path] = []
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        if not reference_dir.exists():
            return skipped
        for path in sorted(reference_dir.rglob("*")):
            if not path.is_file():
                continue
            try:
                zf.write(path, Path(reference_dir.name) / path.relative_to(reference_dir))
            except PermissionError:
                skipped.append(path)
    return skipped


def refresh_zip(core_dir: Path, zip_path: Path) -> Path:
    try:
        skipped = write_core_zip(core_dir, zip_path)
        written = zip_path
    except PermissionError:
        written = zip_path.with_name(
            f"{zip_path.stem}_새로생성_{dt.datetime.now().strftime('%H%M%S')}{zip_path.suffix}"
        )
        skipped = write_core_zip(core_dir, written)
        print(f"warning: {zip_path.name} is open or locked; wrote {written.name} instead")
    for path in skipped:
        print(f"warning: skipped locked file while zipping: {path.relative_to(core_dir)}")
    return written


def refresh_reference_zip(reference_dir: Path, zip_path: Path) -> Path:
    try:
        skipped = write_reference_zip(reference_dir, zip_path)
        written = zip_path
    except PermissionError:
        written = zip_path.with_name(
            f"{zip_path.stem}_새로생성_{dt.datetime.now().strftime('%H%M%S')}{zip_path.suffix}"
        )
        skipped = write_reference_zip(reference_dir, written)
        print(f"warning: {zip_path.name} is open or locked; wrote {written.name} instead")
    for path in skipped:
        print(f"warning: skipped locked reference file while zipping: {path.relative_to(reference_dir)}")
    return written


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--no-zip", action="store_true", help="Do not refresh the core output zip.")
    args = parser.parse_args()

    root = args.root.resolve()
    out_dir = find_child(root, "30_", want_dir=True)
    core_dir = find_core_dir(out_dir)
    archive_dir = out_dir / "80_압축본"
    delivery_dir = find_child(find_child(root, "40_", want_dir=True), "01_", want_dir=True)
    body_csv = find_reference_file(core_dir, "06_", ".csv")
    number_csv = find_reference_file(core_dir, "02_", ".csv")

    records = build_records(read_csv(body_csv), read_csv(number_csv))
    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    queue_path = build_workbook(core_dir / ADMIN_QUEUE_XLSX, records, generated_at)
    reviewer_paths: dict[str, Path] = {}
    for reviewer in REVIEWERS:
        reviewer_path = core_dir / f"01_{reviewer}_검토큐.xlsx"
        reviewer_paths[reviewer] = build_reviewer_workbook(reviewer_path, reviewer, records, generated_at)

    stale_queue = core_dir / "00_담당자별_검토큐.xlsx"
    if stale_queue.exists() and stale_queue != queue_path:
        try:
            stale_queue.unlink()
        except PermissionError:
            pass

    start_md = out_dir / "00_검토_시작.md"
    write_text(start_md, build_start_markdown(root, core_dir, queue_path, records, generated_at))
    write_text(core_dir / "00_README_먼저열기.md", build_core_readme(root, queue_path, generated_at))

    for reviewer in REVIEWERS:
        reviewer_dir = delivery_dir / reviewer
        if reviewer_dir.exists():
            queue_for_readme = reviewer_paths[reviewer]
            personal_queue = reviewer_dir / f"00_{reviewer}_검토큐.xlsx"
            try:
                queue_for_readme = copy_with_fallback(reviewer_paths[reviewer], personal_queue)
            except PermissionError:
                print(f"warning: could not copy locked reviewer queue: {personal_queue.relative_to(root)}")
            write_text(
                reviewer_dir / "00_README_먼저열기.md",
                build_reviewer_readme(reviewer, root, queue_for_readme, records, generated_at),
            )

    moved = move_reference_outputs(core_dir)
    if moved:
        print(f"moved reference files: {len(moved)}")

    if not args.no_zip:
        archive_dir.mkdir(parents=True, exist_ok=True)
        refresh_zip(core_dir, archive_dir / "00_검토용_핵심산출물.zip")
        refresh_reference_zip(core_dir / "99_원천산출물_참고용", archive_dir / "99_원천산출물_참고용.zip")

    print(f"review queue: {queue_path}")
    for path in reviewer_paths.values():
        print(f"reviewer queue: {path}")
    print(f"start guide: {start_md}")
    print(f"records: {len(records)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
